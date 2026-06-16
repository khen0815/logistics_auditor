import io
import math
import re
from itertools import permutations
from pathlib import Path

import pandas as pd
import pdfplumber
import plotly.express as px
import streamlit as st
from aramex_html_processor import clean_selected_table, select_shipment_table
from fpdf import FPDF
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from supabase import create_client


APP_DIR = Path(__file__).resolve().parent
MONTHLY_SNAPSHOT_PATH = APP_DIR / "monthly_audit_snapshots.csv"


st.set_page_config(
    page_title="360° Margin Diagnostic Engine",
    layout="wide",
    initial_sidebar_state="expanded",
)


STANDARD_COLUMNS = [
    "Order_ID",
    "Waybill_ID",
    "SKU",
    "Province",
    "Actual_Weight_KG",
    "Billed_Vol_KG",
    "Billed_Cost_ZAR",
    "Length_cm",
    "Width_cm",
    "Height_cm",
]

NUMERIC_COLUMNS = [
    "Actual_Weight_KG",
    "Billed_Vol_KG",
    "Billed_Cost_ZAR",
    "Length_cm",
    "Width_cm",
    "Height_cm",
]

REQUIRED_AUDIT_COLUMNS = [
    "Actual_Weight_KG",
    "Billed_Cost_ZAR",
]

OPTIONAL_NUMERIC_DEFAULTS = {
    "Billed_Vol_KG": 0,
    "Length_cm": 0,
    "Width_cm": 0,
    "Height_cm": 0,
}

COLUMN_SYNONYMS = {
    "Order_ID": [
        "order id", "order_id", "order no", "order number", "order", "waybill", "waybill no",
        "waybill number", "hawb", "hawb no", "hawb number", "tracking", "tracking number",
        "shipment id", "parcel id", "reference",
    ],
    "SKU": ["sku", "product", "product code", "item", "item code", "description", "product description"],
    "Province": ["province", "destination province", "dest province", "region", "zone", "destination", "ship province"],
    "Actual_Weight_KG": [
        "actual kg", "actual weight", "actual weight kg", "actual_weight_kg", "actualweight", "physical weight",
        "physical_weight", "physicalweight", "wgt", "actual mass", "dead weight", "scale weight", "parcel weight",
        "submitted weight kg", "submitted weight", "submitted weight (kg)", "tot kg", "actualweight",
    ],
    "Billed_Vol_KG": [
        "vol wgt", "vol weight", "volumetric weight", "volumetric weight kg", "volumetric_weight_kg", "billed volumetric",
        "billed volumetric weight", "billed vol", "billed vol kg", "billed_vol_kg", "chargeable weight",
        "billed weight", "billed_weight", "billedweight", "csv volumetric weight", "vol kg", "charged weight kg",
        "charged weight", "charged wt", "charged wt kg", "charged weight kg", "charged weight (kg)",
        "submitted volumetric weight kg", "submitted volumetric weight (kg)", "tot vol", "mass charged",
        "inc mass", "weight", "weight kg",
    ],
    "Billed_Cost_ZAR": [
        "billed amount", "cost zar", "billed cost", "billed cost zar", "amount", "charge", "shipping cost",
        "shipping cost zar", "billed shipping cost", "billed shipping cost zar", "billed_shipping_cost_zar", "total", "total amount",
        "totalamount", "invoice amount", "courier cost", "freight charge", "charged rate incl vat",
        "charged rate (incl vat)", "charged rate incl. vat", "accepted rate incl vat", "accepted rate (incl vat)",
        "courier confirmed rate incl vat", "courier confirmed rate (incl vat)", "exc vat", "exc vat ", "rate", "cost",
        "amount",
    ],
    "Length_cm": ["l", "len", "length", "length cm", "length (cm)", "length_cm", "parcel length"],
    "Width_cm": ["w", "wid", "width", "width cm", "width (cm)", "width_cm", "parcel width"],
    "Height_cm": ["h", "hei", "height", "height cm", "height (cm)", "height_cm", "parcel height"],
    "Dimensions": ["dimensions", "dims", "parcel dimensions", "size", "l x w x h", "lxwxh", "submitted dimensions", "charged dimensions"],
    "Cube": ["cube", "cubic", "cubic volume", "volume", "cbm", "m3"],
    "Revenue": ["revenue", "sales value", "item price", "selling price", "order value"],
}

BOB_GO_DIRECT_MAPPINGS = {
    "submitted weight kg": "Actual_Weight_KG",
    "submitted weight": "Actual_Weight_KG",
    "submitted weight kg": "Actual_Weight_KG",
    "charged weight kg": "Billed_Vol_KG",
    "charged weight": "Billed_Vol_KG",
    "charged rate incl vat": "Billed_Cost_ZAR",
    "charged rate": "Billed_Cost_ZAR",
    "accepted rate incl vat": "Billed_Cost_ZAR",
    "courier confirmed rate incl vat": "Billed_Cost_ZAR",
}

FORCED_UPLOAD_RENAME = {
    "Wb No": "Tracking_Number",
    "HAWB": "Tracking_Number",
    "AWB": "Tracking_Number",
    "Destination Town": "Destination",
    "Destination": "Destination",
    "Mass Charged": "weight",
    "Tot KG": "weight",
    "ActualWeight": "Actual_Weight_KG",
    "Physical_Weight": "Actual_Weight_KG",
    "Weight": "Billed_Vol_KG",
    "Billed_Weight": "Billed_Vol_KG",
    "Exc Vat": "cost",
    "Base Charge": "cost",
    "TotalAmount": "cost",
    "Cost": "cost",
    "Dimensions": "dimensions",
    "Cube": "Cube",
}

FLEXIBLE_COURIER_MAPPINGS = {
    "Tracking_Number": ["wb no", "hawb", "awb", "tracking number", "waybill number", "short waybill number", "alternative tracking references"],
    "Destination": ["destination town", "destination", "delivery branch", "delivery city"],
    "Actual_Weight_KG": ["actual weight", "actualweight", "actual weight kg", "actual_weight_kg", "physical weight", "physical_weight"],
    "Billed_Vol_KG": ["weight", "weight kg", "billed weight", "billed_weight", "chargeable weight", "charged weight", "volumetric_weight_kg", "volumetric weight kg"],
    "weight": ["mass charged", "tot kg"],
    "cost": ["exc vat", "base charge", "total amount", "totalamount", "cost", "rate", "amount", "billed_shipping_cost_zar", "billed shipping cost zar"],
    "dimensions": ["dimensions"],
    "Cube": ["cube", "cubic", "cubic volume", "volume", "cbm", "m3"],
}

BOB_GO_DIMENSION_PATTERN = re.compile(
    r"^\s*(?P<waybill>[^\s\-–—]+).*?"
    r"(?P<length>\d+(?:\.\d+)?)\s*cm\s*x\s*"
    r"(?P<width>\d+(?:\.\d+)?)\s*cm\s*x\s*"
    r"(?P<height>\d+(?:\.\d+)?)\s*cm",
    re.I,
)
DIMENSION_ONLY_PATTERN = re.compile(
    r"(?P<length>\d+(?:\.\d+)?)\s*(?:cm)?\s*x\s*"
    r"(?P<width>\d+(?:\.\d+)?)\s*(?:cm)?\s*x\s*"
    r"(?P<height>\d+(?:\.\d+)?)\s*(?:cm)?",
    re.I,
)

DEFAULT_PACKAGING_MATRIX = pd.DataFrame(
    [
        {"Package": "A4 Flyer", "L": 30.0, "W": 21.0, "H": 2.0, "cost": 1.50, "Fragile/Void Fill Required": False},
        {"Package": "A3 Flyer", "L": 42.0, "W": 30.0, "H": 2.0, "cost": 2.50, "Fragile/Void Fill Required": False},
        {"Package": "Small Box", "L": 25.0, "W": 15.0, "H": 10.0, "cost": 5.00, "Fragile/Void Fill Required": False},
        {"Package": "Medium Box", "L": 35.0, "W": 25.0, "H": 15.0, "cost": 8.00, "Fragile/Void Fill Required": False},
        {"Package": "Large Corrugated Box", "L": 45.0, "W": 35.0, "H": 25.0, "cost": 12.00, "Fragile/Void Fill Required": False},
    ]
)

RATE_CARD_COLUMNS = [
    "Courier",
    "Service_Level",
    "Origin_Zone",
    "Destination_Zone",
    "Min_Weight_KG",
    "Max_Weight_KG",
    "Base_Rate_ZAR",
    "Per_KG_Rate_ZAR",
    "Minimum_Charge_ZAR",
    "Fuel_Surcharge_Pct",
    "VAT_Included",
    "Rounding_Increment_KG",
    "Volumetric_Divisor",
]
DEFAULT_RATE_CARD = pd.DataFrame(columns=RATE_CARD_COLUMNS)
RATE_CARD_ALIAS_GROUPS = {
    "Courier": ["courier", "carrier", "provider"],
    "Service_Level": ["service", "service level", "product", "product group", "service type", "delivery service"],
    "Origin_Zone": ["origin zone", "origin", "from zone", "collection zone", "sender zone"],
    "Destination_Zone": ["destination zone", "destination", "to zone", "zone", "area", "region", "route", "delivery zone"],
    "Min_Weight_KG": ["min kg", "minimum kg", "from kg", "weight from", "kg from", "min weight"],
    "Max_Weight_KG": ["max kg", "maximum kg", "to kg", "weight to", "kg to", "max weight"],
    "Weight_Band": ["weight band", "kg band", "band", "weight", "kg"],
    "Base_Rate_ZAR": ["base", "base rate", "basic", "rate", "charge", "price", "tariff", "amount"],
    "Per_KG_Rate_ZAR": ["per kg", "additional kg", "add kg", "rate per kg", "kg rate", "excess kg"],
    "Minimum_Charge_ZAR": ["minimum", "minimum charge", "min charge", "minimum rate"],
    "Fuel_Surcharge_Pct": ["fuel", "fuel surcharge", "fuel %", "surcharge %", "faf", "fuel levy"],
    "VAT_Included": ["vat", "vat included", "incl vat", "including vat"],
    "Rounding_Increment_KG": ["round", "rounding", "round kg", "increment", "rounding increment"],
    "Volumetric_Divisor": ["divisor", "volumetric divisor", "volume divisor", "vol divisor"],
}

DISTANT_ZONE_PATTERN = re.compile(r"limpopo|mpumalanga|northern cape|eastern cape|free state|rural|remote|outlying", re.I)
FLYER_PATTERN = re.compile(r"flyer|satchel|bag", re.I)
CORRUGATED_PATTERN = re.compile(r"box|corrugated|carton", re.I)

COURIER_EXPLANATIONS = {
    "the courier guy": {
        "title": "The Courier Guy / Bob Go rate-card checks",
        "notes": [
            "Volumetric weight is a rate-card calculation: length × width × height ÷ the client's negotiated divisor.",
            "When an export contains accepted and charged rates, the app treats the charged-minus-accepted amount as the direct dispute-ready recovery.",
            "Weight and dimension differences are kept as operational exposure because service level, zone, fuel and account surcharges can legitimately change the final invoice.",
        ],
    },
    "bob go": {
        "title": "Bob Go aggregated courier checks",
        "notes": [
            "Bob Go-style exports often include submitted weight, submitted volumetric weight, charged weight, accepted rate and charged/confirmed rate.",
            "The safest refund evidence is a direct rate delta on the same waybill, not an estimated rand-per-kg conversion.",
            "Dimension or charged-weight inflation is still useful for packaging and account-manager queries, but it is separated from the dispute pack.",
        ],
    },
    "aramex": {
        "title": "Aramex chargeable-weight checks",
        "notes": [
            "Aramex exports may expose actual, volumetric, chargeable or billed weights depending on the portal format.",
            "The app normalizes these fields for diagnostics, but direct recoveries should be verified against the account's invoice/rate card.",
            "Rows without direct billed-vs-expected money fields are labelled as operational exposure rather than certain credits.",
        ],
    },
    "skynet": {
        "title": "Skynet financial-only checks",
        "notes": [
            "The Skynet path intentionally bypasses volumetric/packaging math and compares identical start town, destination, service and weight groups.",
            "Only charges more than 10% above the common same-lane cost are shown as financial/routing anomalies.",
            "These rows should be queried with the account manager for surcharge, routing or credit explanations.",
        ],
    },
    "custom": {
        "title": "Custom courier assumptions",
        "notes": [
            "Use the client's actual rate card and divisor before treating any weight delta as recoverable.",
            "Direct billed-vs-accepted money fields are dispute-ready; estimated kg exposure is operational unless validated by the courier.",
        ],
    },
}


@st.cache_resource
def get_supabase_client():
    try:
        return create_client(st.secrets["SUPABASE_URL"], st.secrets["SUPABASE_KEY"])
    except Exception:
        return None


def normalize_label(value):
    return re.sub(r"[^a-z0-9]+", " ", str(value).strip().lower()).strip()


def build_alias_lookup():
    lookup = {}
    for standard_name, aliases in COLUMN_SYNONYMS.items():
        lookup[normalize_label(standard_name)] = standard_name
        for alias in aliases:
            lookup[normalize_label(alias)] = standard_name
    return lookup


ALIAS_LOOKUP = build_alias_lookup()
FLEXIBLE_COURIER_LOOKUP = {
    normalize_label(alias): standard_name
    for standard_name, aliases in FLEXIBLE_COURIER_MAPPINGS.items()
    for alias in aliases
}


def parse_numeric(value):
    if pd.isna(value):
        return pd.NA
    cleaned = re.sub(r"[^0-9.\-]", "", str(value))
    return pd.to_numeric(cleaned, errors="coerce")


def parse_dimensions(value):
    if pd.isna(value):
        return None
    numbers = re.findall(r"\d+(?:\.\d+)?", str(value))
    if len(numbers) < 3:
        return None
    return [float(numbers[0]), float(numbers[1]), float(numbers[2])]


def parse_bob_go_dimensions(value):
    if pd.isna(value):
        return {"Waybill_ID": pd.NA, "Length_cm": pd.NA, "Width_cm": pd.NA, "Height_cm": pd.NA}
    text = str(value)
    match = BOB_GO_DIMENSION_PATTERN.search(text)
    waybill = pd.NA
    if not match:
        match = DIMENSION_ONLY_PATTERN.search(text)
    else:
        waybill = match.group("waybill")
    if not match:
        return {"Waybill_ID": pd.NA, "Length_cm": pd.NA, "Width_cm": pd.NA, "Height_cm": pd.NA}
    return {
        "Waybill_ID": waybill,
        "Length_cm": float(match.group("length")),
        "Width_cm": float(match.group("width")),
        "Height_cm": float(match.group("height")),
    }


def safe_float(value, default=0.0):
    value = pd.to_numeric(value, errors="coerce")
    return float(value) if pd.notna(value) else default


def safe_text(value, default="Unknown"):
    if pd.isna(value) or str(value).strip() == "":
        return default
    return str(value).strip()


def format_currency(value, currency_prefix):
    return f"{currency_prefix}{safe_float(value):,.2f}"


def format_currency_whole(value, currency_prefix):
    return f"{currency_prefix}{safe_float(value):,.0f}"


def display_currency_columns(df, currency_symbol):
    return df.rename(
        columns={
            "Billed_Cost_ZAR": f"Billed_Cost_{currency_symbol}",
            "Estimated_Loss_ZAR": f"Estimated_Loss_{currency_symbol}",
            "Recoverable_Overcharge_ZAR": f"Recoverable_Overcharge_{currency_symbol}",
            "Avoidable_Volumetric_Leak_ZAR": f"Avoidable_Volumetric_Leak_{currency_symbol}",
            "Financial_Excess_ZAR": f"Financial_Excess_{currency_symbol}",
            "Courier_Recoverable_ZAR": f"Courier_Recoverable_{currency_symbol}",
            "Operational_Exposure_ZAR": f"Operational_Exposure_{currency_symbol}",
            "Rate_Delta_ZAR": f"Rate_Delta_{currency_symbol}",
            "Expected_Rate_ZAR": f"Expected_Rate_{currency_symbol}",
            "Rate_Card_Delta_ZAR": f"Rate_Card_Delta_{currency_symbol}",
            "Accepted_Rate_ZAR": f"Accepted_Rate_{currency_symbol}",
            "Charged_Rate_ZAR": f"Charged_Rate_{currency_symbol}",
            "Courier_Confirmed_Rate_ZAR": f"Courier_Confirmed_Rate_{currency_symbol}",
            "SKU_Total_Shipping_Cost_ZAR": f"SKU_Total_Shipping_Cost_{currency_symbol}",
            "SKU_Total_Leakage_ZAR": f"SKU_Total_Leakage_{currency_symbol}",
            "Multi_Item_Leak_ZAR": f"Multi_Item_Leak_{currency_symbol}",
        }
    )


def convert_df_to_styled_excel(df):
    """Return a professionally styled in-memory Excel export for Streamlit downloads."""
    output = io.BytesIO()
    export_df = df.copy() if df is not None else pd.DataFrame()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        sheet_name = "Diagnostic Export"
        export_df.to_excel(writer, index=False, sheet_name=sheet_name)
        worksheet = writer.sheets[sheet_name]

        header_fill = PatternFill("solid", fgColor="1F4E79")
        header_font = Font(name="Calibri", bold=True, color="FFFFFF")
        body_font = Font(name="Calibri", color="111827")
        border_side = Side(style="thin", color="D9E2EC")
        cell_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
        center_columns = {
            "Tracking_Number",
            "Order_ID",
            "Waybill_ID",
            "Service_Level",
            "Srv",
            "Start_Town",
        }
        currency_columns = {
            "cost",
            "Billed_Cost_ZAR",
            "Financial_Excess_ZAR",
            "Estimated_Loss_ZAR",
            "Recoverable_Overcharge_ZAR",
            "Courier_Recoverable_ZAR",
            "Operational_Exposure_ZAR",
            "Rate_Delta_ZAR",
            "Expected_Rate_ZAR",
            "Rate_Card_Delta_ZAR",
            "Accepted_Rate_ZAR",
            "Charged_Rate_ZAR",
            "Courier_Confirmed_Rate_ZAR",
            "Avoidable_Volumetric_Leak_ZAR",
            "SKU_Total_Shipping_Cost_ZAR",
            "SKU_Total_Leakage_ZAR",
            "Multi_Item_Leak_ZAR",
        }
        weight_columns = {"weight", "Actual_Weight_KG", "Billed_Vol_KG", "Financial_Weight_KG", "Billed_Weight_KG", "Excess_Weight_KG"}

        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = cell_border

        header_lookup = {cell.column: str(cell.value) for cell in worksheet[1] if cell.value is not None}
        for row in worksheet.iter_rows(min_row=2):
            for cell in row:
                column_name = header_lookup.get(cell.column, "")
                cell.font = body_font
                cell.border = cell_border
                cell.alignment = Alignment(
                    horizontal="center" if column_name in center_columns else "left",
                    vertical="center",
                )
                if column_name in currency_columns:
                    cell.number_format = '"R" #,##0.00'
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                elif column_name in weight_columns:
                    cell.number_format = '0.00" kg"'
                    cell.alignment = Alignment(horizontal="right", vertical="center")

        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions

        for column_cells in worksheet.columns:
            column_letter = get_column_letter(column_cells[0].column)
            max_length = 0
            for cell in column_cells:
                value = "" if cell.value is None else str(cell.value)
                max_length = max(max_length, len(value))
            worksheet.column_dimensions[column_letter].width = min(max(max_length + 3, 12), 45)

    output.seek(0)
    return output.getvalue()


def parse_numeric_or_zero(value):
    """Convert messy weights/currency to float while tolerating blanks, NaN, and pd.NA."""
    if pd.isna(value):
        return 0.0
    cleaned = re.sub(r"[^0-9.\-]", "", str(value).replace("\n", " ").strip())
    if cleaned in {"", "-", "."}:
        return 0.0
    numeric_value = pd.to_numeric(cleaned, errors="coerce")
    return float(numeric_value) if pd.notna(numeric_value) else 0.0


def clean_numeric_series(series):
    """Vector-safe numeric cleanup for Excel/PDF exports with page breaks and blank cells."""
    cleaned = (
        series
        .fillna("")
        .astype(str)
        .str.replace("\n", " ", regex=False)
        .str.replace(r"[^0-9.\-]", "", regex=True)
    )
    return pd.to_numeric(cleaned, errors="coerce").fillna(0.0).astype(float)


def clean_money_series(series):
    return clean_numeric_series(series)


def coalesce_duplicate_columns(df):
    """Collapse duplicate column labels created by aggressive courier header mapping."""
    if not df.columns.duplicated().any():
        return df

    coalesced = pd.DataFrame(index=df.index)
    for column in dict.fromkeys(df.columns):
        matching = df.loc[:, df.columns == column]
        if matching.shape[1] == 1:
            coalesced[column] = matching.iloc[:, 0]
        else:
            coalesced[column] = matching.bfill(axis=1).iloc[:, 0]
    return coalesced


def apply_bouncer_bypass(df):
    """Inject legacy strict-cleanup columns immediately after courier header renaming."""
    bypassed = df.copy()
    bypassed.columns = bypassed.columns.astype(str).str.strip()

    if "weight" not in bypassed.columns:
        bypassed["weight"] = bypassed["Actual_Weight_KG"] if "Actual_Weight_KG" in bypassed.columns else 0.0
    if "cost" not in bypassed.columns:
        bypassed["cost"] = bypassed["Billed_Cost_ZAR"] if "Billed_Cost_ZAR" in bypassed.columns else 0.0
    if "billed volumetric weight" not in bypassed.columns:
        bypassed["billed volumetric weight"] = bypassed["Billed_Vol_KG"] if "Billed_Vol_KG" in bypassed.columns else 0.0
    if "dimensions" not in bypassed.columns:
        bypassed["dimensions"] = bypassed["Dimensions"] if "Dimensions" in bypassed.columns else 0.0

    bypassed["weight"] = pd.to_numeric(bypassed["weight"], errors="coerce").fillna(0.0)
    bypassed["cost"] = pd.to_numeric(bypassed["cost"], errors="coerce").fillna(0.0)
    bypassed["billed volumetric weight"] = pd.to_numeric(bypassed["billed volumetric weight"], errors="coerce").fillna(0.0)
    bypassed["dimensions"] = bypassed["dimensions"].fillna("")
    return bypassed


def normalize_courier_export_columns(df):
    """Force courier-specific upload headers into the app's standard working variables."""
    cleaned = df.copy().dropna(how="all")
    cleaned.columns = cleaned.columns.astype(str).str.strip()
    cleaned = cleaned.dropna(axis=1, how="all")

    exact_rename = {column: FORCED_UPLOAD_RENAME[column] for column in cleaned.columns if column in FORCED_UPLOAD_RENAME}
    cleaned = cleaned.rename(columns=exact_rename)
    cleaned = coalesce_duplicate_columns(cleaned)

    flexible_rename = {}
    used_targets = set(exact_rename.values())
    for column in cleaned.columns:
        if column in used_targets:
            continue
        normalized = normalize_label(column)
        target = FLEXIBLE_COURIER_LOOKUP.get(normalized)
        if target and target not in used_targets:
            flexible_rename[column] = target
            used_targets.add(target)
    cleaned = cleaned.rename(columns=flexible_rename)
    cleaned = coalesce_duplicate_columns(cleaned)

    cleaned = apply_bouncer_bypass(cleaned)

    if "Tracking_Number" in cleaned.columns:
        cleaned["HAWB"] = cleaned.get("HAWB", cleaned["Tracking_Number"])
        cleaned["Order_ID"] = cleaned.get("Order_ID", cleaned["Tracking_Number"])
        cleaned["Waybill_ID"] = cleaned.get("Waybill_ID", cleaned["Tracking_Number"])
    if "Destination" in cleaned.columns:
        cleaned["Province"] = cleaned.get("Province", cleaned["Destination"])
    if "weight" in cleaned.columns:
        cleaned["Actual_Weight_KG"] = cleaned.get("Actual_Weight_KG", cleaned["weight"])
    if "cost" in cleaned.columns:
        cleaned["Billed_Cost_ZAR"] = cleaned.get("Billed_Cost_ZAR", cleaned["cost"])
    if "dimensions" in cleaned.columns:
        cleaned["Dimensions"] = cleaned.get("Dimensions", cleaned["dimensions"])

    return cleaned


def process_skynet_file(uploaded_file):
    """Read Skynet Excel exports into the exact schema used by financial anomaly detection."""
    preview = pd.read_excel(uploaded_file, header=None, nrows=15)
    header_index = None
    for index, row in preview.iterrows():
        row_text = " ".join(str(value).replace(" ", " ").strip() for value in row.dropna().tolist())
        if re.search(r"\b(Wb\s*No|Tot\s*KG)\b", row_text, re.I):
            header_index = index
            break

    if header_index is None:
        raise ValueError("Could not find Skynet header row containing Wb No or Tot KG.")

    uploaded_file.seek(0)
    df = pd.read_excel(uploaded_file, header=header_index)
    df.columns = df.columns.astype(str).str.strip()
    df = df.dropna(how="all")

    required_columns = {
        "Wb No": "Tracking_Number",
        "Start Town": "Start_Town",
        "Destination Town": "Destination",
        "Srv": "Service_Level",
        "Mass Charged": "weight",
        "Exc Vat": "cost",
    }
    missing = [column for column in required_columns if column not in df.columns]
    if missing:
        raise ValueError(f"Skynet file is missing required columns: {', '.join(missing)}")

    df = df[list(required_columns)].rename(columns=required_columns).copy()
    df["weight"] = pd.to_numeric(df["weight"], errors="coerce")
    df["cost"] = pd.to_numeric(df["cost"], errors="coerce")
    df = df[df["weight"].notna() & (df["weight"] > 0) & df["cost"].notna() & (df["cost"] > 0)].copy()

    df.attrs["pipeline"] = "skynet_financial_only"
    return df.reset_index(drop=True)


def read_excel_with_dynamic_header(file_buffer, scan_rows=15):
    """Find the real courier export header row below invoice/letterhead rows."""
    preview = pd.read_excel(file_buffer, header=None, nrows=scan_rows)
    header_index = None
    header_markers = re.compile(
        r"\b(AWB|HAWB|Weight|ActualWeight|Actual_Weight_kg|Volumetric_Weight_kg|Billed_Shipping_Cost_ZAR|TotalAmount|Wb\s*No|Tot\s*KG|Mass\s*Charged|Exc\s*Vat|Charged\s*weight|Rate)\b",
        re.I,
    )
    for index, row in preview.iterrows():
        row_text = " ".join(str(value) for value in row.dropna().tolist())
        if header_markers.search(row_text):
            header_index = index
            break

    file_buffer.seek(0)
    if header_index is None:
        st.warning("Could not auto-detect an Excel header row containing courier fields; falling back to the second row as headers.")
        header_index = 1
    else:
        st.caption(f"Excel header detected at row {header_index + 1}; skipped {header_index} letterhead/metadata row(s).")

    data = pd.read_excel(file_buffer, header=header_index)
    data.columns = data.columns.astype(str).str.strip()
    data = data.dropna(how="all").dropna(axis=1, how="all")
    return normalize_courier_export_columns(data)


def load_and_clean_data(df):
    """Robust courier ingestion that keeps valid rows even when dimension columns are absent."""
    return clean_and_standardize_data(df)


def first_available_column(df, candidates):
    for candidate in candidates:
        if candidate in df.columns:
            return df[candidate]
    return None


def get_courier_explanation(provider_name):
    normalized_provider = normalize_label(provider_name)
    for key, explanation in COURIER_EXPLANATIONS.items():
        if normalize_label(key) in normalized_provider:
            return explanation
    return COURIER_EXPLANATIONS["custom"]


def first_present_column(df, candidates):
    for candidate in candidates:
        if candidate in df.columns:
            return candidate
    return None


def calculate_rate_card_deltas(df):
    """Calculate direct billed-vs-accepted money deltas when the export provides them."""
    result = pd.DataFrame(index=df.index)
    accepted_col = first_present_column(df, ["Accepted rate (incl. VAT)", "Accepted rate incl VAT", "Accepted_Rate_ZAR"])
    charged_col = first_present_column(df, ["Courier confirmed rate (incl. VAT)", "Courier confirmed rate incl VAT", "Charged rate (incl. VAT)", "Charged rate incl VAT", "Billed_Cost_ZAR"])

    result["Accepted_Rate_ZAR"] = clean_money_series(df[accepted_col]) if accepted_col else 0.0
    result["Charged_Rate_ZAR"] = clean_money_series(df[charged_col]) if charged_col else 0.0
    result["Courier_Confirmed_Rate_ZAR"] = result["Charged_Rate_ZAR"]
    result["Rate_Delta_ZAR"] = 0.0
    result["Rate_Delta_Flag"] = False
    result["Rate_Delta_Reason"] = "No accepted-vs-charged rate fields were available for direct recovery calculation."

    if accepted_col and charged_col:
        result["Rate_Delta_ZAR"] = (result["Charged_Rate_ZAR"] - result["Accepted_Rate_ZAR"]).clip(lower=0).round(2)
        result["Rate_Delta_Flag"] = result["Rate_Delta_ZAR"] > 0
        result.loc[result["Rate_Delta_Flag"], "Rate_Delta_Reason"] = "Charged/courier-confirmed rate exceeds the accepted quote/rate for this waybill."
        result.loc[~result["Rate_Delta_Flag"], "Rate_Delta_Reason"] = "Charged/courier-confirmed rate does not exceed the accepted quote/rate."

    return result


def rate_card_template_frame(courier_provider="The Courier Guy", fallback_divisor=5000):
    return pd.DataFrame([
        {
            "Courier": courier_provider,
            "Service_Level": "ECO",
            "Origin_Zone": "All",
            "Destination_Zone": "All",
            "Min_Weight_KG": 0,
            "Max_Weight_KG": 2,
            "Base_Rate_ZAR": 65,
            "Per_KG_Rate_ZAR": 0,
            "Minimum_Charge_ZAR": 65,
            "Fuel_Surcharge_Pct": 0,
            "VAT_Included": True,
            "Rounding_Increment_KG": 1,
            "Volumetric_Divisor": fallback_divisor,
        },
        {
            "Courier": courier_provider,
            "Service_Level": "ECO",
            "Origin_Zone": "All",
            "Destination_Zone": "All",
            "Min_Weight_KG": 2.01,
            "Max_Weight_KG": 5,
            "Base_Rate_ZAR": 65,
            "Per_KG_Rate_ZAR": 12,
            "Minimum_Charge_ZAR": 0,
            "Fuel_Surcharge_Pct": 0,
            "VAT_Included": True,
            "Rounding_Increment_KG": 1,
            "Volumetric_Divisor": fallback_divisor,
        },
    ], columns=RATE_CARD_COLUMNS)


def generate_rate_card_template_xlsx(courier_provider, fallback_divisor):
    return convert_df_to_styled_excel(rate_card_template_frame(courier_provider, fallback_divisor))


def parse_weight_band(value):
    text = str(value or "").lower().replace(",", ".")
    numbers = [float(number) for number in re.findall(r"\d+(?:\.\d+)?", text)]
    if not numbers:
        return pd.NA, pd.NA
    if any(marker in text for marker in ["+", "plus", "over", "above"]):
        return numbers[0], 999999.0
    if len(numbers) >= 2:
        return numbers[0], numbers[1]
    return 0.0, numbers[0]


def infer_rate_card_zone_from_filename(name):
    label = normalize_label(name)
    if "local" in label:
        return "Local"
    if "regional" in label:
        return "Regional"
    if "main" in label:
        return "Main"
    return "All"


def strip_duplicate_column_suffix(label):
    return re.sub(r"\.\d+$", "", str(label or "")).strip()


def clean_matrix_service_label(label):
    text = re.sub(r"\([^)]*\)", "", str(label or "")).strip()
    return re.sub(r"\s+", " ", text) or "All"


def looks_like_weight_rate_matrix(table):
    if table is None or table.empty or len(table) < 2 or len(table.columns) < 2:
        return False
    first_header = normalize_label(table.columns[0])
    first_value = normalize_label(table.iloc[0, 0])
    return "courier" in first_header and "weight" in first_value


def normalize_rate_matrix(table, source_name, fallback_divisor):
    zone = infer_rate_card_zone_from_filename(source_name)
    rows = []
    matrix = table.copy() if table is not None else pd.DataFrame()
    if not looks_like_weight_rate_matrix(matrix):
        return pd.DataFrame(columns=RATE_CARD_COLUMNS)

    service_row = matrix.iloc[0]
    rate_rows = matrix.iloc[1:].copy()
    weight_column = matrix.columns[0]
    weights = clean_numeric_series(rate_rows[weight_column])
    previous_weight = 0.0

    for row_position, (_, rate_row) in enumerate(rate_rows.iterrows()):
        max_weight = safe_float(weights.iloc[row_position])
        if max_weight <= 0:
            continue
        min_weight = 0.0 if previous_weight == 0 else previous_weight + 0.0001
        for column in matrix.columns[1:]:
            price = parse_numeric_or_zero(rate_row[column])
            if price <= 0:
                continue
            rows.append({
                "Courier": strip_duplicate_column_suffix(column),
                "Service_Level": clean_matrix_service_label(service_row[column]),
                "Origin_Zone": "All",
                "Destination_Zone": zone,
                "Min_Weight_KG": round(min_weight, 4),
                "Max_Weight_KG": max_weight,
                "Base_Rate_ZAR": price,
                "Per_KG_Rate_ZAR": 0.0,
                "Minimum_Charge_ZAR": 0.0,
                "Fuel_Surcharge_Pct": 0.0,
                "VAT_Included": True,
                "Rounding_Increment_KG": 1.0,
                "Volumetric_Divisor": fallback_divisor,
            })
        previous_weight = max(previous_weight, max_weight)

    return pd.DataFrame(rows, columns=RATE_CARD_COLUMNS)


def normalize_rate_card_candidate(candidate, courier_provider, fallback_divisor, source_name):
    if candidate.get("kind") == "matrix":
        normalized = normalize_rate_matrix(candidate["table"], source_name, fallback_divisor)
        return normalized, {"Matrix": "Courier/service/weight grid"}
    return normalize_rate_card_columns(candidate["table"], courier_provider, fallback_divisor)


def rate_card_alias_lookup():
    lookup = {}
    for target, aliases in RATE_CARD_ALIAS_GROUPS.items():
        lookup[normalize_label(target)] = target
        for alias in aliases:
            lookup[normalize_label(alias)] = target
    return lookup


def score_rate_card_table(table):
    if table is None or table.empty:
        return 0, []
    header_text = " ".join(normalize_label(column) for column in table.columns)
    sample_text = " ".join(table.head(5).astype(str).fillna("").agg(" ".join, axis=1).tolist()).lower()
    combined = f"{header_text} {sample_text}"
    signals = {
        "service": 12,
        "zone": 10,
        "weight": 12,
        "kg": 12,
        "rate": 14,
        "charge": 10,
        "base": 8,
        "fuel": 6,
        "vat": 5,
        "divisor": 5,
    }
    matched = [signal for signal in signals if signal in combined]
    score = sum(signals[signal] for signal in matched) + min(len(table), 20)
    return score, matched


def build_rate_card_candidate(label, table):
    score, matched = score_rate_card_table(table)
    kind = "matrix" if looks_like_weight_rate_matrix(table) else "flat"
    if kind == "matrix":
        matched = sorted(set(matched + ["matrix", "courier", "service", "weight"]))
        score += 50
    return {"label": label, "table": table, "score": score, "matched": matched, "kind": kind}


def read_rate_card_upload(uploaded_file):
    if uploaded_file is None:
        return [], "No rate card uploaded."
    name = uploaded_file.name
    lower_name = name.lower()
    raw_bytes = uploaded_file.getvalue()
    candidates = []
    try:
        if lower_name.endswith(".csv"):
            table = pd.read_csv(io.BytesIO(raw_bytes))
            candidates.append(build_rate_card_candidate("CSV table", table))
        elif lower_name.endswith((".xlsx", ".xls")):
            workbook = pd.ExcelFile(io.BytesIO(raw_bytes))
            for sheet_name in workbook.sheet_names:
                table = pd.read_excel(workbook, sheet_name=sheet_name)
                candidates.append(build_rate_card_candidate(f"Sheet: {sheet_name}", table))
        elif lower_name.endswith((".html", ".htm")):
            for index, table in enumerate(pd.read_html(io.BytesIO(raw_bytes))):
                candidates.append(build_rate_card_candidate(f"HTML table {index + 1}", table))
        elif lower_name.endswith(".pdf"):
            with pdfplumber.open(io.BytesIO(raw_bytes)) as pdf:
                for page_index, page in enumerate(pdf.pages):
                    for table_index, table_rows in enumerate(page.extract_tables() or []):
                        if table_rows and len(table_rows) > 1:
                            table = pd.DataFrame(table_rows[1:], columns=[str(cell or "").strip() for cell in table_rows[0]])
                            candidates.append(build_rate_card_candidate(f"PDF page {page_index + 1} table {table_index + 1}", table))
        else:
            return [], "Unsupported rate-card file type."
    except Exception as exc:
        return [], f"Rate card upload could not be read: {exc}"
    candidates = sorted(candidates, key=lambda candidate: candidate["score"], reverse=True)
    return candidates, f"Detected {len(candidates):,} candidate table(s) in {name}."


def normalize_rate_card_columns(raw_table, courier_provider, fallback_divisor):
    table = raw_table.copy() if raw_table is not None else pd.DataFrame()
    normalized = pd.DataFrame(index=table.index)
    lookup = rate_card_alias_lookup()
    mapped_columns = {}
    for column in table.columns:
        target = lookup.get(normalize_label(column))
        if target and target not in mapped_columns:
            mapped_columns[target] = column

    for target, source in mapped_columns.items():
        if target != "Weight_Band":
            normalized[target] = table[source]

    if "Weight_Band" in mapped_columns:
        bands = table[mapped_columns["Weight_Band"]].apply(parse_weight_band)
        band_frame = pd.DataFrame(bands.tolist(), index=table.index, columns=["Min_Weight_KG", "Max_Weight_KG"])
        normalized["Min_Weight_KG"] = normalized.get("Min_Weight_KG", band_frame["Min_Weight_KG"])
        normalized["Max_Weight_KG"] = normalized.get("Max_Weight_KG", band_frame["Max_Weight_KG"])

    for column in RATE_CARD_COLUMNS:
        if column not in normalized.columns:
            normalized[column] = pd.NA

    normalized["Courier"] = normalized["Courier"].fillna(courier_provider).replace("", courier_provider)
    normalized["Service_Level"] = normalized["Service_Level"].fillna("All").replace("", "All")
    normalized["Origin_Zone"] = normalized["Origin_Zone"].fillna("All").replace("", "All")
    normalized["Destination_Zone"] = normalized["Destination_Zone"].fillna("All").replace("", "All")
    normalized["VAT_Included"] = normalized["VAT_Included"].fillna(True)
    normalized["Rounding_Increment_KG"] = normalized["Rounding_Increment_KG"].fillna(1)
    normalized["Volumetric_Divisor"] = normalized["Volumetric_Divisor"].fillna(fallback_divisor)

    for column in ["Min_Weight_KG", "Max_Weight_KG", "Base_Rate_ZAR", "Per_KG_Rate_ZAR", "Minimum_Charge_ZAR", "Fuel_Surcharge_Pct", "Rounding_Increment_KG", "Volumetric_Divisor"]:
        normalized[column] = clean_numeric_series(normalized[column])

    normalized["VAT_Included"] = normalized["VAT_Included"].astype(str).str.lower().isin(["true", "yes", "y", "1", "incl", "included"])
    return normalized[RATE_CARD_COLUMNS], mapped_columns


def validate_rate_card_readiness(cleaned_card):
    warnings = []
    card = cleaned_card if cleaned_card is not None else pd.DataFrame()
    if card.empty:
        return ["No usable rate-card rows yet. Upload a rate card, paste rows, or use the template."]
    if not ((card["Max_Weight_KG"] >= card["Min_Weight_KG"]) & (card["Max_Weight_KG"] > 0)).all():
        warnings.append("Some weight bands are missing or invalid.")
    if not ((card["Base_Rate_ZAR"] > 0) | (card["Per_KG_Rate_ZAR"] > 0) | (card["Minimum_Charge_ZAR"] > 0)).any():
        warnings.append("No positive base, per-kg, or minimum charges were detected.")
    if card["Service_Level"].fillna("").astype(str).str.strip().eq("").all():
        warnings.append("Service levels are blank; matching will be broad and lower-confidence.")
    if card["Destination_Zone"].fillna("").astype(str).str.lower().isin(["", "all"]).all():
        warnings.append("No destination zones were detected; zone-specific courier pricing may be missed.")
    if card["Volumetric_Divisor"].fillna(0).le(0).any():
        warnings.append("Some rows are missing volumetric divisors.")
    if card["Fuel_Surcharge_Pct"].fillna(0).eq(0).all():
        warnings.append("Fuel surcharge is zero/blank; confirm whether rates already include fuel.")
    return warnings or ["Rate card looks usable for Tier B expected-rate checks. Confirm effective dates before disputing."]


def clean_rate_card(rate_card_df, fallback_divisor):
    """Normalize user-supplied client rate cards without creating claims from blank/template data."""
    card = rate_card_df.copy() if rate_card_df is not None else pd.DataFrame()
    for column in RATE_CARD_COLUMNS:
        if column not in card.columns:
            card[column] = pd.NA
    card = card[RATE_CARD_COLUMNS].dropna(how="all").copy()
    if card.empty:
        return card

    for column in ["Courier", "Service_Level", "Origin_Zone", "Destination_Zone"]:
        card[column] = card[column].fillna("").astype(str).str.strip()
    for column in ["Min_Weight_KG", "Max_Weight_KG", "Base_Rate_ZAR", "Per_KG_Rate_ZAR", "Minimum_Charge_ZAR", "Fuel_Surcharge_Pct", "Rounding_Increment_KG", "Volumetric_Divisor"]:
        card[column] = pd.to_numeric(card[column], errors="coerce")
    card["VAT_Included"] = card["VAT_Included"].fillna(True).astype(bool)
    card["Min_Weight_KG"] = card["Min_Weight_KG"].fillna(0)
    card["Max_Weight_KG"] = card["Max_Weight_KG"].fillna(999999)
    card["Base_Rate_ZAR"] = card["Base_Rate_ZAR"].fillna(0)
    card["Per_KG_Rate_ZAR"] = card["Per_KG_Rate_ZAR"].fillna(0)
    card["Minimum_Charge_ZAR"] = card["Minimum_Charge_ZAR"].fillna(0)
    card["Fuel_Surcharge_Pct"] = card["Fuel_Surcharge_Pct"].fillna(0)
    card["Rounding_Increment_KG"] = card["Rounding_Increment_KG"].fillna(1).clip(lower=0.01)
    card["Volumetric_Divisor"] = card["Volumetric_Divisor"].fillna(fallback_divisor)
    card = card[(card["Max_Weight_KG"] >= card["Min_Weight_KG"]) & ((card["Base_Rate_ZAR"] > 0) | (card["Per_KG_Rate_ZAR"] > 0) | (card["Minimum_Charge_ZAR"] > 0))]
    return card.reset_index(drop=True)


def normalized_match(value, target):
    target_norm = normalize_label(target)
    if target_norm in {"", "all", "any", "*"}:
        return True
    value_norm = normalize_label(value)
    return value_norm == target_norm or (target_norm and target_norm in value_norm)


def row_expected_rate_for_sort(row):
    return calculate_expected_rate(pd.Series({"Actual_Weight_KG": row.get("Max_Weight_KG", 0), "Billed_Vol_KG": row.get("Max_Weight_KG", 0)}), row)["Expected_Rate_ZAR"]


def extract_route_bucket(value):
    normalized = normalize_label(value)
    for bucket in ["local", "regional", "main"]:
        if bucket in normalized:
            return bucket
    return ""


def match_rate_card_row(shipment_row, rate_card_df, courier_provider):
    if rate_card_df is None or rate_card_df.empty:
        return None, "No client rate card rows loaded.", 0
    weight = get_rate_card_chargeable_weight(shipment_row)
    service = safe_text(shipment_row.get("Service_Level", shipment_row.get("Service level", shipment_row.get("Srv", ""))), "")
    destination = safe_text(shipment_row.get("Zone", shipment_row.get("Province", shipment_row.get("Destination", shipment_row.get("Delivery branch", "")))), "")
    route_bucket = extract_route_bucket(" ".join(str(shipment_row.get(column, "")) for column in ["Zone", "Province", "Destination", "Delivery branch", "Destination_Zone", "Route"]))

    candidates = rate_card_df[(rate_card_df["Min_Weight_KG"] <= weight) & (rate_card_df["Max_Weight_KG"] >= weight)].copy()
    if candidates.empty:
        return None, f"No rate-card weight band matched {weight:.2f} kg.", 0

    candidates = candidates[candidates["Service_Level"].apply(lambda value: normalized_match(service, value))]
    if candidates.empty:
        return None, f"No rate-card service level matched '{service or 'blank'}'.", 0

    confidence = 78
    courier_candidates = candidates[candidates["Courier"].apply(lambda value: normalized_match(courier_provider, value))]
    if not courier_candidates.empty:
        candidates = courier_candidates
        confidence += 5

    zone_candidates = candidates[candidates["Destination_Zone"].apply(lambda value: normalized_match(destination, value))]
    if not zone_candidates.empty:
        candidates = zone_candidates
        confidence = max(confidence, 88)

    if route_bucket:
        bucket_candidates = candidates[candidates["Destination_Zone"].apply(lambda value: extract_route_bucket(value) == route_bucket)]
        if not bucket_candidates.empty:
            candidates = bucket_candidates
            confidence = max(confidence, 92)

    service_zone_pairs = candidates[["Service_Level", "Destination_Zone"]].drop_duplicates() if not candidates.empty else pd.DataFrame()
    ambiguous_route = False
    if len(service_zone_pairs) > 1 and not route_bucket:
        same_service_count = candidates["Service_Level"].apply(normalize_label).nunique()
        if same_service_count == 1:
            ambiguous_route = True
            candidates = candidates.assign(_Expected_For_Sort=candidates.apply(row_expected_rate_for_sort, axis=1)).sort_values(["_Expected_For_Sort", "Max_Weight_KG", "Min_Weight_KG"])
            confidence = min(confidence, 82)
        else:
            candidates = candidates.sort_values(["Max_Weight_KG", "Min_Weight_KG"])
    else:
        candidates = candidates.sort_values(["Max_Weight_KG", "Min_Weight_KG"])

    match = candidates.iloc[0]
    band = f"{safe_float(match['Min_Weight_KG']):.2f}-{safe_float(match['Max_Weight_KG']):.2f} kg"
    zone_text = safe_text(match.get("Destination_Zone", "Any"), "Any")
    reason = f"Matched service '{service or 'any'}', destination '{destination or 'any'}', route bucket '{zone_text}', weight band {band}."
    if ambiguous_route:
        reason += " Multiple HEL route buckets matched; no shipment route bucket was available, so the conservative lowest eligible rate was used. Confirm local/main/regional before dispute."
    return match, reason, min(confidence, 95)


def get_rate_card_chargeable_weight(row):
    """Return the weight used for rate-card reconstruction without inventing a new courier weight."""
    return max(safe_float(row.get("Billed_Vol_KG", row.get("weight", 0))), safe_float(row.get("Actual_Weight_KG", 0)))


def calculate_expected_rate(row, rate_card_row):
    chargeable_weight = get_rate_card_chargeable_weight(row)
    increment = max(safe_float(rate_card_row.get("Rounding_Increment_KG", 1)), 0.01)
    rounded_weight = math.ceil(chargeable_weight / increment) * increment
    base_component = safe_float(rate_card_row.get("Base_Rate_ZAR", 0))
    weight_component = rounded_weight * safe_float(rate_card_row.get("Per_KG_Rate_ZAR", 0))
    pre_minimum_amount = base_component + weight_component
    minimum_charge = safe_float(rate_card_row.get("Minimum_Charge_ZAR", 0))
    minimum_applied_amount = max(pre_minimum_amount, minimum_charge)
    fuel_component = minimum_applied_amount * (safe_float(rate_card_row.get("Fuel_Surcharge_Pct", 0)) / 100)
    expected_total = minimum_applied_amount + fuel_component
    return {
        "Chargeable_Weight_KG": round(chargeable_weight, 4),
        "Rounded_Weight_KG": round(rounded_weight, 4),
        "Base_Component_ZAR": round(base_component, 2),
        "Weight_Component_ZAR": round(weight_component, 2),
        "Pre_Minimum_ZAR": round(pre_minimum_amount, 2),
        "Minimum_Applied_ZAR": round(minimum_applied_amount, 2),
        "Fuel_Component_ZAR": round(fuel_component, 2),
        "Expected_Rate_ZAR": round(expected_total, 2),
    }


def apply_rate_card_analysis(analysis_data, rate_card_df, courier_provider, volumetric_divisor):
    working = analysis_data.copy()
    cleaned_card = clean_rate_card(rate_card_df, volumetric_divisor)
    defaults = {
        "Rate_Card_Matched": False,
        "Rate_Card_Confidence": 0,
        "Rate_Card_Reason": "No client rate card loaded; no Tier B claim created.",
        "Expected_Rate_ZAR": 0.0,
        "Rate_Card_Delta_ZAR": 0.0,
        "Rate_Card_Service_Level": "",
        "Rate_Card_Weight_Band": "",
        "Rate_Card_Rounded_Weight_KG": 0.0,
        "Rate_Card_Base_Component_ZAR": 0.0,
        "Rate_Card_Weight_Component_ZAR": 0.0,
        "Rate_Card_Fuel_Component_ZAR": 0.0,
    }
    for column, default in defaults.items():
        working[column] = default
    if cleaned_card.empty or working.empty:
        return working, cleaned_card

    for idx, row in working.iterrows():
        match, reason, confidence = match_rate_card_row(row, cleaned_card, courier_provider)
        if match is None:
            working.at[idx, "Rate_Card_Reason"] = reason
            continue
        expected_parts = calculate_expected_rate(row, match)
        expected = expected_parts["Expected_Rate_ZAR"]
        rounded_weight = expected_parts["Rounded_Weight_KG"]
        charged = safe_float(row.get("Charged_Rate_ZAR", row.get("Billed_Cost_ZAR", row.get("cost", 0))))
        delta = max(0, charged - expected)
        formula = (
            f"Expected {expected:.2f} = max(min {safe_float(match.get('Minimum_Charge_ZAR', 0)):.2f}, "
            f"base {expected_parts['Base_Component_ZAR']:.2f} + rounded kg {rounded_weight:.2f} × per kg {safe_float(match.get('Per_KG_Rate_ZAR', 0)):.2f}) "
            f"+ fuel {expected_parts['Fuel_Component_ZAR']:.2f}."
        )
        working.at[idx, "Rate_Card_Matched"] = True
        working.at[idx, "Rate_Card_Confidence"] = confidence
        working.at[idx, "Rate_Card_Reason"] = f"{reason} {formula}"
        working.at[idx, "Expected_Rate_ZAR"] = expected
        working.at[idx, "Rate_Card_Delta_ZAR"] = round(delta, 2)
        working.at[idx, "Rate_Card_Service_Level"] = safe_text(match.get("Service_Level", "Any"), "Any")
        working.at[idx, "Rate_Card_Weight_Band"] = f"{safe_float(match.get('Min_Weight_KG', 0)):.2f}-{safe_float(match.get('Max_Weight_KG', 0)):.2f} kg; rounded {rounded_weight:.2f} kg"
        working.at[idx, "Rate_Card_Rounded_Weight_KG"] = rounded_weight
        working.at[idx, "Rate_Card_Base_Component_ZAR"] = expected_parts["Base_Component_ZAR"]
        working.at[idx, "Rate_Card_Weight_Component_ZAR"] = expected_parts["Weight_Component_ZAR"]
        working.at[idx, "Rate_Card_Fuel_Component_ZAR"] = expected_parts["Fuel_Component_ZAR"]
    return working, cleaned_card


EVIDENCE_TIERS = {
    "A": "A_DIRECT_MONETARY_PROOF",
    "B": "B_RATE_CARD_REPRODUCIBLE",
    "C": "C_STATISTICAL_OR_PATTERN_ANOMALY",
    "D": "D_OPERATIONAL_ESTIMATE",
    "E": "E_DATA_QUALITY_WARNING",
}

RECOVERY_DIRECT = "DIRECT_RECOVERY"
RECOVERY_QUERY = "ACCOUNT_MANAGER_QUERY"
RECOVERY_OPERATIONAL = "OPERATIONAL_EXPOSURE"
RECOVERY_DATA_QUALITY = "DATA_QUALITY_ONLY"

LABEL_OVERRIDES = {
    "DIRECT_RECOVERY": "Direct Recovery",
    "ACCOUNT_MANAGER_QUERY": "Account Manager Query",
    "OPERATIONAL_EXPOSURE": "Operational Exposure",
    "DATA_QUALITY_ONLY": "Data Quality Only",
    "DIRECT_RATE_DELTA": "Direct Rate Delta",
    "RATE_CARD_REPRODUCIBLE": "Rate Card Reproducible",
    "WEIGHT_EXPOSURE_ESTIMATE": "Weight Exposure Estimate",
    "SKYNET_SAME_LANE_SPIKE": "Same-Lane Cost Spike",
    "MULTI_ITEM_PACKAGING_BLOAT": "Multi-Item Packaging Bloat",
    "PACKAGING_VOL_LEAK": "Packaging Volumetric Leak",
    "DATA_QUALITY_RISK": "Data Quality Risk",
    EVIDENCE_TIERS["A"]: "Tier A - Direct Monetary Proof",
    EVIDENCE_TIERS["B"]: "Tier B - Rate Card Reproducible",
    EVIDENCE_TIERS["C"]: "Tier C - Statistical or Pattern Anomaly",
    EVIDENCE_TIERS["D"]: "Tier D - Operational Estimate",
    EVIDENCE_TIERS["E"]: "Tier E - Data Quality Warning",
}


def humanize_label(value):
    if pd.isna(value):
        return ""
    text = str(value).strip()
    if text in LABEL_OVERRIDES:
        return LABEL_OVERRIDES[text]
    if re.match(r"^[A-E]_", text):
        tier, rest = text.split("_", 1)
        return f"Tier {tier} - {rest.replace('_', ' ').title()}"
    return text.replace("_", " ").title()


def humanize_bool(value):
    if isinstance(value, str):
        return "Yes" if value.strip().lower() in {"true", "yes", "1", "y"} else "No"
    return "Yes" if bool(value) else "No"


def client_column_label(column_name, currency_symbol):
    labels = {
        "Finding_Category": "Finding Type",
        "Evidence_Tier": "Evidence Tier",
        "Confidence_Score": "Confidence Score",
        "Recovery_Type": "Recovery Type",
        "Dispute_Ready": "Dispute Ready",
        "Manual_Review_Required": "Manual Review Required",
        "Direct_Recovery_ZAR": f"Recovery Candidate ({currency_symbol})",
        "Operational_Exposure_ZAR": f"Operational Exposure ({currency_symbol})",
        "Waybill_ID": "Waybill ID",
        "Order_ID": "Order ID",
        "Service_Level": "Service Level",
        "Actual_Weight_KG": "Actual Weight (kg)",
        "Billed_Weight_KG": "Billed Weight (kg)",
        "Expected_Weight_KG": "Expected Weight (kg)",
        "Accepted_Rate_ZAR": f"Accepted Rate ({currency_symbol})",
        "Charged_Rate_ZAR": f"Charged Rate ({currency_symbol})",
        "Expected_Rate_ZAR": f"Expected Rate ({currency_symbol})",
        "Rate_Card_Delta_ZAR": f"Rate Card Difference ({currency_symbol})",
        "Rate_Card_Confidence": "Rate Card Confidence",
        "Rate_Card_Reason": "Rate Card Reason",
        "Claim_Amount_ZAR": f"Claim Amount ({currency_symbol})",
        "Recommended_Action": "Recommended Action",
        "Explainability_Text": "Plain-English Explanation",
        "Evidence_Fields": "Evidence Fields Present",
    }
    return labels.get(column_name, humanize_label(column_name))


def prepare_client_facing_findings(frame, currency_symbol):
    export_frame = frame.copy() if frame is not None else pd.DataFrame()
    if export_frame.empty:
        return export_frame.rename(columns=lambda column: client_column_label(column, currency_symbol))

    for column in ["Finding_Category", "Evidence_Tier", "Recovery_Type"]:
        if column in export_frame.columns:
            export_frame[column] = export_frame[column].apply(humanize_label)
    for column in ["Dispute_Ready", "Manual_Review_Required"]:
        if column in export_frame.columns:
            export_frame[column] = export_frame[column].apply(humanize_bool)
    if "Evidence_Fields" in export_frame.columns:
        export_frame["Evidence_Fields"] = export_frame["Evidence_Fields"].fillna("").astype(str).str.replace("_", " ", regex=False)

    preferred_order = [
        "Waybill_ID", "Order_ID", "Courier", "Service_Level", "Destination", "Finding_Category",
        "Evidence_Tier", "Recovery_Type", "Dispute_Ready", "Manual_Review_Required", "Confidence_Score",
        "Direct_Recovery_ZAR", "Operational_Exposure_ZAR", "Accepted_Rate_ZAR", "Charged_Rate_ZAR",
        "Expected_Rate_ZAR", "Rate_Card_Delta_ZAR", "Reason", "Recommended_Action",
        "Explainability_Text", "Rate_Card_Reason", "Evidence_Fields",
    ]
    ordered_columns = [column for column in preferred_order if column in export_frame.columns]
    ordered_columns += [column for column in export_frame.columns if column not in ordered_columns]
    export_frame = export_frame[ordered_columns]
    return export_frame.rename(columns=lambda column: client_column_label(column, currency_symbol))


def score_finding_confidence(evidence_tier, has_waybill=True, known_courier=True, source_quality="clean"):
    base_scores = {
        EVIDENCE_TIERS["A"]: 95,
        EVIDENCE_TIERS["B"]: 85,
        EVIDENCE_TIERS["C"]: 70,
        EVIDENCE_TIERS["D"]: 55,
        EVIDENCE_TIERS["E"]: 25,
    }
    score = base_scores.get(evidence_tier, 40)
    if not has_waybill:
        score -= 15
    if not known_courier:
        score -= 10
    if source_quality == "weak":
        score -= 10
    return max(0, min(100, int(score)))


def build_finding_explanation(category, amount, reason, evidence_tier, recovery_type, currency_prefix):
    amount_text = format_currency(amount, currency_prefix)
    if recovery_type == RECOVERY_DIRECT:
        return f"{amount_text} is classified as direct recovery because the upload contains direct monetary evidence. {reason} Evidence tier: {evidence_tier}."
    if recovery_type == RECOVERY_QUERY:
        return f"{amount_text} should be queried with the courier account manager. The row is pattern-based rather than a guaranteed credit. {reason} Evidence tier: {evidence_tier}."
    if recovery_type == RECOVERY_OPERATIONAL:
        return f"{amount_text} is operational exposure, not an automatic courier refund. Use it to discuss packaging, divisor, weight capture, or rate-card rules. {reason} Evidence tier: {evidence_tier}."
    return f"This row is a data-quality warning. {reason} Evidence tier: {evidence_tier}."


def make_finding(row, category, evidence_tier, recovery_type, direct_amount, operational_amount, reason, action, currency_prefix, manual_review=False, force_dispute_ready=False, confidence_override=None):
    waybill = safe_text(row.get("Waybill_ID", row.get("Tracking_Number", row.get("Order_ID", ""))), "")
    order_id = safe_text(row.get("Order_ID", waybill), "")
    confidence = int(confidence_override) if confidence_override is not None else score_finding_confidence(evidence_tier, bool(waybill or order_id), True)
    dispute_ready = recovery_type == RECOVERY_DIRECT and direct_amount > 0 and confidence >= 80 and (evidence_tier == EVIDENCE_TIERS["A"] or force_dispute_ready)
    amount = direct_amount if direct_amount > 0 else operational_amount
    return {
        "Finding_Category": category,
        "Evidence_Tier": evidence_tier,
        "Confidence_Score": confidence,
        "Recovery_Type": recovery_type,
        "Dispute_Ready": bool(dispute_ready),
        "Manual_Review_Required": bool(manual_review or confidence < 70),
        "Direct_Recovery_ZAR": round(safe_float(direct_amount), 2),
        "Operational_Exposure_ZAR": round(safe_float(operational_amount), 2),
        "Waybill_ID": waybill,
        "Order_ID": order_id,
        "SKU": safe_text(row.get("SKU", ""), ""),
        "Courier": safe_text(row.get("Courier", row.get("courier", "")), ""),
        "Service_Level": safe_text(row.get("Service_Level", row.get("Service level", row.get("Srv", ""))), ""),
        "Origin": safe_text(row.get("Start_Town", row.get("Origin", "")), ""),
        "Destination": safe_text(row.get("Destination", row.get("Province", "")), ""),
        "Actual_Weight_KG": safe_float(row.get("Actual_Weight_KG", row.get("weight", 0))),
        "Billed_Weight_KG": safe_float(row.get("Billed_Vol_KG", row.get("weight", 0))),
        "Expected_Weight_KG": safe_float(row.get("Expected_Weight_KG", row.get("Expected_Weight", 0))),
        "Accepted_Rate_ZAR": safe_float(row.get("Accepted_Rate_ZAR", 0)),
        "Charged_Rate_ZAR": safe_float(row.get("Charged_Rate_ZAR", row.get("Billed_Cost_ZAR", row.get("cost", 0)))),
        "Expected_Rate_ZAR": safe_float(row.get("Expected_Rate_ZAR", 0)),
        "Rate_Card_Delta_ZAR": safe_float(row.get("Rate_Card_Delta_ZAR", 0)),
        "Rate_Card_Confidence": safe_float(row.get("Rate_Card_Confidence", 0)),
        "Rate_Card_Reason": safe_text(row.get("Rate_Card_Reason", ""), ""),
        "Claim_Amount_ZAR": round(safe_float(direct_amount), 2),
        "Reason": reason,
        "Recommended_Action": action,
        "Explainability_Text": build_finding_explanation(category, amount, reason, evidence_tier, recovery_type, currency_prefix),
        "Evidence_Fields": "; ".join([field for field in ["accepted_rate" if safe_float(row.get("Accepted_Rate_ZAR", 0)) else "", "charged_rate" if safe_float(row.get("Charged_Rate_ZAR", 0)) else "", "weight" if safe_float(row.get("Billed_Vol_KG", row.get("weight", 0))) else ""] if field]),
    }


def build_dispute_findings(analysis_data, financial_anomaly_rows, packaging_rows, skipped_rows, courier_provider, currency_prefix):
    findings = []

    if analysis_data is not None and not analysis_data.empty:
        direct_amounts = pd.to_numeric(analysis_data["Courier_Recoverable_ZAR"], errors="coerce").fillna(0) if "Courier_Recoverable_ZAR" in analysis_data.columns else pd.Series(0, index=analysis_data.index)
        operational_amounts = pd.to_numeric(analysis_data["Operational_Exposure_ZAR"], errors="coerce").fillna(0) if "Operational_Exposure_ZAR" in analysis_data.columns else pd.Series(0, index=analysis_data.index)
        direct_rows = analysis_data[direct_amounts > 0]
        for _, row in direct_rows.iterrows():
            findings.append(make_finding(
                row,
                "DIRECT_RATE_DELTA",
                EVIDENCE_TIERS["A"],
                RECOVERY_DIRECT,
                safe_float(row.get("Courier_Recoverable_ZAR", 0)),
                0,
                safe_text(row.get("Anomaly_Reason", row.get("Rate_Delta_Reason", "Charged rate exceeds accepted rate.")), "Charged rate exceeds accepted rate."),
                "Send this waybill in the dispute pack and request credit or written surcharge evidence.",
                currency_prefix,
            ))

        rate_card_amounts = pd.to_numeric(analysis_data["Rate_Card_Delta_ZAR"], errors="coerce").fillna(0) if "Rate_Card_Delta_ZAR" in analysis_data.columns else pd.Series(0, index=analysis_data.index)
        rate_card_matched = analysis_data["Rate_Card_Matched"].astype(bool) if "Rate_Card_Matched" in analysis_data.columns else pd.Series(False, index=analysis_data.index)
        rate_card_rows = analysis_data[(rate_card_matched) & (rate_card_amounts > 0) & (direct_amounts <= 0)]
        for _, row in rate_card_rows.iterrows():
            confidence = safe_float(row.get("Rate_Card_Confidence", 0))
            recovery_type = RECOVERY_DIRECT if confidence >= 85 else RECOVERY_QUERY
            findings.append(make_finding(
                row,
                "RATE_CARD_REPRODUCIBLE",
                EVIDENCE_TIERS["B"],
                recovery_type,
                safe_float(row.get("Rate_Card_Delta_ZAR", 0)),
                0,
                safe_text(row.get("Rate_Card_Reason", "Charged rate exceeds expected rate from the loaded client rate card."), "Charged rate exceeds expected rate from the loaded client rate card."),
                "Verify rate-card effective date, then send this expected-vs-charged calculation to the courier account manager.",
                currency_prefix,
                manual_review=confidence < 85,
                force_dispute_ready=confidence >= 85,
                confidence_override=confidence,
            ))

        operational_rows = analysis_data[(operational_amounts > 0) & (direct_amounts <= 0) & ~(rate_card_matched & (rate_card_amounts > 0))]
        for _, row in operational_rows.iterrows():
            findings.append(make_finding(
                row,
                "WEIGHT_EXPOSURE_ESTIMATE",
                EVIDENCE_TIERS["D"],
                RECOVERY_OPERATIONAL,
                0,
                safe_float(row.get("Operational_Exposure_ZAR", 0)),
                safe_text(row.get("Anomaly_Reason", "Charged weight exceeds expected weight, but no direct money delta is present."), "Charged weight exceeds expected weight, but no direct money delta is present."),
                "Use for packaging/rate-card review; do not submit as a standalone refund claim without courier evidence.",
                currency_prefix,
                manual_review=True,
            ))

    if financial_anomaly_rows is not None and not financial_anomaly_rows.empty:
        for _, row in financial_anomaly_rows.iterrows():
            findings.append(make_finding(
                row,
                "SKYNET_SAME_LANE_SPIKE",
                EVIDENCE_TIERS["C"],
                RECOVERY_QUERY,
                safe_float(row.get("Financial_Excess_ZAR", 0)),
                0,
                safe_text(row.get("Financial_Anomaly_Reason", "Cost is high versus identical lane/service/weight peers."), "Cost is high versus identical lane/service/weight peers."),
                "Ask the courier account manager to explain the same-lane cost spike or issue a credit if unsupported.",
                currency_prefix,
                manual_review=True,
            ))

    if packaging_rows is not None and not packaging_rows.empty:
        for _, row in packaging_rows.iterrows():
            category = "MULTI_ITEM_PACKAGING_BLOAT" if bool(row.get("Is_Multi_Item", False)) else "PACKAGING_VOL_LEAK"
            findings.append(make_finding(
                row,
                category,
                EVIDENCE_TIERS["D"],
                RECOVERY_OPERATIONAL,
                0,
                safe_float(row.get("Avoidable_Volumetric_Leak_ZAR", 0)),
                safe_text(row.get("Packaging_Reason", "Packaging choice appears to create avoidable volumetric exposure."), "Packaging choice appears to create avoidable volumetric exposure."),
                "Fix the pack-station rule or packaging matrix before treating this as courier recovery.",
                currency_prefix,
                manual_review=True,
            ))

    if skipped_rows is not None and not skipped_rows.empty:
        for _, row in skipped_rows.head(100).iterrows():
            findings.append(make_finding(
                row,
                "DATA_QUALITY_RISK",
                EVIDENCE_TIERS["E"],
                RECOVERY_DATA_QUALITY,
                0,
                0,
                safe_text(row.get("Data_Quality_Note", "Missing critical audit fields."), "Missing critical audit fields."),
                "Fix or request cleaner source data before relying on this row.",
                currency_prefix,
                manual_review=True,
            ))

    columns = [
        "Finding_Category", "Evidence_Tier", "Confidence_Score", "Recovery_Type", "Dispute_Ready",
        "Manual_Review_Required", "Direct_Recovery_ZAR", "Operational_Exposure_ZAR", "Waybill_ID",
        "Order_ID", "SKU", "Courier", "Service_Level", "Origin", "Destination", "Actual_Weight_KG",
        "Billed_Weight_KG", "Expected_Weight_KG", "Accepted_Rate_ZAR", "Charged_Rate_ZAR",
        "Expected_Rate_ZAR", "Rate_Card_Delta_ZAR", "Rate_Card_Confidence", "Rate_Card_Reason",
        "Claim_Amount_ZAR", "Reason", "Recommended_Action", "Explainability_Text", "Evidence_Fields",
    ]
    if not findings:
        return pd.DataFrame(columns=columns)
    return pd.DataFrame(findings)[columns].sort_values(["Dispute_Ready", "Confidence_Score", "Direct_Recovery_ZAR", "Operational_Exposure_ZAR"], ascending=[False, False, False, False]).reset_index(drop=True)


def clean_and_standardize_data(df):
    """Normalize courier data while only rejecting rows with missing/zero weight or cost."""
    cleaned = normalize_courier_export_columns(df)
    cleaned.columns = [str(col).strip() for col in cleaned.columns]

    bob_go_dimensions_col = next((col for col in ["Charged dimensions", "Submitted dimensions"] if col in cleaned.columns), None)
    bob_go_parsed = None
    if bob_go_dimensions_col:
        bob_go_parsed = pd.DataFrame(
            cleaned[bob_go_dimensions_col].apply(parse_bob_go_dimensions).tolist(),
            index=cleaned.index,
            columns=["Waybill_ID", "Length_cm", "Width_cm", "Height_cm"],
        )

    rename_map = {}
    used_standard_names = set()
    for original in cleaned.columns:
        if original in {"Tracking_Number", "Destination", "weight", "cost", "dimensions"}:
            continue
        normalized = normalize_label(original)
        standard = BOB_GO_DIRECT_MAPPINGS.get(normalized) or ALIAS_LOOKUP.get(normalized)
        if standard and standard not in used_standard_names and standard not in cleaned.columns:
            rename_map[original] = standard
            used_standard_names.add(standard)
    cleaned = cleaned.rename(columns=rename_map)
    cleaned = coalesce_duplicate_columns(cleaned)

    cleaned = apply_bouncer_bypass(cleaned)

    if "Tracking_Number" in cleaned.columns:
        cleaned["Order_ID"] = cleaned.get("Order_ID", cleaned["Tracking_Number"])
        cleaned["Waybill_ID"] = cleaned.get("Waybill_ID", cleaned["Tracking_Number"])
    if "Destination" in cleaned.columns:
        cleaned["Province"] = cleaned.get("Province", cleaned["Destination"])

    actual_weight_source = first_available_column(cleaned, ["Submitted weight (kg)", "Submitted weight", "Tot KG", "ActualWeight", "Actual_Weight_KG", "Actual_Weight_kg", "weight"])
    cost_source = first_available_column(cleaned, ["Charged rate (incl. VAT)", "Courier confirmed rate (incl. VAT)", "Accepted rate (incl. VAT)", "TotalAmount", "Exc Vat ", "Exc Vat", "Billed_Cost_ZAR", "Billed_Shipping_Cost_ZAR", "Rate", "Amount", "cost"])
    billed_vol_source = first_available_column(cleaned, ["Charged weight (kg)", "Charged weight", "Submitted volumetric weight (kg)", "Volumetric_Weight_kg", "Mass Charged", "Inc Mass ", "Tot Vol", "Billed_Vol_KG", "billed volumetric weight"])
    dimensions_source = first_available_column(cleaned, ["Charged dimensions", "Submitted dimensions", "Dimensions", "dimensions"])

    cleaned["Actual_Weight_KG"] = actual_weight_source if actual_weight_source is not None else (billed_vol_source if billed_vol_source is not None else 0.0)
    cleaned["Billed_Cost_ZAR"] = cost_source if cost_source is not None else 0.0
    cleaned["Billed_Vol_KG"] = billed_vol_source if billed_vol_source is not None else 0.0
    cleaned["Dimensions"] = dimensions_source if dimensions_source is not None else ""

    if bob_go_parsed is not None:
        for col in ["Waybill_ID", "Length_cm", "Width_cm", "Height_cm"]:
            cleaned[col] = cleaned[col].combine_first(bob_go_parsed[col]) if col in cleaned.columns else bob_go_parsed[col]
        cleaned["Order_ID"] = cleaned.get("Order_ID", cleaned["Waybill_ID"]).combine_first(cleaned["Waybill_ID"])

    if "Dimensions" in cleaned.columns:
        parsed_dims = cleaned["Dimensions"].apply(parse_dimensions)
        if "Length_cm" not in cleaned.columns:
            cleaned["Length_cm"] = parsed_dims.apply(lambda dims: dims[0] if dims else pd.NA)
        if "Width_cm" not in cleaned.columns:
            cleaned["Width_cm"] = parsed_dims.apply(lambda dims: dims[1] if dims else pd.NA)
        if "Height_cm" not in cleaned.columns:
            cleaned["Height_cm"] = parsed_dims.apply(lambda dims: dims[2] if dims else pd.NA)

    if "Cube" in cleaned.columns and "Billed_Vol_KG" in cleaned.columns:
        cube_values = clean_numeric_series(cleaned["Cube"])
        cleaned["Billed_Vol_KG"] = cleaned["Billed_Vol_KG"].where(cleaned["Billed_Vol_KG"].notna(), cube_values)

    for col in STANDARD_COLUMNS:
        if col not in cleaned.columns:
            cleaned[col] = 0.0 if col in {"Billed_Vol_KG", "Length_cm", "Width_cm", "Height_cm"} else pd.NA

    cleaned["Order_ID"] = cleaned["Order_ID"].fillna(pd.Series([f"ROW-{i + 1}" for i in range(len(cleaned))], index=cleaned.index))
    cleaned["Waybill_ID"] = cleaned["Waybill_ID"].fillna(cleaned["Order_ID"])
    cleaned["SKU"] = cleaned["SKU"].fillna("Unknown SKU")
    cleaned["Province"] = cleaned["Province"].fillna("Unknown")

    cleaned["Actual_Weight_KG"] = clean_numeric_series(cleaned["Actual_Weight_KG"])
    cleaned["Billed_Cost_ZAR"] = clean_money_series(cleaned["Billed_Cost_ZAR"])
    for col in ["Billed_Vol_KG", "Length_cm", "Width_cm", "Height_cm"]:
        cleaned[col] = clean_numeric_series(cleaned[col]) if col in cleaned.columns else 0.0

    cleaned["Billed_Vol_KG"] = cleaned["Billed_Vol_KG"].fillna(0.0)
    missing_actual_with_chargeable = (cleaned["Actual_Weight_KG"] <= 0) & (cleaned["Billed_Vol_KG"] > 0)
    cleaned.loc[missing_actual_with_chargeable, "Actual_Weight_KG"] = cleaned.loc[missing_actual_with_chargeable, "Billed_Vol_KG"]
    for col in ["Length_cm", "Width_cm", "Height_cm"]:
        cleaned[col] = cleaned[col].fillna(0.0)

    missing_weight_or_cost = (cleaned["Actual_Weight_KG"] <= 0) | (cleaned["Billed_Cost_ZAR"] <= 0)
    cleaned["Data_Quality_Issue"] = missing_weight_or_cost
    cleaned["Data_Quality_Note"] = "OK"
    cleaned.loc[missing_weight_or_cost, "Data_Quality_Note"] = "Missing or zero weight/cost; row skipped from audit math."

    return cleaned


def fits_in_package(item_dims, package_dims):
    if any(pd.isna(dim) or safe_float(dim) <= 0 for dim in item_dims + package_dims):
        return False
    return any(all(item <= package for item, package in zip(ordering, package_dims)) for ordering in permutations(item_dims))


def clean_packaging_matrix(packaging_matrix):
    matrix = packaging_matrix.copy()
    required_columns = ["Package", "L", "W", "H", "cost"]
    for col in required_columns:
        if col not in matrix.columns:
            matrix[col] = pd.NA
    for col in ["L", "W", "H", "cost"]:
        matrix[col] = pd.to_numeric(matrix[col], errors="coerce")
    matrix = matrix.dropna(subset=required_columns)
    matrix = matrix[(matrix["L"] > 0) & (matrix["W"] > 0) & (matrix["H"] > 0)]
    matrix["Volume_cm3"] = matrix["L"] * matrix["W"] * matrix["H"]
    return matrix.sort_values("Volume_cm3").reset_index(drop=True)


def run_single_item_repack(row, packaging_matrix, penalty_rate, divisor):
    actual_weight = safe_float(row["Actual_Weight_KG"])
    billed_vol = safe_float(row["Billed_Vol_KG"])
    item_dims = [safe_float(row["Length_cm"]), safe_float(row["Width_cm"]), safe_float(row["Height_cm"])]
    billed_weight = max(actual_weight, billed_vol)
    current_penalty = max(0, billed_weight - actual_weight) * penalty_rate

    base = {
        "Packaging_Flag": False,
        "Packaging_Reason": "No lower-cost flyer fit found.",
        "Recommended_Package": "No recommendation",
        "Optimized_Vol_KG": pd.NA,
        "Avoidable_Volumetric_Leak_ZAR": 0.0,
    }

    try:
        if actual_weight <= 0 or billed_vol <= 0 or any(dim <= 0 for dim in item_dims):
            return {**base, "Packaging_Reason": "Missing or invalid single-item dimensions/weight."}

        if packaging_matrix.empty:
            return {**base, "Packaging_Reason": "Packaging matrix has no valid numeric package rows."}

        for _, package in packaging_matrix.iterrows():
            usable_factor = 0.85 if bool(package.get("Fragile/Void Fill Required", False)) else 1.0
            package_dims = [safe_float(package["L"]) * usable_factor, safe_float(package["W"]) * usable_factor, safe_float(package["H"]) * usable_factor]
            if fits_in_package(item_dims, package_dims):
                package_name = str(package["Package"])
                optimized_vol = (safe_float(package["L"]) * safe_float(package["W"]) * safe_float(package["H"])) / max(divisor, 1)
                optimized_billed_weight = max(actual_weight, optimized_vol)
                optimized_penalty = max(0, optimized_billed_weight - actual_weight) * penalty_rate
                net_savings = max(0, current_penalty - optimized_penalty - safe_float(package["cost"]))
                flyer_fit = bool(FLYER_PATTERN.search(package_name))
                return {
                    "Packaging_Flag": flyer_fit and net_savings > 0,
                    "Packaging_Reason": "Single-item order fits into a standard flyer." if flyer_fit and net_savings > 0 else "Fits smaller packaging but no net flyer leak after material cost.",
                    "Recommended_Package": package_name,
                    "Optimized_Vol_KG": optimized_vol,
                    "Avoidable_Volumetric_Leak_ZAR": net_savings if flyer_fit else 0.0,
                }
        return {**base, "Packaging_Reason": "No configured package fits this order; marked as custom/oversized.", "Recommended_Package": "Custom/Oversized"}
    except Exception as exc:
        order_id = row.get("Order_ID", "unknown")
        st.warning(f"Packaging calculation failed for row {row.name} / order {order_id}: {exc}")
        return {**base, "Packaging_Reason": f"Packaging calculation error: {exc}", "Recommended_Package": "Review manually"}


def find_standard_column(df, aliases):
    normalized_lookup = {normalize_label(column): column for column in df.columns}
    for alias in aliases:
        normalized_alias = normalize_label(alias)
        if normalized_alias in normalized_lookup:
            return normalized_lookup[normalized_alias]
    for column in df.columns:
        normalized_column = normalize_label(column)
        if any(normalize_label(alias) in normalized_column for alias in aliases):
            return column
    return None


def generate_client_summary(df, anomalies_df):
    shipment_count = len(df)
    if anomalies_df is None or anomalies_df.empty:
        return (
            f"Courier audit complete: {shipment_count:,} shipments were reviewed.\n\n"
            "No material financial or weight anomalies were detected in the current ruleset. "
            "This means the file did not show surcharge spikes or same-weight base-charge mismatches above the configured thresholds."
        )

    financial_recovery = safe_float(anomalies_df.get("Financial_Excess_ZAR", pd.Series(dtype=float)).sum())
    courier_recovery = safe_float(anomalies_df.get("Courier_Recoverable_ZAR", anomalies_df.get("Recoverable_Overcharge_ZAR", pd.Series(dtype=float))).sum())
    operational_exposure = safe_float(anomalies_df.get("Operational_Exposure_ZAR", pd.Series(dtype=float)).sum())
    total_recovery = financial_recovery + courier_recovery

    amount_columns = [column for column in ["Financial_Excess_ZAR", "Courier_Recoverable_ZAR", "Recoverable_Overcharge_ZAR"] if column in anomalies_df.columns]
    ranked = anomalies_df.copy()
    ranked["_Largest_Direct_Amount"] = ranked[amount_columns].apply(pd.to_numeric, errors="coerce").fillna(0).max(axis=1) if amount_columns else 0
    largest = ranked.sort_values("_Largest_Direct_Amount", ascending=False).iloc[0]
    largest_amount = safe_float(largest.get("_Largest_Direct_Amount", 0))

    weight = safe_float(largest.get("Financial_Weight_KG", largest.get("Actual_Weight_KG", 0)))
    destination = safe_text(largest.get("Destination", largest.get("Province", "Unknown destination")), "Unknown destination")
    reason = safe_text(largest.get("Financial_Anomaly_Reason", largest.get("Anomaly_Reason", "unexplained courier charge")), "unexplained courier charge")

    return (
        f"Courier audit complete: {shipment_count:,} shipments were reviewed.\n\n"
        f"The audit identified {format_currency(total_recovery, currency_prefix)} in direct, evidence-backed recovery candidates. "
        f"A further {format_currency(operational_exposure, currency_prefix)} is shown as estimated operational exposure, not automatic courier credit.\n\n"
        f"The largest direct anomaly was a {weight:.2f} kg package to {destination} with a charge difference of {format_currency(largest_amount, currency_prefix)}. "
        f"Reason flagged: {reason}\n\n"
        "Recommended next step: send the direct recovery rows to the courier account manager, and use the operational exposure rows to discuss packaging, divisor, route and surcharge rules."
    )


def calculate_true_overcharges(df, vol_divisor=5000, penalty_per_kg=15.00, tolerance_kg=1.0):
    """Return only mathematically defensible weight overcharges using courier chargeable-weight logic."""
    if df is None or df.empty:
        return pd.DataFrame()

    working = df.copy()
    column_aliases = {
        "Weight": ["Weight", "Billed_Vol_KG", "Billed_Weight_KG", "weight"],
        "ActualWeight": ["ActualWeight", "Actual_Weight_KG", "Physical_Weight"],
        "Cube": ["Cube", "Physical_Volume_cm3"],
    }

    normalized_inputs = {}
    for required_name, candidates in column_aliases.items():
        source_column = next((column for column in candidates if column in working.columns), None)
        if source_column is None:
            raise ValueError(
                f"Missing required column for true overcharge calculation: {required_name}. "
                f"Available columns: {list(working.columns)}"
            )
        normalized_inputs[required_name] = pd.to_numeric(working[source_column], errors="coerce")

    working["Weight"] = normalized_inputs["Weight"]
    working["ActualWeight"] = normalized_inputs["ActualWeight"].fillna(0)
    working["Cube"] = normalized_inputs["Cube"]

    working = working.dropna(subset=["Weight", "Cube"])
    working = working[(working["Weight"] > 0) & (working["Cube"] >= 0) & (working["ActualWeight"] >= 0)].copy()
    if working.empty:
        return working

    working["Calculated_Vol_Weight"] = working["Cube"] / max(float(vol_divisor), 1.0)
    working["Expected_Weight"] = working[["ActualWeight", "Calculated_Vol_Weight"]].max(axis=1)
    working["Weight_Difference_KG"] = working["Weight"] - working["Expected_Weight"]
    working["Overcharge_Flag"] = working["Weight_Difference_KG"] > float(tolerance_kg)

    overcharges = working[working["Overcharge_Flag"]].copy()
    if overcharges.empty:
        return overcharges

    overcharges["True_Overcharge_ZAR"] = overcharges["Weight_Difference_KG"] * float(penalty_per_kg)
    for column in ["Calculated_Vol_Weight", "Expected_Weight", "Weight_Difference_KG", "True_Overcharge_ZAR"]:
        overcharges[column] = overcharges[column].round(2)

    return overcharges.sort_values("True_Overcharge_ZAR", ascending=False).reset_index(drop=True)


def calculate_financial_anomalies(df):
    """Flag Skynet financial anomalies within identical route, service, and weight groups."""
    if df is None or df.empty:
        return pd.DataFrame()

    required_columns = ["Start_Town", "Destination", "Service_Level", "weight", "cost"]
    missing = [column for column in required_columns if column not in df.columns]
    if missing:
        st.warning(f"Financial anomaly calculation skipped safely: missing columns {', '.join(missing)}")
        return pd.DataFrame()

    working = df.copy()
    working["weight"] = pd.to_numeric(working["weight"], errors="coerce")
    working["cost"] = pd.to_numeric(working["cost"], errors="coerce")
    working = working.dropna(subset=required_columns)
    working = working[(working["weight"] > 0) & (working["cost"] > 0)].copy()

    anomaly_frames = []
    for _, group in working.groupby(["Start_Town", "Destination", "Service_Level", "weight"], dropna=True):
        if len(group) <= 1:
            continue

        try:
            mode_values = group["cost"].mode(dropna=True)
            if mode_values.empty:
                continue
            std_cost = mode_values.iloc[0]
            if pd.isna(std_cost) or std_cost <= 0:
                continue
        except Exception:
            continue

        anomalies = group[group["cost"] > std_cost * 1.1].copy()
        if anomalies.empty:
            continue

        anomalies["Financial_Anomaly_Flag"] = True
        anomalies["Financial_Anomaly_Reason"] = "Cost is more than 10% above the mode for identical route, service, and weight."
        anomalies["Financial_Excess_ZAR"] = anomalies["cost"] - std_cost
        anomaly_frames.append(anomalies)

    if not anomaly_frames:
        return pd.DataFrame()

    return pd.concat(anomaly_frames, ignore_index=True, sort=False)


def classify_velocity(data):
    sku_summary = (
        data.groupby("SKU", as_index=False)
        .agg(
            SKU_Order_Frequency=("Order_ID", "nunique"),
            SKU_Line_Count=("Order_ID", "count"),
            SKU_Total_Shipping_Cost_ZAR=("Billed_Cost_ZAR", "sum"),
            SKU_Distant_Zone_Lines=("Is_Distant_Zone", "sum"),
            SKU_Total_Leakage_ZAR=("Estimated_Loss_ZAR", "sum"),
        )
        .sort_values(["SKU_Order_Frequency", "SKU_Line_Count"], ascending=False)
        .reset_index(drop=True)
    )
    sku_count = len(sku_summary)
    if sku_count == 0:
        sku_summary["Velocity_Class"] = pd.Series(dtype=str)
        return sku_summary

    a_count = max(1, math.ceil(sku_count * 0.20))
    c_count = max(1, math.ceil(sku_count * 0.30))
    c_start = max(a_count, sku_count - c_count)
    sku_summary["Velocity_Class"] = "B"
    sku_summary.loc[: a_count - 1, "Velocity_Class"] = "A"
    sku_summary.loc[c_start:, "Velocity_Class"] = "C"
    return sku_summary


def run_triple_pillar_engine(raw_data, packaging_matrix, penalty_rate, volumetric_divisor, negotiated_divisor):
    if raw_data.attrs.get("pipeline") == "skynet_financial_only":
        data = raw_data.copy()
        data["Data_Quality_Issue"] = False
        data["Data_Quality_Note"] = "Skynet financial-only pipeline; legacy cleanup bypassed."
    else:
        already_standardized = set(STANDARD_COLUMNS).issubset(raw_data.columns) and "Data_Quality_Issue" in raw_data.columns
        data = raw_data.copy() if already_standardized else clean_and_standardize_data(raw_data)
    cleaned_packaging_matrix = clean_packaging_matrix(packaging_matrix)
    valid = data[~data["Data_Quality_Issue"]].copy()
    skipped = data[data["Data_Quality_Issue"]].copy()

    if valid.empty:
        return valid, skipped, pd.DataFrame()

    valid["Order_ID"] = valid["Order_ID"].apply(lambda value: safe_text(value, "Unknown Order"))
    valid["SKU"] = valid["SKU"].apply(lambda value: safe_text(value, "Unknown SKU"))
    valid["Province"] = valid["Province"].apply(lambda value: safe_text(value, "Unknown"))

    valid["Physical_Volume_cm3"] = valid["Length_cm"] * valid["Width_cm"] * valid["Height_cm"]
    if "Cube" in valid.columns:
        cube_values = pd.to_numeric(valid["Cube"], errors="coerce")
        valid["Physical_Volume_cm3"] = valid["Physical_Volume_cm3"].where(valid["Physical_Volume_cm3"] > 0, cube_values.fillna(0))
    valid["Calculated_Vol_KG"] = valid["Physical_Volume_cm3"] / max(volumetric_divisor, 1)
    valid["Billed_Weight_KG"] = valid["Billed_Vol_KG"]
    valid["Expected_Weight_KG"] = valid[["Actual_Weight_KG", "Calculated_Vol_KG"]].max(axis=1)
    valid["Excess_Weight_KG"] = (valid["Billed_Weight_KG"] - valid["Expected_Weight_KG"]).clip(lower=0).fillna(0)
    valid["Estimated_Loss_ZAR"] = valid["Excess_Weight_KG"] * penalty_rate

    order_line_counts = valid.groupby("Order_ID")["SKU"].transform("count")
    valid["Is_Multi_Item"] = order_line_counts > 1
    valid["Is_Distant_Zone"] = valid["Province"].str.contains(DISTANT_ZONE_PATTERN, na=False)

    weight_exposure_mask = valid["Excess_Weight_KG"] > 1.0
    valid["Calculated_Vol_Weight"] = valid["Calculated_Vol_KG"].round(2)
    valid["Expected_Weight"] = valid["Expected_Weight_KG"].round(2)
    valid["Weight_Difference_KG"] = valid["Excess_Weight_KG"].round(2)
    valid["Estimated_Weight_Exposure_ZAR"] = valid["Estimated_Loss_ZAR"].where(weight_exposure_mask, 0).round(2)
    valid["Operational_Exposure_ZAR"] = valid["Estimated_Weight_Exposure_ZAR"].fillna(0)

    rate_deltas = calculate_rate_card_deltas(valid)
    for column in rate_deltas.columns:
        valid[column] = rate_deltas[column]

    direct_recovery_mask = valid["Rate_Delta_Flag"] & (valid["Rate_Delta_ZAR"] > 0)
    valid["Courier_Recoverable_ZAR"] = valid["Rate_Delta_ZAR"].where(direct_recovery_mask, 0).round(2)
    valid["Recoverable_Overcharge_ZAR"] = valid["Courier_Recoverable_ZAR"].fillna(0)
    valid["True_Overcharge_ZAR"] = valid["Courier_Recoverable_ZAR"].fillna(0)
    valid["Anomaly_Flag"] = direct_recovery_mask
    valid["Anomaly_Reason"] = ""
    valid.loc[direct_recovery_mask, "Anomaly_Reason"] = valid.loc[direct_recovery_mask, "Rate_Delta_Reason"]
    valid.loc[~direct_recovery_mask & weight_exposure_mask, "Anomaly_Reason"] = (
        "Estimated operational exposure only: charged weight exceeds expected chargeable weight, but no direct accepted-vs-charged money delta is available."
    )

    valid["Packaging_Flag"] = False
    valid["Packaging_Reason"] = "No packaging leak detected."
    valid["Recommended_Package"] = "No recommendation"
    valid["Optimized_Vol_KG"] = pd.NA
    valid["Avoidable_Volumetric_Leak_ZAR"] = 0.0

    single_mask = ~valid["Is_Multi_Item"]
    if single_mask.any():
        try:
            single_results = valid.loc[single_mask].apply(
                lambda row: run_single_item_repack(row, cleaned_packaging_matrix, penalty_rate, negotiated_divisor), axis=1
            )
            for idx, result in single_results.items():
                for key, value in result.items():
                    valid.at[idx, key] = value
        except Exception as exc:
            st.error(f"Packaging optimization failed before completion: {exc}")
            valid.loc[single_mask, "Packaging_Reason"] = f"Packaging optimization failed: {exc}"
            valid.loc[single_mask, "Recommended_Package"] = "Review manually"

    multi_orders = (
        valid[valid["Is_Multi_Item"]]
        .groupby("Order_ID", as_index=False)
        .agg(
            Combined_Physical_Volume_cm3=("Physical_Volume_cm3", "sum"),
            Combined_Actual_Weight_KG=("Actual_Weight_KG", "sum"),
            Billed_Vol_KG=("Billed_Vol_KG", "max"),
            Line_Count=("SKU", "count"),
        )
    )
    if not multi_orders.empty:
        multi_orders["Allowed_Vol_KG_With_20pct_Buffer"] = (multi_orders["Combined_Physical_Volume_cm3"] * 1.20) / max(volumetric_divisor, 1)
        multi_orders["Multi_Item_Excess_Vol_KG"] = (multi_orders["Billed_Vol_KG"] - multi_orders["Allowed_Vol_KG_With_20pct_Buffer"]).clip(lower=0)
        multi_orders["Multi_Item_Leak_ZAR"] = multi_orders["Multi_Item_Excess_Vol_KG"] * penalty_rate
        multi_lookup = multi_orders.set_index("Order_ID").to_dict("index")
        for order_id, info in multi_lookup.items():
            order_mask = valid["Order_ID"].eq(order_id)
            if info["Multi_Item_Leak_ZAR"] > 0:
                per_line_leak = info["Multi_Item_Leak_ZAR"] / max(info["Line_Count"], 1)
                valid.loc[order_mask, "Packaging_Flag"] = True
                valid.loc[order_mask, "Packaging_Reason"] = "Multi-Item Packaging Bloat: billed volumetric weight exceeds combined item volume plus 20% void-fill allowance."
                valid.loc[order_mask, "Recommended_Package"] = "Review pack station: mixed-basket carton selection"
                valid.loc[order_mask, "Optimized_Vol_KG"] = info["Allowed_Vol_KG_With_20pct_Buffer"]
                valid.loc[order_mask, "Avoidable_Volumetric_Leak_ZAR"] = per_line_leak

    sku_summary = classify_velocity(valid)
    valid = valid.merge(sku_summary, on="SKU", how="left")
    high_shipping_threshold = sku_summary["SKU_Total_Shipping_Cost_ZAR"].median() if not sku_summary.empty else 0
    valid["Capital_Trap_Flag"] = (
        valid["Velocity_Class"].eq("C")
        & (valid["SKU_Total_Shipping_Cost_ZAR"] >= high_shipping_threshold)
        & (valid["SKU_Distant_Zone_Lines"] > 0)
    ).fillna(False)

    return valid, skipped, sku_summary


def extract_pdf_rows(pdf_source):
    rows = []
    with pdfplumber.open(pdf_source) as pdf:
        for page in pdf.pages:
            for table in page.extract_tables() or []:
                if not table or len(table) < 2:
                    continue
                headers = [str(header or "").strip() for header in table[0]]
                for row in table[1:]:
                    if row and any(row):
                        rows.append({headers[i]: row[i] for i in range(min(len(headers), len(row)))})
            text = page.extract_text() or ""
            for line in text.splitlines():
                dims = parse_dimensions(line)
                order_match = re.search(r"\b[A-Z]{1,4}[- ]?\d{4,}|\b\d{8,}\b", line)
                numbers = re.findall(r"\d+(?:\.\d+)?", line)
                if order_match and dims and len(numbers) >= 5:
                    rows.append({"Order_ID": order_match.group(0), "Dimensions": " x ".join(map(str, dims)), "Actual_Weight_KG": numbers[-3], "Billed_Cost_ZAR": numbers[-1]})
    if not rows:
        raise ValueError("No invoice rows found")
    return pd.DataFrame(rows)


def extract_html_rows(file_source):
    try:
        tables = pd.read_html(file_source)
    except ValueError as exc:
        raise ValueError("No HTML tables found in the uploaded file.") from exc
    except Exception as exc:
        raise ValueError(f"Could not parse uploaded HTML file: {exc}") from exc

    if not tables:
        raise ValueError("No HTML tables found in the uploaded file.")

    debug_rows = []
    for index, table in enumerate(tables):
        score, matched_columns = score_table_for_display(table)
        debug_rows.append(f"table {index}: score={score}, shape={table.shape}, matched={matched_columns}")
    st.caption("HTML table scan: " + " | ".join(debug_rows[:6]))
    if len(debug_rows) > 6:
        st.caption(f"HTML table scan found {len(debug_rows)} tables; showing first 6 in the caption.")

    selected_table = select_shipment_table(tables)
    cleaned = clean_selected_table(selected_table)
    if cleaned.empty:
        raise ValueError("HTML file was parsed, but the selected shipment table produced no usable rows.")
    return cleaned


def score_table_for_display(table):
    try:
        from aramex_html_processor import score_table

        return score_table(table)
    except Exception as exc:
        return 0, [f"score failed: {exc}"]


INGESTION_SCHEMA_VERSION = "client_export_mapper_v6"


@st.cache_data(show_spinner="Loading courier shipment data...")
def load_data(file_name, file_bytes, ingestion_schema_version=INGESTION_SCHEMA_VERSION):
    if file_bytes is None:
        return clean_and_standardize_data(pd.read_csv(APP_DIR / "mock_shipping_data.csv", sep=None, engine="python", encoding="utf-8-sig"))
    file_buffer = io.BytesIO(file_bytes)
    lower_name = file_name.lower()
    if lower_name.endswith(".csv"):
        return clean_and_standardize_data(pd.read_csv(file_buffer, sep=None, engine="python", encoding="utf-8-sig"))
    if lower_name.endswith(".xlsx"):
        return clean_and_standardize_data(read_excel_with_dynamic_header(file_buffer))
    if lower_name.endswith(".pdf"):
        return clean_and_standardize_data(extract_pdf_rows(file_buffer))
    if lower_name.endswith((".html", ".htm")):
        return clean_and_standardize_data(extract_html_rows(file_buffer))
    raise ValueError("Unsupported file type. Upload CSV, XLSX, PDF, HTML, or HTM.")


class MarginPDF(FPDF):
    NAVY = (31, 41, 55)
    TEXT = (38, 38, 38)
    MUTED = (90, 99, 110)
    BORDER = (215, 221, 229)
    ZEBRA = (246, 248, 250)
    WHITE = (255, 255, 255)

    def header(self):
        self.set_y(10)
        self.set_font("Helvetica", "B", 10)
        self.set_text_color(*self.NAVY)
        self.cell(0, 6, "Margin Diagnostic & Recovery Blueprint", ln=1, align="R")
        self.set_draw_color(*self.BORDER)
        self.line(self.l_margin, 20, self.w - self.r_margin, 20)
        self.ln(8)

    def footer(self):
        self.set_y(-14)
        self.set_font("Helvetica", "", 8)
        self.set_text_color(*self.MUTED)
        self.cell(95, 6, "Prepared by Kyle | Logistics Audit Advisory", align="L")
        self.cell(0, 6, f"Page {self.page_no()} of {{nb}}", align="R")


def generate_dispute_pack_xlsx(dispute_rows, currency_symbol, currency_prefix):
    """Generate a multi-sheet evidence workbook from normalized dispute findings."""
    findings = dispute_rows.copy() if dispute_rows is not None else pd.DataFrame()
    output = io.BytesIO()

    if findings.empty:
        findings = pd.DataFrame(columns=[
            "Waybill_ID", "Order_ID", "Finding_Category", "Evidence_Tier", "Confidence_Score",
            "Recovery_Type", "Direct_Recovery_ZAR", "Operational_Exposure_ZAR", "Recommended_Action",
            "Explainability_Text",
        ])

    dispute_mask = findings["Dispute_Ready"].eq(True) if "Dispute_Ready" in findings.columns else pd.Series(False, index=findings.index)
    query_mask = findings["Recovery_Type"].eq(RECOVERY_QUERY) if "Recovery_Type" in findings.columns else pd.Series(False, index=findings.index)
    operational_amounts = findings["Operational_Exposure_ZAR"] if "Operational_Exposure_ZAR" in findings.columns else pd.Series(dtype=float)
    confidence_scores = findings["Confidence_Score"] if "Confidence_Score" in findings.columns else pd.Series(dtype=float)

    direct_recovery_values = pd.to_numeric(findings.get("Direct_Recovery_ZAR", pd.Series(0, index=findings.index)), errors="coerce").fillna(0)
    operational_values = pd.to_numeric(findings.get("Operational_Exposure_ZAR", pd.Series(0, index=findings.index)), errors="coerce").fillna(0)
    data_quality_mask = findings["Recovery_Type"].eq(RECOVERY_DATA_QUALITY) if "Recovery_Type" in findings.columns else pd.Series(False, index=findings.index)
    operational_mask = findings["Recovery_Type"].eq(RECOVERY_OPERATIONAL) if "Recovery_Type" in findings.columns else pd.Series(False, index=findings.index)

    summary = pd.DataFrame([
        {"Metric": "Dispute-ready claims", "Value": f"{int(dispute_mask.sum()):,}", "What it means": "Rows closest to courier submission, still subject to courier validation."},
        {"Metric": "Dispute-ready recovery candidate", "Value": format_currency(direct_recovery_values[dispute_mask].sum(), currency_prefix), "What it means": "Potential credit amount from rows with the strongest evidence."},
        {"Metric": "Account-manager query rows", "Value": f"{int(query_mask.sum()):,}", "What it means": "Rows to ask the courier/account manager to explain before treating as credits."},
        {"Metric": "Account-manager query amount", "Value": format_currency(direct_recovery_values[query_mask].sum(), currency_prefix), "What it means": "Potential value to query, not guaranteed recovery."},
        {"Metric": "Operational exposure rows", "Value": f"{int(operational_mask.sum()):,}", "What it means": "Prevention/control opportunities, not courier refund claims."},
        {"Metric": "Operational exposure amount", "Value": format_currency(operational_values[operational_mask].sum(), currency_prefix), "What it means": "Estimated preventable leakage from packaging, weight, or process."},
        {"Metric": "Data-quality warning rows", "Value": f"{int(data_quality_mask.sum()):,}", "What it means": "Rows needing cleaner source data before relying on the maths."},
        {"Metric": "Average confidence score", "Value": f"{safe_float(confidence_scores.mean()):.0f}/100", "What it means": "Average evidence strength across all exported rows."},
        {"Metric": "Rows in this workbook", "Value": f"{len(findings):,}", "What it means": "Includes claims, queries, prevention items, and data-quality warnings."},
    ])

    methodology = pd.DataFrame([
        {"Evidence Tier": humanize_label(EVIDENCE_TIERS["A"]), "Meaning": "Direct monetary proof such as accepted rate versus charged or courier-confirmed rate.", "Use": "Best candidate for a dispute-ready claim, subject to courier validation."},
        {"Evidence Tier": humanize_label(EVIDENCE_TIERS["B"]), "Meaning": "Recreated from the loaded client-specific rate card.", "Use": "Strong when the signed rate card and route/service context are confirmed; otherwise query first."},
        {"Evidence Tier": humanize_label(EVIDENCE_TIERS["C"]), "Meaning": "Statistical or same-lane anomaly.", "Use": "Ask the account manager to explain the charge or surcharge."},
        {"Evidence Tier": humanize_label(EVIDENCE_TIERS["D"]), "Meaning": "Operational estimate from weight, packaging, or dimension exposure.", "Use": "Use for prevention and process improvement, not as an automatic refund."},
        {"Evidence Tier": humanize_label(EVIDENCE_TIERS["E"]), "Meaning": "Data-quality warning.", "Use": "Fix source data before relying on row-level maths."},
    ])
    provenance = pd.DataFrame([
        {"Item": "Direct recovery calculation", "Value": "Charged rate minus accepted rate, never below zero."},
        {"Item": "Dispute-ready recovery candidate", "Value": safe_float(direct_recovery_values[dispute_mask].sum())},
        {"Item": "Account-manager query amount", "Value": safe_float(direct_recovery_values[query_mask].sum())},
        {"Item": "Operational exposure amount", "Value": safe_float(operational_values[operational_mask].sum())},
        {"Item": "Operational exposure rule", "Value": "Weight, dimension, and packaging exposure is kept separate and is not treated as direct recovery."},
        {"Item": "Rate-card reconstruction rule", "Value": "Rate-card findings require the client's actual rate card plus courier confirmation before final recovery claims."},
        {"Item": "Credit status", "Value": "All amounts are candidates until a courier response or credit note confirms them."},
    ])

    with pd.ExcelWriter(output, engine="xlsxwriter") as writer:
        workbook = writer.book
        money_fmt = workbook.add_format({"num_format": f"{currency_prefix} #,##0.00"})
        header_fmt = workbook.add_format({"bold": True, "bg_color": "#1F4E79", "font_color": "#FFFFFF", "border": 1})

        sheets = {
            "Executive Summary": summary,
            "Dispute Ready Claims": prepare_client_facing_findings(findings[findings["Dispute_Ready"]].copy() if "Dispute_Ready" in findings.columns else findings.iloc[0:0].copy(), currency_symbol),
            "Rate Card Repro Claims": prepare_client_facing_findings(findings[findings["Evidence_Tier"].eq(EVIDENCE_TIERS["B"])].copy() if "Evidence_Tier" in findings.columns else findings.iloc[0:0].copy(), currency_symbol),
            "Account Manager Queries": prepare_client_facing_findings(findings[findings["Recovery_Type"].eq(RECOVERY_QUERY)].copy() if "Recovery_Type" in findings.columns else findings.iloc[0:0].copy(), currency_symbol),
            "Operational Exposure": prepare_client_facing_findings(findings[findings["Recovery_Type"].eq(RECOVERY_OPERATIONAL)].copy() if "Recovery_Type" in findings.columns else findings.iloc[0:0].copy(), currency_symbol),
            "Data Quality Exceptions": prepare_client_facing_findings(findings[findings["Recovery_Type"].eq(RECOVERY_DATA_QUALITY)].copy() if "Recovery_Type" in findings.columns else findings.iloc[0:0].copy(), currency_symbol),
            "Recovery Provenance": provenance,
            "Methodology": methodology,
        }

        for sheet_name, frame in sheets.items():
            export_frame = frame.copy()
            export_frame.to_excel(writer, index=False, sheet_name=sheet_name)
            worksheet = writer.sheets[sheet_name]
            for col_idx, column_name in enumerate(export_frame.columns):
                worksheet.write(0, col_idx, column_name, header_fmt)
                width = min(max(len(str(column_name)) + 4, 14), 60)
                if not export_frame.empty:
                    width = min(max(width, export_frame[column_name].astype(str).str.len().max() + 4), 60)
                is_money_column = str(column_name).endswith("_ZAR") or f"({currency_symbol})" in str(column_name) or str(column_name) == "Value"
                worksheet.set_column(col_idx, col_idx, width, money_fmt if is_money_column and sheet_name not in {"Executive Summary", "Recovery Provenance"} else None)
            worksheet.freeze_panes(1, 0)
            if len(export_frame.columns) > 0:
                worksheet.autofilter(0, 0, max(len(export_frame), 1), len(export_frame.columns) - 1)

    output.seek(0)
    return output.getvalue()


def choose_pitch_angle(findings_df, operational_exposure_loss, pillar_a_loss, skipped_rows, benchmark_table):
    """Choose a cautious consultant pitch angle from deterministic audit signals."""
    skipped_count = len(skipped_rows) if skipped_rows is not None else 0
    operational_exposure_loss = safe_float(operational_exposure_loss)
    pillar_a_loss = safe_float(pillar_a_loss)

    urgent_benchmark = False
    if benchmark_table is not None and not benchmark_table.empty and "Benchmark_Status" in benchmark_table.columns:
        urgent_benchmark = benchmark_table["Benchmark_Status"].astype(str).str.lower().isin(["urgent", "material"]).any()

    if pillar_a_loss >= 5000 and pillar_a_loss >= operational_exposure_loss * 0.40:
        return {
            "Angle": "Cash recovery first",
            "Why": "The strongest opening is the evidence-backed direct recovery candidate total. Lead with the dispute pack, then widen the conversation into prevention and monthly control.",
            "Caution": "Say this is a direct recovery candidate, not guaranteed credited cash until the courier confirms each waybill.",
        }
    if operational_exposure_loss >= max(5000, pillar_a_loss * 1.5):
        return {
            "Angle": "Governance and prevention",
            "Why": "The biggest signal is operational exposure. Position the work as stopping recurring leakage through rate-card control, packaging rules, and exception monitoring.",
            "Caution": "Do not frame operational exposure as an automatic refund. It is a control and prevention opportunity.",
        }
    if skipped_count >= 10:
        return {
            "Angle": "Data cleanup and control",
            "Why": "Dirty or incomplete rows limit how confidently the client can challenge courier charges. Lead with clean evidence and a repeatable monthly process.",
            "Caution": "Keep claims conservative until source exports and rate-card fields are cleaner.",
        }
    if urgent_benchmark:
        return {
            "Angle": "Monthly monitoring retainer",
            "Why": "The benchmark view shows the client is outside at least one internal control band. Position the offer as ongoing shipping-margin governance.",
            "Caution": "Label benchmarks as internal heuristics until enough outcome history is collected.",
        }
    return {
        "Angle": "Baseline trust builder",
        "Why": "This upload is best used to create a low-risk baseline, educate the client, and earn the next dataset or rate card.",
        "Caution": "Avoid dramatic recovery language. Sell repeatable control and evidence quality.",
    }


def build_call_script(client_name, courier_provider, pitch_angle, direct_recovery, operational_exposure, packaging_leak, financial_query, currency_prefix):
    client_name = safe_text(client_name, "the client") or "the client"
    angle = safe_text(pitch_angle.get("Angle", "Baseline trust builder") if isinstance(pitch_angle, dict) else pitch_angle, "Baseline trust builder")
    return [
        f"Opening: I have reviewed {client_name}'s courier data and split the findings into what we can safely query now versus what should be treated as operational exposure.",
        f"Positioning: The strongest angle is '{angle}'. Lead with the clearest business risk, not every row in the spreadsheet.",
        f"Numbers: Direct recovery candidates total {format_currency(direct_recovery, currency_prefix)}. Operational exposure totals {format_currency(operational_exposure, currency_prefix)}. Packaging leakage totals {format_currency(packaging_leak, currency_prefix)}. Financial/routing query amounts total {format_currency(financial_query, currency_prefix)}.",
        f"Courier context: Because the selected courier is {courier_provider}, ask for the signed rate card, surcharge rules, and the account manager's explanation before treating rate-card or same-lane items as confirmed recoveries.",
        "Dashboard walkthrough: Start with the Founder Dashboard, then open the Dispute Command Centre to show evidence tiers. Explain that Tier A is closest to a claim; Tier D is a prevention opportunity.",
        "Close: Ask permission to send the dispute pack, request the rate card, fix the highest-value packaging rule, and schedule the next monthly audit.",
    ]


def build_objection_responses(currency_prefix):
    return [
        ("Is this definitely recoverable?", "No. The dispute-ready number is an evidence-backed recovery candidate. The courier still needs to validate the waybill, surcharge, credit-note, or rate-card context."),
        ("Why is operational exposure not a refund?", "Because weight, dimension, or packaging signals can show avoidable cost without proving the courier charged incorrectly. Use it to change process and negotiate better controls."),
        ("Could the courier have valid surcharges?", "Yes. That is why Tier C and some Tier B items are positioned as account-manager queries until the signed rate card and surcharge rules are checked."),
        ("Why pay monthly if the direct recovery is small?", f"The monthly value is governance: fewer billing surprises, cleaner data, faster dispute packs, packaging discipline, and a growing benchmark history. It is not only about this month's {currency_prefix} recovery."),
        ("Can we do this ourselves?", "They can try, but the value is the repeatable evidence model, tiering discipline, exports, benchmarks, and knowing what not to overclaim."),
    ]


def format_top_finding_rows(findings_df, limit=5):
    if findings_df is None or findings_df.empty:
        return []
    frame = findings_df.copy()
    frame["Sort_Amount"] = 0.0
    for col in ["Direct_Recovery_ZAR", "Operational_Exposure_ZAR"]:
        if col in frame.columns:
            frame["Sort_Amount"] += pd.to_numeric(frame[col], errors="coerce").fillna(0)
    frame = frame.sort_values("Sort_Amount", ascending=False).head(limit)
    rows = []
    for _, row in frame.iterrows():
        rows.append({
            "Waybill": safe_text(row.get("Waybill_ID", ""), "n/a") or "n/a",
            "Order": safe_text(row.get("Order_ID", ""), "n/a") or "n/a",
            "Tier": safe_text(row.get("Evidence_Tier", ""), "n/a"),
            "Type": safe_text(row.get("Recovery_Type", ""), "n/a"),
            "Amount": safe_float(row.get("Direct_Recovery_ZAR", 0)) or safe_float(row.get("Operational_Exposure_ZAR", 0)),
            "Confidence": safe_float(row.get("Confidence_Score", 0)),
            "Talk_Point": safe_text(row.get("Recommended_Action", row.get("Explainability_Text", "Review this row with the client.")), "Review this row with the client."),
        })
    return rows


def build_intro_stage_offer(estimated_monthly_spend, direct_monthly_leakage, operational_exposure, currency_prefix):
    """Return realistic introduction-stage offer guidance for early client acquisition."""
    estimated_monthly_spend = safe_float(estimated_monthly_spend)
    direct_monthly_leakage = safe_float(direct_monthly_leakage)
    operational_exposure = safe_float(operational_exposure)
    value_signal = max(direct_monthly_leakage, operational_exposure * 0.25, estimated_monthly_spend * 0.03)

    packages = {
        "Starter Leak Snapshot": {
            "condition": f"Best below {format_currency_whole(50_000, currency_prefix)} monthly courier spend or for cold prospects.",
            "audit_min": 2500,
            "audit_max": 4500,
            "retainer_min": 1500,
            "retainer_max": 3500,
            "success_fee": "Optional 10% of credited recoveries if you help submit/follow up.",
            "positioning": "Low-friction proof audit: find the leak, separate claim candidates from operational exposure, and earn trust.",
        },
        "Recovery Assist": {
            "condition": f"Best around {format_currency_whole(50_000, currency_prefix)}-{format_currency_whole(150_000, currency_prefix)} monthly courier spend or where direct recovery candidates are visible.",
            "audit_min": 4500,
            "audit_max": 9500,
            "retainer_min": 3500,
            "retainer_max": 7500,
            "success_fee": "10%-20% of credited recoveries, capped or negotiable for beta clients.",
            "positioning": "Audit plus dispute-pack support: help the client turn evidence into courier/account-manager actions.",
        },
        "Margin Control Retainer": {
            "condition": f"Best above {format_currency_whole(150_000, currency_prefix)} monthly courier spend or when repeat leakage/operational exposure is material.",
            "audit_min": 9500,
            "audit_max": 18000,
            "retainer_min": 7500,
            "retainer_max": 15000,
            "success_fee": "Optional 10%-15% success fee on credited recoveries if retainer is discounted.",
            "positioning": "Monthly shipping-margin control: dispute workflow, rate-card control, packaging prevention, and benchmark history.",
        },
    }

    if estimated_monthly_spend < 50_000 and direct_monthly_leakage < 8_000:
        key = "Starter Leak Snapshot"
    elif estimated_monthly_spend <= 150_000 or direct_monthly_leakage >= 8_000:
        key = "Recovery Assist"
    else:
        key = "Margin Control Retainer"

    offer = packages[key].copy()
    offer["package_name"] = key
    offer["recommended_audit_fee"] = int(round(min(offer["audit_max"], max(offer["audit_min"], value_signal * 0.35)) / 500) * 500)
    offer["recommended_retainer"] = int(round(min(offer["retainer_max"], max(offer["retainer_min"], value_signal * 0.18)) / 500) * 500)
    offer["audit_range"] = f"{format_currency_whole(offer['audit_min'], currency_prefix)}-{format_currency_whole(offer['audit_max'], currency_prefix)}"
    offer["retainer_range"] = f"{format_currency_whole(offer['retainer_min'], currency_prefix)}-{format_currency_whole(offer['retainer_max'], currency_prefix)}/mo"
    offer["cta"] = "Book a Courier Leak Audit. Send us your courier export and we'll separate recoverable claim candidates from operational leakage within 48 hours."
    offer["intro_note"] = "Introduction-stage pricing: reduce adoption friction, earn outcome data, and ask for testimonial/case-study permission when discounting."
    offer["alternatives"] = packages
    return offer


PIPELINE_STAGES = [
    "Prospect identified",
    "Data requested",
    "Data received",
    "Audit completed",
    "Memo sent",
    "Call booked",
    "Dispute pack sent",
    "Outcome pending",
    "Credit received",
    "Monthly monitoring proposed",
    "Converted",
    "Lost / nurture",
]


def build_client_pipeline_template(client_name, courier_provider, estimated_monthly_spend, direct_recovery, credited_amount, offer, currency_prefix):
    """Build a one-row editable client pipeline seed from the current audit."""
    client_name = safe_text(client_name, "Local Client") or "Local Client"
    offer = offer or {}
    stage = "Audit completed" if safe_float(direct_recovery) > 0 else "Data received"
    if safe_float(credited_amount) > 0:
        stage = "Credit received"
    return pd.DataFrame([{
        "Client": client_name,
        "Stage": stage,
        "Courier": courier_provider,
        "Monthly_Spend_ZAR": safe_float(estimated_monthly_spend),
        "Direct_Candidates_ZAR": safe_float(direct_recovery),
        "Credited_ZAR": safe_float(credited_amount),
        "Recommended_Package": offer.get("package_name", "Starter Leak Snapshot"),
        "Audit_Fee_ZAR": safe_float(offer.get("recommended_audit_fee", 0)),
        "Monthly_Retainer_ZAR": safe_float(offer.get("recommended_retainer", 0)),
        "Next_Follow_Up": "",
        "Next_Action": "Send client audit memo and book a 15-minute review call.",
        "Notes": "",
    }])


def summarize_pipeline(pipeline_df):
    if pipeline_df is None or pipeline_df.empty:
        return {
            "client_count": 0,
            "active_count": 0,
            "converted_count": 0,
            "credited_total": 0.0,
            "projected_setup": 0.0,
            "projected_mrr": 0.0,
        }
    frame = pipeline_df.copy()
    stage = frame.get("Stage", pd.Series("", index=frame.index)).astype(str)
    active_mask = ~stage.isin(["Converted", "Lost / nurture"])
    return {
        "client_count": len(frame),
        "active_count": int(active_mask.sum()),
        "converted_count": int(stage.eq("Converted").sum()),
        "credited_total": safe_float(pd.to_numeric(frame.get("Credited_ZAR", 0), errors="coerce").fillna(0).sum()),
        "projected_setup": safe_float(pd.to_numeric(frame.get("Audit_Fee_ZAR", 0), errors="coerce").fillna(0).sum()),
        "projected_mrr": safe_float(pd.to_numeric(frame.get("Monthly_Retainer_ZAR", 0), errors="coerce").fillna(0).sum()),
    }


def recommend_pipeline_action(stage):
    stage = safe_text(stage, "").lower()
    if "prospect" in stage:
        return "Send the Courier Leak Audit intro message and ask for one courier export."
    if "data requested" in stage:
        return "Follow up for the export, rate card, and account-manager contact."
    if "data received" in stage:
        return "Run the audit and prepare the client audit memo."
    if "audit completed" in stage:
        return "Send the memo, proof angle, and book the walkthrough call."
    if "memo sent" in stage:
        return "Follow up with two time options for a 15-minute review call."
    if "call booked" in stage:
        return "Prepare the briefing playbook and objection responses before the call."
    if "dispute pack" in stage:
        return "Ask the courier for row-level accepted/rejected/credited responses."
    if "outcome pending" in stage:
        return "Follow up every 7 days until credited, rejected, or more evidence is requested."
    if "credit received" in stage:
        return "Record the credited amount and ask permission for an anonymized case study."
    if "monitoring proposed" in stage:
        return "Anchor monthly monitoring around prevention, outcome tracking, and rate-card control."
    if "converted" in stage:
        return "Schedule next month's audit and keep tracking outcomes."
    return "Keep in nurture with a lighter baseline audit or referral ask."


def score_prospect_fit(monthly_spend, monthly_orders, courier_count, has_rate_card, has_export_access, uses_aggregator, has_dispute_process, has_packaging_rules, pain_level, urgency):
    monthly_spend = safe_float(monthly_spend)
    monthly_orders = safe_float(monthly_orders)
    courier_count = max(1, int(safe_float(courier_count, 1)))
    pain_level = int(safe_float(pain_level))
    urgency = int(safe_float(urgency))
    score = 0
    risks = []

    score += min(25, int(monthly_spend / 4000))
    score += min(15, int(monthly_orders / 80))
    score += min(10, courier_count * 3)
    score += 15 if has_export_access else 0
    score += 10 if has_rate_card else 0
    score += 8 if uses_aggregator else 4
    score += max(0, 8 - (4 if has_dispute_process else 0))
    score += max(0, 8 - (4 if has_packaging_rules else 0))
    score += min(12, pain_level * 2)
    score += min(12, urgency * 2)
    score = min(100, int(score))

    if not has_export_access:
        risks.append("No export/invoice access yet; start with education and a simple data request.")
    if not has_rate_card:
        risks.append("No signed rate card yet; Tier B rate-card reconstruction will be weaker until requested.")
    if monthly_spend < 15000:
        risks.append("Courier spend may be too low for a high-touch retainer; use a baseline audit or nurture.")
    if has_dispute_process and has_packaging_rules and pain_level <= 2:
        risks.append("Prospect may feel they already control the problem; lead with independent evidence and benchmark history.")

    if score >= 75:
        band = "Prime target"
        why = "High spend/data complexity and enough pain or urgency to justify an audit plus monitoring conversation."
        best_offer_hint = "Lead with Recovery Assist or Margin Control Retainer if spend supports it."
        pitch_angle = "Monthly courier margin control"
    elif score >= 55:
        band = "Good pilot"
        why = "Good fit for a low-risk first audit; use the result to create proof and monitoring appetite."
        best_offer_hint = "Lead with Starter Leak Snapshot, then convert if outcomes or prevention signals are visible."
        pitch_angle = "Baseline audit and evidence split"
    elif score >= 35:
        band = "Educate first"
        why = "The category likely needs explanation before the prospect will understand why the audit matters."
        best_offer_hint = "Send education copy and ask for one export before discussing a larger package."
        pitch_angle = "Courier leakage education"
    else:
        band = "Low priority"
        why = "Spend, urgency, or data access is not strong enough for active selling yet."
        best_offer_hint = "Nurture, referral, or wait until courier spend/data access improves."
        pitch_angle = "Light nurture"

    first_data_ask = "Please send one courier export and invoice/statement for the same month. Add the signed rate card if available."
    moat_note = "The strongest moat is built when this prospect enters the monthly loop: audit, query, track outcomes, refine rules, and build benchmark history."
    return {
        "score": score,
        "band": band,
        "why": why,
        "risks": risks or ["No major qualification risks detected."],
        "best_offer_hint": best_offer_hint,
        "first_data_ask": first_data_ask,
        "pitch_angle": pitch_angle,
        "moat_note": moat_note,
    }


def build_prospect_outreach_from_score(prospect_name, company_name, qualification, sales_kit, currency_prefix):
    prospect_name = safe_text(prospect_name, "there") or "there"
    company_name = safe_text(company_name, "your brand") or "your brand"
    qualification = qualification or {}
    sales_kit = sales_kit or {}
    band = qualification.get("band", "Good pilot")
    pitch_angle = qualification.get("pitch_angle", "Baseline audit and evidence split")
    data_ask = qualification.get("first_data_ask", "Please send one courier export and invoice/statement for the same month.")
    education = sales_kit.get("category_explainer", {}).get("why_it_matters", "Courier leakage hides in accepted-vs-charged rates, billed-weight jumps, rate-card mismatches, and packaging decisions.")
    cta = sales_kit.get("offer_sheet", {}).get("cta", "Book a Courier Leak Audit.")
    return {
        "first_message": f"Hi {prospect_name}, I help e-commerce brands check whether courier bills are quietly leaking margin. For {company_name}, the angle I would test is: {pitch_angle}. {education} Worth checking one month of courier data?",
        "email_subject": f"Courier leakage check for {company_name}",
        "email_body": f"Hi {prospect_name},\n\nCourier leakage is hard to spot because it is spread across accepted-vs-charged rates, rate-card mismatches, billed-weight changes, fuel/remote-zone rules, and packaging choices.\n\nBased on the fit signals, I would treat {company_name} as: {band}. The first step is low-risk: {data_ask}\n\nI will return a plain-English split between dispute-ready/query candidates, operational exposure, and data gaps - without overclaiming refunds.\n\n{cta}\n\nKyle",
        "data_request": data_ask,
        "follow_up": f"Hi {prospect_name}, just following up. Even if there is no immediate recovery, the baseline shows whether {company_name} has courier billing control, packaging exposure, or data gaps before spend scales further.",
        "close_cta": cta,
    }


def build_prospect_qualification_pack_text(qualification, outreach):
    qualification = qualification or {}
    outreach = outreach or {}
    lines = [
        "Prospect Qualification & First Pitch Pack",
        "========================================",
        "",
        f"Score: {qualification.get('score', 0)}/100",
        f"Band: {qualification.get('band', 'n/a')}",
        f"Pitch angle: {qualification.get('pitch_angle', 'n/a')}",
        f"Why: {qualification.get('why', '')}",
        f"Best offer hint: {qualification.get('best_offer_hint', '')}",
        f"First data ask: {qualification.get('first_data_ask', '')}",
        "",
        "RISKS",
    ]
    lines.extend(f"- {risk}" for risk in qualification.get("risks", []))
    lines.extend(["", "MOAT NOTE", qualification.get("moat_note", ""), "", "FIRST MESSAGE", outreach.get("first_message", ""), "", "EMAIL", f"Subject: {outreach.get('email_subject', '')}", outreach.get("email_body", ""), "", "FOLLOW-UP", outreach.get("follow_up", ""), "", "CTA", outreach.get("close_cta", "")])
    return "\n".join(lines).strip() + "\n"


def determine_client_stage(file_name, cleaned_rate_card, findings_df, dispute_ready_findings, query_findings, skipped_rows, evidence_readiness, outcome_summary, rate_card_matched_count):
    findings_df = findings_df if findings_df is not None else pd.DataFrame()
    dispute_ready_findings = dispute_ready_findings if dispute_ready_findings is not None else pd.DataFrame()
    query_findings = query_findings if query_findings is not None else pd.DataFrame()
    skipped_rows = skipped_rows if skipped_rows is not None else pd.DataFrame()
    cleaned_rate_card = cleaned_rate_card if cleaned_rate_card is not None else pd.DataFrame()
    evidence_readiness = evidence_readiness or {}
    outcome_summary = outcome_summary or {}

    evidence_score = int(safe_float(evidence_readiness.get("score", 0)))
    credited_amount = safe_float(outcome_summary.get("credited_amount", 0))
    submitted_amount = safe_float(outcome_summary.get("submitted_amount", 0))
    rejection_count = int(safe_float(outcome_summary.get("rejection_count", 0)))
    missing_evidence = evidence_readiness.get("missing_evidence", []) or []
    blockers = []
    assets_to_send = []

    if not file_name:
        blockers.append("No client courier export uploaded yet; current results are based on mock/demo data.")
    if cleaned_rate_card.empty:
        blockers.append("No usable client rate card loaded; rate-card recoveries will be weaker.")
    if not findings_df.empty and dispute_ready_findings.empty and query_findings.empty:
        blockers.append("No dispute-ready or account-manager query rows yet; lead with operational/control findings.")
    if len(skipped_rows) > 0:
        blockers.append(f"{len(skipped_rows):,} rows were skipped or need cleanup before relying on them.")
    for item in missing_evidence[:4]:
        blockers.append(item)

    if credited_amount > 0 and evidence_score >= 80:
        stage = "Case Study Opportunity"
        stage_reason = "Credited outcomes are recorded and the evidence posture is strong enough to request permission for proof assets."
        next_action = "Ask for case-study/testimonial permission and propose monthly monitoring."
        assets_to_send = ["Case Study/Follow-Up Pack", "Monthly Monitoring Offer", "Outcome Tracker"]
        follow_up_timing = "Today"
        primary_angle = "Proof and retainer conversion"
    elif credited_amount > 0:
        stage = "Proof / Monitoring Opportunity"
        stage_reason = "A credited outcome is recorded; convert the result into proof and a monthly control conversation."
        next_action = "Send outcome summary, ask for permission to use anonymized proof, and propose monthly monitoring."
        assets_to_send = ["Case Study/Follow-Up Pack", "Outcome Tracker", "Monthly Monitoring Offer"]
        follow_up_timing = "Today"
        primary_angle = "Outcome-backed monitoring"
    elif submitted_amount > 0 or rejection_count > 0:
        stage = "Outcome Pending"
        stage_reason = "Rows have been submitted or responses started, but credited outcomes are not fully recorded yet."
        next_action = "Follow up for row-level courier responses: credited, rejected with reason, or needs more evidence."
        assets_to_send = ["Outcome Tracker", "Courier Follow-Up Email", "Credit Note Confirmation Request"]
        follow_up_timing = "3-7 days"
        primary_angle = "Close the evidence loop"
    elif not dispute_ready_findings.empty or not query_findings.empty:
        stage = "Client Pack Ready"
        stage_reason = "The audit has claim/query candidates ready for a cautious client conversation."
        next_action = "Send the client audit memo and evidence-scored dispute pack, then request courier row-level responses."
        assets_to_send = ["Client Audit Memo", "Evidence-Scored Dispute Pack", "Rate Card Request Email", "Outcome Tracker"]
        follow_up_timing = "3 days"
        primary_angle = "Cautious recovery and query workflow"
    elif findings_df.empty:
        stage = "Baseline / No Findings"
        stage_reason = "No findings were generated from this upload; use it as a baseline and request stronger evidence if needed."
        next_action = "Send a baseline note, request invoice/rate-card context, and offer monthly monitoring if spend justifies it."
        assets_to_send = ["Rate Card Request Email", "Client Intake Checklist", "Baseline Follow-Up"]
        follow_up_timing = "5 days"
        primary_angle = "Baseline trust builder"
    elif evidence_score < 40 or len(skipped_rows) > max(10, len(findings_df)):
        stage = "Evidence Cleanup Needed"
        stage_reason = "Evidence quality is too weak for strong recovery language."
        next_action = "Request cleaner exports, invoice/rate-card context, and missing fields before presenting major claims."
        assets_to_send = ["Rate Card Request Email", "Client Intake Checklist"]
        follow_up_timing = "2 days"
        primary_angle = "Data cleanup and control"
    else:
        stage = "Prospect / Demo Mode" if not file_name else "Audit Completed"
        stage_reason = "The audit is loaded; use the dashboard to decide whether to lead with recovery, prevention, or data cleanup."
        next_action = evidence_readiness.get("recommended_next_action", "Review evidence posture and prepare the client conversation.")
        assets_to_send = ["Client Conversation Guide", "Client Audit Memo"]
        follow_up_timing = "3 days"
        primary_angle = evidence_readiness.get("band", "Baseline trust builder")

    if rate_card_matched_count > 0 and "Rate Card Diagnostics" not in assets_to_send:
        assets_to_send.append("Rate Card Diagnostics")
    return {
        "stage": stage,
        "stage_reason": stage_reason,
        "next_action": next_action,
        "blockers": blockers or ["No major blockers detected."],
        "assets_to_send": assets_to_send,
        "follow_up_timing": follow_up_timing,
        "primary_angle": primary_angle,
        "evidence_score": evidence_score,
        "readiness_band": evidence_readiness.get("band", "Not scored"),
    }


def build_next_action_message(client_name, courier_provider, command_state, currency_prefix):
    client_name = safe_text(client_name, "the client") or "the client"
    stage = command_state.get("stage", "Audit Review")
    next_action = command_state.get("next_action", "Review the audit and prepare next steps.")
    assets = ", ".join(command_state.get("assets_to_send", []))
    blockers = "; ".join(command_state.get("blockers", []))
    email = f"Hi {{Name}},\n\nI have reviewed {client_name}'s courier data and the current stage is: {stage}.\n\nRecommended next step: {next_action}\n\nAssets to use now: {assets}.\n\nImportant context: {command_state.get('stage_reason', '')}\n\nIf we send this to {courier_provider}, the goal is to get row-level responses: credited, rejected with reason, or needs more evidence.\n\nKyle"
    whatsapp = f"Hi {{Name}}, quick audit update for {client_name}: {stage}. Next step: {next_action} Assets to use: {assets}. Follow up in {command_state.get('follow_up_timing', '3 days')}."
    internal_note = f"Stage: {stage}\nReason: {command_state.get('stage_reason', '')}\nBlockers: {blockers}\nPrimary angle: {command_state.get('primary_angle', '')}\nFollow-up timing: {command_state.get('follow_up_timing', '')}"
    return {
        "message_title": f"{stage}: next action",
        "email": email,
        "whatsapp": whatsapp,
        "internal_note": internal_note,
    }


def build_command_center_pack_text(command_state, message_pack):
    command_state = command_state or {}
    message_pack = message_pack or {}
    lines = [
        "Client Command Center - Next Action Pack",
        "========================================",
        "",
        f"Stage: {command_state.get('stage', 'n/a')}",
        f"Evidence score: {command_state.get('evidence_score', 0)}/100",
        f"Readiness band: {command_state.get('readiness_band', 'n/a')}",
        f"Primary angle: {command_state.get('primary_angle', 'n/a')}",
        f"Follow-up timing: {command_state.get('follow_up_timing', 'n/a')}",
        "",
        "WHY THIS STAGE",
        command_state.get("stage_reason", ""),
        "",
        "NEXT ACTION",
        command_state.get("next_action", ""),
        "",
        "BLOCKERS",
    ]
    lines.extend(f"- {item}" for item in command_state.get("blockers", []))
    lines.extend(["", "ASSETS TO SEND"])
    lines.extend(f"- {item}" for item in command_state.get("assets_to_send", []))
    lines.extend(["", "EMAIL", message_pack.get("email", ""), "", "WHATSAPP", message_pack.get("whatsapp", ""), "", "INTERNAL NOTE", message_pack.get("internal_note", "")])
    return "\n".join(lines).strip() + "\n"


def build_delivery_pack(command_state, evidence_readiness, sales_kit=None):
    command_state = command_state or {}
    evidence_readiness = evidence_readiness or {}
    stage = command_state.get("stage", "Audit Review")
    missing = evidence_readiness.get("missing_evidence", []) or []
    base_warnings = ["Do not describe operational exposure as guaranteed courier recovery.", "Keep all amounts framed as candidates until courier response or credit note confirms them."]
    mapping = {
        "Prospect / Demo Mode": {
            "pack_name": "Prospect Education Pack",
            "objective": "Educate the prospect on courier leakage and earn the first data upload.",
            "attach_or_generate": ["Outreach Pack", "Client Intake Checklist", "What-I-Do Explainer", "Data Request Message"],
            "ask_for": ["One courier export", "Matching invoice/statement", "Current rate card if available"],
            "next_stage": "Data Requested",
        },
        "Evidence Cleanup Needed": {
            "pack_name": "Evidence Cleanup Pack",
            "objective": "Request missing evidence before making strong recovery claims.",
            "attach_or_generate": ["Missing Evidence Checklist", "Rate Card Request Email", "Cleaner Export Request", "Data Quality Explanation"],
            "ask_for": missing or ["Cleaner export", "Signed rate card", "Surcharge schedule"],
            "next_stage": "Data Received",
        },
        "Baseline / No Findings": {
            "pack_name": "Baseline Control Pack",
            "objective": "Position this as a baseline and set up the next monthly control check.",
            "attach_or_generate": ["Baseline Note", "Rate Card Request Email", "Next Month Checklist", "Monitoring Explanation"],
            "ask_for": ["Latest rate card", "Next month export", "Courier invoice/statement"],
            "next_stage": "Monthly Monitoring Proposed",
        },
        "Client Pack Ready": {
            "pack_name": "Client Audit Pack",
            "objective": "Send a cautious client-ready audit pack and start courier/account-manager response tracking.",
            "attach_or_generate": ["Client Audit Memo", "Evidence-Scored Dispute Pack", "Rate Card Request Email", "Outcome Tracker"],
            "ask_for": ["Row-level courier responses", "Signed rate card", "Surcharge schedule", "Credit-note confirmation"],
            "next_stage": "Dispute Pack Sent",
        },
        "Outcome Pending": {
            "pack_name": "Courier Follow-Up Pack",
            "objective": "Get every submitted row marked as credited, rejected with reason, or needs more evidence.",
            "attach_or_generate": ["Courier Follow-Up Email", "Outcome Tracker", "Credit-Note Confirmation Request", "Rejection Reason Request"],
            "ask_for": ["Courier row-level response", "Credit note", "Rejected reasons", "Evidence requests"],
            "next_stage": "Credit Received or Proof Opportunity",
        },
        "Proof / Monitoring Opportunity": {
            "pack_name": "Monitoring Conversion Pack",
            "objective": "Turn outcomes into proof and convert the client into monthly monitoring.",
            "attach_or_generate": ["Case Study/Follow-Up Pack", "Monthly Monitoring Pack", "Snapshot Trend Narrative", "Monitoring Proposal"],
            "ask_for": ["Permission for anonymized proof", "Next month export", "Updated rate card/surcharge schedule"],
            "next_stage": "Monthly Monitoring Proposed",
        },
        "Case Study Opportunity": {
            "pack_name": "Proof & Case Study Pack",
            "objective": "Capture permissioned proof and use it to support retention/referrals.",
            "attach_or_generate": ["Named/Anonymized Case Study", "Permission Request", "LinkedIn Proof Post", "Referral Partner Proof", "Monthly Monitoring Offer"],
            "ask_for": ["Case-study permission", "Testimonial permission", "Referral introduction", "Next monitoring cycle approval"],
            "next_stage": "Converted or Referral Loop",
        },
    }
    pack = mapping.get(stage, {
        "pack_name": "Audit Review Pack",
        "objective": "Review the current audit state and choose the strongest next action.",
        "attach_or_generate": command_state.get("assets_to_send", ["Client Conversation Guide", "Client Audit Memo"]),
        "ask_for": missing or ["Confirm next data/evidence requirement"],
        "next_stage": "Next Action Completed",
    }).copy()
    pack["stage"] = stage
    pack["warnings"] = base_warnings + (["Evidence gaps remain: " + "; ".join(missing[:3])] if missing else [])
    pack["follow_up_timing"] = command_state.get("follow_up_timing", "3 days")
    pack["delivery_checklist"] = [
        "Confirm the stage and objective before sending.",
        "Generate or attach the listed assets.",
        "Paste the recommended message and customize the greeting.",
        "Record the send date and next follow-up date.",
        "Update outcome/snapshot history after the client or courier responds.",
    ]
    return pack


def build_delivery_pack_text(delivery_pack, command_message_pack):
    delivery_pack = delivery_pack or {}
    command_message_pack = command_message_pack or {}
    lines = [
        "Client Delivery Pack Instructions",
        "=================================",
        "",
        f"Stage: {delivery_pack.get('stage', 'n/a')}",
        f"Pack: {delivery_pack.get('pack_name', 'n/a')}",
        f"Objective: {delivery_pack.get('objective', '')}",
        f"Follow-up timing: {delivery_pack.get('follow_up_timing', 'n/a')}",
        f"Next stage: {delivery_pack.get('next_stage', 'n/a')}",
        "",
        "ATTACH OR GENERATE",
    ]
    lines.extend(f"[ ] {item}" for item in delivery_pack.get("attach_or_generate", []))
    lines.extend(["", "ASK FOR"])
    lines.extend(f"[ ] {item}" for item in delivery_pack.get("ask_for", []))
    lines.extend(["", "WARNINGS"])
    lines.extend(f"- {item}" for item in delivery_pack.get("warnings", []))
    lines.extend(["", "DELIVERY CHECKLIST"])
    lines.extend(f"[ ] {item}" for item in delivery_pack.get("delivery_checklist", []))
    lines.extend(["", "EMAIL MESSAGE", command_message_pack.get("email", ""), "", "WHATSAPP MESSAGE", command_message_pack.get("whatsapp", ""), "", "INTERNAL NOTE", command_message_pack.get("internal_note", "")])
    return "\n".join(lines).strip() + "\n"


def build_recovery_provenance_summary(analysis_data, findings_df, currency_prefix):
    """Explain where direct recovery comes from and what is deliberately excluded."""
    analysis_data = analysis_data if analysis_data is not None else pd.DataFrame()
    findings_df = findings_df if findings_df is not None else pd.DataFrame()
    direct_findings = findings_df[findings_df["Recovery_Type"].eq(RECOVERY_DIRECT)] if not findings_df.empty and "Recovery_Type" in findings_df.columns else pd.DataFrame()
    operational_findings = findings_df[findings_df["Recovery_Type"].eq(RECOVERY_OPERATIONAL)] if not findings_df.empty and "Recovery_Type" in findings_df.columns else pd.DataFrame()
    tier_b = findings_df[findings_df["Evidence_Tier"].eq(EVIDENCE_TIERS["B"])] if not findings_df.empty and "Evidence_Tier" in findings_df.columns else pd.DataFrame()
    query_findings = findings_df[findings_df["Recovery_Type"].eq(RECOVERY_QUERY)] if not findings_df.empty and "Recovery_Type" in findings_df.columns else pd.DataFrame()
    leakage_totals = calculate_leakage_totals(analysis_data, findings_df, pd.DataFrame(), pd.DataFrame())
    direct_total = leakage_totals["direct_recovery_zar"]
    operational_total = leakage_totals["operational_exposure_zar"]
    tier_b_total = safe_float(tier_b.get("Direct_Recovery_ZAR", pd.Series(dtype=float)).sum()) if not tier_b.empty else 0.0
    query_total = safe_float(query_findings.get("Direct_Recovery_ZAR", pd.Series(dtype=float)).sum()) if not query_findings.empty else 0.0
    direct_with_money_fields = 0
    if not direct_findings.empty:
        accepted = pd.to_numeric(direct_findings.get("Accepted_Rate_ZAR", pd.Series(0, index=direct_findings.index)), errors="coerce").fillna(0)
        charged = pd.to_numeric(direct_findings.get("Charged_Rate_ZAR", pd.Series(0, index=direct_findings.index)), errors="coerce").fillna(0)
        direct_with_money_fields = int((accepted.gt(0) & charged.gt(0)).sum())
    return {
        "direct_total": direct_total,
        "operational_total": operational_total,
        "tier_b_total": tier_b_total,
        "query_total": query_total,
        "direct_rows": len(direct_findings),
        "operational_rows": len(operational_findings),
        "tier_b_rows": len(tier_b),
        "query_rows": len(query_findings),
        "direct_with_money_fields": direct_with_money_fields,
        "verdict": f"Direct recovery candidates total {format_currency(direct_total, currency_prefix)} and are based on accepted-vs-charged money deltas. Packaging/weight exposure of {format_currency(operational_total, currency_prefix)} is separated as operational exposure and is not included in the direct recovery total. If direct recovery is zero but operational exposure is positive, the upload is not necessarily 100% accurate; it means the current evidence supports prevention/control work rather than an immediate courier credit claim.",
        "formula": "Direct recovery candidate = max(Charged_Rate_ZAR - Accepted_Rate_ZAR, 0).",
    }


def build_direct_recovery_sanity_checks(analysis_data, findings_df):
    analysis_data = analysis_data if analysis_data is not None else pd.DataFrame()
    findings_df = findings_df if findings_df is not None else pd.DataFrame()
    direct_findings = findings_df[findings_df["Recovery_Type"].eq(RECOVERY_DIRECT)] if not findings_df.empty and "Recovery_Type" in findings_df.columns else pd.DataFrame()
    has_accepted = "Accepted_Rate_ZAR" in analysis_data.columns and pd.to_numeric(analysis_data["Accepted_Rate_ZAR"], errors="coerce").fillna(0).gt(0).any()
    has_charged = "Charged_Rate_ZAR" in analysis_data.columns and pd.to_numeric(analysis_data["Charged_Rate_ZAR"], errors="coerce").fillna(0).gt(0).any()
    operational_separate = not findings_df.empty and "Operational_Exposure_ZAR" in findings_df.columns
    tier_b_separate = direct_findings.empty or not direct_findings.get("Evidence_Tier", pd.Series(dtype=str)).eq(EVIDENCE_TIERS["B"]).any()
    return pd.DataFrame([
        {"Check": "Accepted rate field present", "Status": "Pass" if has_accepted else "Review", "Explanation": "Direct recovery needs accepted/quoted rate evidence."},
        {"Check": "Charged rate field present", "Status": "Pass" if has_charged else "Review", "Explanation": "Direct recovery needs charged/courier-confirmed rate evidence."},
        {"Check": "Direct delta formula", "Status": "Pass", "Explanation": "Direct recovery uses max(Charged_Rate_ZAR - Accepted_Rate_ZAR, 0)."},
        {"Check": "Not packaging/weight based", "Status": "Pass", "Explanation": "Packaging dimensions and weight differences are not included in direct recovery; they are operational exposure unless rate-card/courier evidence confirms a money delta."},
        {"Check": "Operational exposure separated", "Status": "Pass" if operational_separate else "Review", "Explanation": "Weight/dimension exposure is stored separately as Operational_Exposure_ZAR."},
        {"Check": "Tier B separated", "Status": "Pass" if tier_b_separate else "Review", "Explanation": "Rate-card reproducible Tier B claims are separate from Tier A accepted-vs-charged deltas."},
    ])


def calculate_evidence_readiness(analysis_data, findings_df, skipped_rows, cleaned_rate_card, outcome_summary, rate_card_matched_count):
    """Score how safe this audit is to present, dispute, and turn into proof."""
    analysis_data = analysis_data if analysis_data is not None else pd.DataFrame()
    findings_df = findings_df if findings_df is not None else pd.DataFrame()
    skipped_rows = skipped_rows if skipped_rows is not None else pd.DataFrame()
    cleaned_rate_card = cleaned_rate_card if cleaned_rate_card is not None else pd.DataFrame()
    outcome_summary = outcome_summary or {}

    total_rows = len(analysis_data) + len(skipped_rows)
    skipped_rate = len(skipped_rows) / total_rows if total_rows else 0
    avg_confidence = safe_float(findings_df.get("Confidence_Score", pd.Series(dtype=float)).mean()) if not findings_df.empty else 0
    direct_recovery = safe_float(findings_df.get("Direct_Recovery_ZAR", pd.Series(dtype=float)).sum()) if not findings_df.empty else 0
    credited_amount = safe_float(outcome_summary.get("credited_amount", 0))
    submitted_amount = safe_float(outcome_summary.get("submitted_amount", 0))

    has_waybills = "Waybill_ID" in analysis_data.columns and analysis_data["Waybill_ID"].astype(str).str.strip().ne("").any()
    has_accepted_rate = "Accepted_Rate_ZAR" in analysis_data.columns and pd.to_numeric(analysis_data["Accepted_Rate_ZAR"], errors="coerce").fillna(0).gt(0).any()
    has_charged_rate = "Charged_Rate_ZAR" in analysis_data.columns and pd.to_numeric(analysis_data["Charged_Rate_ZAR"], errors="coerce").fillna(0).gt(0).any()
    has_rate_card = cleaned_rate_card is not None and not cleaned_rate_card.empty

    score = 0
    missing = []
    if has_waybills:
        score += 12
    else:
        missing.append("Waybill IDs for courier row-level dispute matching")
    if has_accepted_rate:
        score += 12
    else:
        missing.append("Accepted/quoted rate field")
    if has_charged_rate:
        score += 12
    else:
        missing.append("Charged/courier-confirmed rate field")
    if has_rate_card or safe_float(rate_card_matched_count) > 0:
        score += 12
    else:
        missing.append("Signed courier rate card and surcharge schedule")
    if skipped_rate <= 0.02:
        score += 10
    elif skipped_rate <= 0.08:
        score += 6
    else:
        missing.append("Cleaner export with fewer skipped/dirty rows")
    if direct_recovery > 0:
        score += 12
    else:
        missing.append("Evidence-backed direct recovery candidates")
    score += min(12, int(avg_confidence / 100 * 12))
    if submitted_amount > 0:
        score += 8
    else:
        missing.append("Recorded courier submission outcomes")
    if credited_amount > 0:
        score += 10
    else:
        missing.append("Confirmed credit notes / credited outcomes")

    if score >= 80:
        band = "Case-study / retainer ready"
        posture = "Strong evidence posture. Use confirmed outcomes and client approval to support case studies and monthly monitoring."
        next_action = "Ask for testimonial/anonymized case-study permission and propose monthly monitoring."
    elif score >= 60:
        band = "Dispute pack ready"
        posture = "Strong enough to submit a cautious dispute pack and ask for row-level courier responses."
        next_action = "Send the dispute pack, track accepted/rejected/credited rows, and request missing rate-card evidence."
    elif score >= 40:
        band = "Client conversation ready"
        posture = "Good for a client conversation, but avoid aggressive recovery claims until missing evidence is supplied."
        next_action = "Lead with cautious recovery candidates and ask for rate card, surcharge rules, and cleaner exports."
    else:
        band = "Baseline only"
        posture = "Use this as a baseline/control conversation. Do not lead with strong refund language."
        next_action = "Request missing evidence before presenting major recovery claims."

    return {
        "score": min(100, int(score)),
        "band": band,
        "posture": posture,
        "missing_evidence": missing,
        "recommended_next_action": next_action,
        "case_study_ready": score >= 80 and credited_amount > 0,
    }


def build_rate_card_request_pack(readiness):
    missing = readiness.get("missing_evidence", []) if isinstance(readiness, dict) else []
    email = """Hi {Name},

To complete the courier leakage audit safely, please send the evidence below:
- signed courier rate card
- fuel and surcharge schedule
- remote/outlying area surcharge rules
- volumetric divisor and VAT inclusion/exclusion
- effective dates for the rate card
- courier account manager contact
- latest invoice/export for the period under review

This helps us separate recoverable claim candidates from operational exposure and avoids overclaiming.

Thanks,
Kyle"""
    whatsapp = "Hi {Name}, to finish the courier audit safely, please send the signed rate card, surcharge schedule, divisor/VAT rules, latest invoice/export, and courier account manager contact. This lets us separate real recovery candidates from operational exposure."
    return {"email": email, "whatsapp": whatsapp, "missing_evidence": missing}


def build_business_swot(readiness, outcome_summary, pipeline_summary, sales_kit, offer):
    return {
        "Strengths": [
            "Defensible evidence-tier approach separates recovery candidates from operational noise.",
            "Clear niche wedge: courier leakage and shipping-margin control for SME e-commerce brands.",
            "Client-ready memos, dispute packs, sales assets, and guided talk tracks already exist.",
            "Outcome tracker and client pipeline create a loop from audit to proof to retainer.",
        ],
        "Weaknesses": [
            "Early-stage credibility is still limited until real credited outcomes and case studies are recorded.",
            "Messy client exports can reduce evidence quality and make onboarding feel manual.",
            "Signed rate-card access is needed for stronger Tier B claims.",
            "Service delivery is still founder-led and may be time-intensive.",
            "The market needs category education: many SMEs do not know to ask for courier leak audits yet.",
        ],
        "Opportunities": [
            "Referral partners: Shopify agencies, accountants, outsourced CFOs, 3PLs, and warehouse consultants.",
            "Monthly monitoring retainers after the first audit proves the workflow.",
            "Anonymized recovery benchmarks by courier, industry, spend band, and evidence tier.",
            "Outcome-based case studies once credit notes are confirmed and clients approve.",
        ],
        "Threats": [
            "Couriers may explain or reject apparent anomalies through valid surcharges or account rules.",
            "Overclaiming could damage trust; cautious evidence language must remain a core differentiator.",
            "Generic consultants can copy surface-level reports, but not outcome history and evidence discipline.",
            "Buyer budget pressure in the current economy means pricing must stay low-friction until proof improves.",
            "Data privacy concerns can slow adoption unless the process is professional and minimal.",
        ],
        "Address Next": [
            "Get 3 pilot clients and record every accepted/rejected/credited outcome.",
            "Ask for testimonial or anonymized case-study permission whenever discounting.",
            "Use the rate-card request pack before making strong Tier B claims.",
            "Build a referral partner list and send the partner pitch from the Sales Kit.",
            "Keep pricing introduction-stage until credited outcomes justify higher retainers.",
            f"Current evidence posture: {readiness.get('band', 'Baseline only')} — {readiness.get('recommended_next_action', 'Request more evidence.')}",
        ],
    }


def build_sales_kit(client_name, courier_provider, direct_recovery, operational_exposure, packaging_leak, financial_query, currency_prefix, offer, outcome_summary=None):
    """Build market-facing copy for the Courier Leak Audit offer."""
    client_name = safe_text(client_name, "your brand") or "your brand"
    direct_recovery = safe_float(direct_recovery)
    operational_exposure = safe_float(operational_exposure)
    packaging_leak = safe_float(packaging_leak)
    financial_query = safe_float(financial_query)
    total_signal = direct_recovery + operational_exposure + packaging_leak + financial_query
    proof_line = (
        f"This audit isolated {format_currency(direct_recovery, currency_prefix)} in direct recovery candidates and "
        f"{format_currency(operational_exposure, currency_prefix)} in operational exposure."
        if total_signal > 0 else
        "This audit creates a clean courier-control baseline before leakage becomes expensive."
    )
    headline = "Find out if your courier bill is leaking profit."
    subheadline = "A Courier Leak Audit separates recoverable claim candidates from operational leakage, so e-commerce brands know what to query, what to fix, and what not to overclaim."
    cta = offer.get("cta", "Book a Courier Leak Audit.") if isinstance(offer, dict) else "Book a Courier Leak Audit."
    package_name = offer.get("package_name", "Starter Leak Snapshot") if isinstance(offer, dict) else "Starter Leak Snapshot"
    audit_fee = offer.get("recommended_audit_fee", 4500) if isinstance(offer, dict) else 4500
    retainer = offer.get("recommended_retainer", 3500) if isinstance(offer, dict) else 3500
    outcome_line = "Outcome tracking turns each courier response into better future benchmarks."
    if outcome_summary and safe_float(outcome_summary.get("credited_amount", 0)) > 0:
        outcome_line = f"Recorded credited outcomes: {format_currency(outcome_summary.get('credited_amount', 0), currency_prefix)}. Publish only with client approval."

    category_explainer = {
        "what_i_do": "I audit courier billing for e-commerce brands. I compare courier exports, invoices, accepted rates, rate cards, billed weights, volumetric rules, and packaging signals to show which rows are query-ready and which issues are operational leakage.",
        "why_it_matters": "Courier leakage is usually hidden across small line-item differences: accepted vs charged rates, rate-card mismatches, remote-zone charges, billed-weight jumps, and packaging choices. One row rarely looks dramatic, but the pattern can quietly erode monthly margin.",
        "why_now": "As order volume grows, courier complexity grows with it. A monthly control process catches issues before they become normalised, and it gives the courier/account manager clean evidence instead of vague complaints.",
        "why_me": "The moat is the evidence workflow: messy export ingestion, rate-card reconstruction, confidence scoring, readable dispute packs, outcome tracking, and a benchmark history that gets smarter each month.",
    }
    landing_page_link = "[Henderson Performance](https://stunning-sunflower-7ee199.netlify.app/)"
    cold_outreach_sequence = {
        "linkedin_connection_note": "Hi {Name}, I help e-commerce brands check whether courier bills are quietly leaking margin. Thought it would be useful to connect.",
        "linkedin_dm_1": f"Hi {{Name}}, most brands know their courier spend, but not which waybills are actually causing the problem. I run a 3 Month Courier Audit that checks overcharges, billed-weight jumps, rate differences and packaging leaks. The details are here: {landing_page_link}. Worth checking {client_name}'s courier data?",
        "email_1_subject": "Courier leakage check for your e-commerce ops",
        "email_1_body": f"Hi {{Name}},\n\nCourier billing leakage is hard to see because it hides in accepted-vs-charged rates, rate-card mismatches, billed-weight jumps, fuel/remote surcharges and packaging decisions.\n\nI run a 3 Month Courier Audit for South African businesses. You send three months of courier data and I return a plain-English findings note showing what to query, what to fix and what evidence is missing.\n\nThe audit details are here: {landing_page_link}\n\n{proof_line}\n\nWorth checking {client_name}'s courier data?\n\nKyle",
        "whatsapp_1": f"Hi {{Name}}, quick one: I help businesses check courier bills for overcharges, billed-weight jumps and packaging leaks. I have put the audit details here: {landing_page_link}. The 3 month pilot audit is R850. Worth checking {client_name}'s courier data?",
        "follow_up_2_day": f"Just bumping this. The useful part is not a generic dashboard. It is a clean split between courier rows to query, operational leaks to fix and evidence gaps to close. Details here: {landing_page_link}",
        "follow_up_5_day": "If helpful, we can start with three months of courier exports, invoice/statement context and the rate card if available. If there is no strong recovery candidate, you still get a control checklist and know what evidence is missing.",
        "breakup_10_day": "I will leave this for now. If courier costs spike, billed weights look odd, or your team wants a second pair of eyes on courier leakage, I can run the R850 3 month audit.",
    }
    persona_scripts = {
        "founder_owner": "For founders: this is margin protection. You do not need another dashboard; you need to know which courier rows can be challenged, which costs are caused by packaging/process, and where monthly control will protect cash.",
        "ops_manager": "For ops: this gives you a practical exception list - billed-weight jumps, service/rate mismatches, packaging exposure, and data-quality gaps - so you can fix the pack station and have a cleaner courier conversation.",
        "finance_manager": "For finance: this creates an evidence trail. It separates potential credit candidates from operational estimates, making courier accruals, credit-note follow-up, and monthly variance checks easier to govern.",
        "agency_or_accountant_referral_partner": "For referral partners: your e-commerce clients may be leaking courier margin without a specialist reviewing rate cards, billed weights, and disputes. You can introduce a low-risk audit and give them a tangible control asset.",
    }
    objection_responses = {
        "Is this just courier disputes?": "No. Disputes are only one output. The bigger value is the control system: rate-card checks, operational leakage separation, packaging signals, readable evidence, and outcome tracking.",
        "We already use an aggregator": "Aggregators help with shipping workflow, but the brand still needs to know whether accepted rates, charged rates, courier confirmations, billed weights, and packaging rules are creating margin leakage.",
        "Will this annoy my courier?": "The opposite if done well. The pack is evidence-scored and cautious, so you send clear waybill-level queries instead of broad accusations.",
        "What data do you need?": "Start with one courier export and invoice/statement for the same month. The signed rate card and surcharge rules make the audit stronger, but the baseline can start small.",
        "What if there is no recovery?": "Then you still get a baseline, data-quality read, packaging/rate-card control checklist, and a benchmark for monthly monitoring. No-recovery months are still useful if they prove the control process is working.",
    }
    data_request_sequence = {
        "short_first_ask": "To start, please send one courier export and invoice/statement for the same month. CSV/XLSX is ideal, but I can work with messy exports.",
        "full_evidence_ask": "For the strongest audit, please also send the signed/current rate card, surcharge schedule, volumetric divisor, VAT/fuel rules, and any previous courier credit notes or disputes.",
        "rate_card_ask": "If the rate card is split across main/local/regional or multiple couriers, send all files. The tool reconstructs expected rates more accurately when the full rate-card context is loaded.",
    }
    moat_positioning = [
        "Messy courier export ingestion across CSV, Excel, PDF, and HTML.",
        "Rate-card reconstruction across courier, service, weight band, and route bucket.",
        "Evidence scoring that separates dispute-ready claims from account-manager queries and operational exposure.",
        "Client-readable dispute packs and memos that avoid overclaiming refunds.",
        "Outcome tracking for credited/rejected rows, creating a benchmark history competitors cannot copy quickly.",
        "Monthly monitoring loop: audit, submit/query, track outcomes, refine rules, prevent repeat leakage.",
    ]

    return {
        "positioning": "We help South African and Botswana e-commerce brands find and prevent courier margin leaks without overclaiming refunds.",
        "headline": headline,
        "subheadline": subheadline,
        "proof_line": proof_line,
        "category_explainer": category_explainer,
        "cold_outreach_sequence": cold_outreach_sequence,
        "persona_scripts": persona_scripts,
        "objection_responses": objection_responses,
        "data_request_sequence": data_request_sequence,
        "moat_positioning": moat_positioning,
        "pain_bullets": [
            "Courier invoices are hard to reconcile against accepted rates, rate cards, and surcharge rules.",
            "Packaging and charged-weight mistakes quietly erode margin.",
            "Most SMEs do not know which rows are dispute-ready and which are only operational exposure.",
        ],
        "deliverables": [
            "Evidence-scored dispute pack",
            "Client audit memo",
            "Operational fix list",
            "Outcome tracker for credited/rejected courier responses",
        ],
        "linkedin_dm": f"Hi {{Name}}, quick question: do you currently check courier invoices against accepted rates, rate cards, billed weights and packaging exposure? I run a 3 Month Courier Audit for South African businesses. Details here: {landing_page_link}. Worth checking {client_name}'s courier data?",
        "email_subjects": ["Courier bill leakage check", f"Quick courier margin audit for {client_name}", "Possible courier overcharges in your exports"],
        "email_body": f"Hi {{Name}},\n\nMost businesses know their courier spend, but not which waybills, rate gaps or packaging decisions created the leak.\n\n{proof_line}\n\nThe 3 Month Courier Audit checks overcharges, billed-weight jumps, rate differences, packaging leaks and missing evidence. The pilot price is R850. Details here: {landing_page_link}\n\nWould it be useful to check {client_name}'s courier data?\n\nKyle",
        "whatsapp": f"Hi {{Name}}, I’m running a 3 Month Courier Audit for South African businesses. It checks overcharges, billed-weight jumps and packaging leaks. Details here: {landing_page_link}. Worth checking {client_name}'s courier data?",
        "discovery_script": [
            "Which courier or aggregator do you use most?",
            "What is your approximate monthly courier spend?",
            "Do you have the signed rate card and surcharge schedule?",
            "Who currently follows up on courier disputes or credit notes?",
            "Are packaging choices rule-based or manual at dispatch?",
            "Have you ever measured credited vs rejected courier disputes?",
        ],
        "referral_pitch": "Your e-commerce clients may be leaking courier margin without knowing which rows are recoverable. We can run a low-risk Courier Leak Audit, give them a clean action pack, and share a referral or co-branded report where appropriate.",
        "post_audit_conversion": f"The first audit gives us the evidence baseline. Monthly monitoring turns that into a control loop: check new exports, submit/query exceptions, track credits and rejections, and prevent the same leakage from repeating. For {client_name}, the logical next step is {package_name} with optional monitoring at {format_currency_whole(retainer, currency_prefix)}/mo.",
        "offer_sheet": {
            "package": package_name,
            "audit_fee": format_currency_whole(audit_fee, currency_prefix),
            "monitoring": f"{format_currency_whole(retainer, currency_prefix)}/mo optional monitoring",
            "timeline": "48 hours after receiving a clean courier export and rate-card context.",
            "cta": cta,
            "outcome_line": outcome_line,
        },
    }


def build_outreach_pack_text(sales_kit):
    sales_kit = sales_kit or {}
    category = sales_kit.get("category_explainer", {})
    sequence = sales_kit.get("cold_outreach_sequence", {})
    personas = sales_kit.get("persona_scripts", {})
    objections = sales_kit.get("objection_responses", {})
    data_request = sales_kit.get("data_request_sequence", {})
    moat = sales_kit.get("moat_positioning", [])
    lines = [
        "Courier Leak Audit Outreach Pack",
        "=================================",
        "",
        "WHAT I DO",
        category.get("what_i_do", ""),
        "",
        "WHY IT MATTERS",
        category.get("why_it_matters", ""),
        "",
        "WHY NOW",
        category.get("why_now", ""),
        "",
        "WHY ME / MOAT",
        category.get("why_me", ""),
        "",
        "COLD OUTREACH SEQUENCE",
    ]
    for label, text in sequence.items():
        lines.extend([f"\n{humanize_label(label)}", str(text)])
    lines.extend(["", "PERSONA-SPECIFIC SCRIPTS"])
    for label, text in personas.items():
        lines.extend([f"\n{humanize_label(label)}", str(text)])
    lines.extend(["", "OBJECTION HANDLING"])
    for question, answer in objections.items():
        lines.extend([f"\nQ: {question}", f"A: {answer}"])
    lines.extend(["", "DATA REQUEST SEQUENCE"])
    for label, text in data_request.items():
        lines.extend([f"\n{humanize_label(label)}", str(text)])
    lines.extend(["", "MOAT POSITIONING"])
    lines.extend(f"- {item}" for item in moat)
    lines.extend([
        "",
        "REFERRAL PARTNER PITCH",
        sales_kit.get("referral_pitch", ""),
        "",
        "POST-AUDIT CONVERSION",
        sales_kit.get("post_audit_conversion", ""),
        "",
        "OFFER CTA",
        sales_kit.get("offer_sheet", {}).get("cta", ""),
    ])
    return "\n".join(lines).strip() + "\n"


def generate_offer_sheet_pdf(client_name, sales_kit, currency_prefix):
    """Generate a one-page-ish offer sheet for the Courier Leak Audit."""
    pdf = MarginPDF()
    pdf.alias_nb_pages()
    pdf.set_margins(15, 24, 15)
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()

    def clean(text):
        return str(text).replace("–", "-").replace("—", "-").replace("‘", "'").replace("’", "'").replace("“", '"').replace("”", '"')

    def h2(text):
        pdf.set_font("Helvetica", "B", 12)
        pdf.set_text_color(*MarginPDF.NAVY)
        pdf.multi_cell(0, 7, clean(text))
        pdf.ln(1)

    def body(text):
        pdf.set_font("Helvetica", "", 10)
        pdf.set_text_color(*MarginPDF.TEXT)
        pdf.multi_cell(0, 6, clean(text))
        pdf.ln(1)

    pdf.set_font("Helvetica", "B", 18)
    pdf.set_text_color(*MarginPDF.NAVY)
    pdf.multi_cell(0, 9, "Courier Leak Audit Offer", align="C")
    pdf.ln(2)
    body(f"Prepared for: {safe_text(client_name, 'Prospective Client')}")
    h2(sales_kit["headline"])
    body(sales_kit["subheadline"])
    body(sales_kit["proof_line"])
    if sales_kit.get("category_explainer"):
        h2("Why this exists")
        body(sales_kit["category_explainer"].get("why_it_matters", ""))
    h2("What the audit checks")
    for item in sales_kit["pain_bullets"]:
        body(f"- {item}")
    h2("What you receive")
    for item in sales_kit["deliverables"]:
        body(f"- {item}")
    sheet = sales_kit["offer_sheet"]
    h2("Commercial option")
    body(f"Package: {sheet['package']}")
    body(f"Audit fee: {sheet['audit_fee']}")
    body(f"Monitoring: {sheet['monitoring']}")
    body(f"Timeline: {sheet['timeline']}")
    body(sheet["outcome_line"])
    h2("Next step")
    body(sheet["cta"])
    body("Note: recoveries are candidates until the courier validates waybills, rate-card context, surcharges, and credit notes.")
    return bytes(pdf.output())


def generate_client_training_brief_pdf(
    client_name,
    courier_provider,
    analysis_data,
    findings_df,
    benchmark_table,
    priority_actions,
    anomaly_rows,
    packaging_rows,
    financial_anomaly_rows,
    capital_trap_skus,
    currency_prefix,
    audit_month,
    audit_year,
):
    """Generate an internal consultant playbook PDF for client call preparation."""
    analysis_data = analysis_data.copy() if analysis_data is not None else pd.DataFrame()
    findings_df = findings_df.copy() if findings_df is not None else pd.DataFrame()
    benchmark_table = benchmark_table.copy() if benchmark_table is not None else pd.DataFrame()
    priority_actions = priority_actions.copy() if priority_actions is not None else pd.DataFrame()
    anomaly_rows = anomaly_rows.copy() if anomaly_rows is not None else pd.DataFrame()
    packaging_rows = packaging_rows.copy() if packaging_rows is not None else pd.DataFrame()
    financial_anomaly_rows = financial_anomaly_rows.copy() if financial_anomaly_rows is not None else pd.DataFrame()
    capital_trap_skus = capital_trap_skus.copy() if capital_trap_skus is not None else pd.DataFrame()

    total_spend = safe_float(analysis_data.get("Billed_Cost_ZAR", pd.Series(dtype=float)).sum()) if not analysis_data.empty else 0.0
    direct_recovery = safe_float(findings_df.get("Direct_Recovery_ZAR", pd.Series(dtype=float)).sum()) if not findings_df.empty else safe_float(anomaly_rows.get("Recoverable_Overcharge_ZAR", pd.Series(dtype=float)).sum())
    operational_exposure = safe_float(findings_df.get("Operational_Exposure_ZAR", pd.Series(dtype=float)).sum()) if not findings_df.empty else safe_float(analysis_data.get("Operational_Exposure_ZAR", pd.Series(dtype=float)).sum())
    packaging_leak = safe_float(packaging_rows.get("Avoidable_Volumetric_Leak_ZAR", pd.Series(dtype=float)).sum()) if not packaging_rows.empty else 0.0
    financial_query = safe_float(financial_anomaly_rows.get("Financial_Excess_ZAR", pd.Series(dtype=float)).sum()) if not financial_anomaly_rows.empty else 0.0
    avg_confidence = safe_float(findings_df.get("Confidence_Score", pd.Series(dtype=float)).mean()) if not findings_df.empty else 0.0
    pitch_angle = choose_pitch_angle(findings_df, operational_exposure, direct_recovery, pd.DataFrame(), benchmark_table)
    call_script = build_call_script(client_name, courier_provider, pitch_angle, direct_recovery, operational_exposure, packaging_leak, financial_query, currency_prefix)

    pdf = MarginPDF()
    pdf.alias_nb_pages()
    pdf.set_margins(15, 24, 15)
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()

    def ensure_space(height):
        if pdf.get_y() + height > pdf.page_break_trigger:
            pdf.add_page()

    def clean(text):
        return str(text).replace("–", "-").replace("—", "-").replace("‘", "'").replace("’", "'").replace("“", '"').replace("”", '"')

    def truncate(value, limit):
        value = "" if pd.isna(value) else clean(value)
        return value if len(value) <= limit else value[: limit - 3] + "..."

    def h1(text):
        ensure_space(18)
        pdf.set_font("Helvetica", "B", 18)
        pdf.set_text_color(*MarginPDF.NAVY)
        pdf.multi_cell(0, 9, clean(text), align="C")
        pdf.ln(2)

    def h2(text):
        ensure_space(15)
        pdf.ln(3)
        pdf.set_font("Helvetica", "B", 12)
        pdf.set_text_color(*MarginPDF.NAVY)
        pdf.multi_cell(0, 7, clean(text))
        pdf.set_draw_color(*MarginPDF.BORDER)
        pdf.line(pdf.l_margin, pdf.get_y(), pdf.w - pdf.r_margin, pdf.get_y())
        pdf.ln(3)

    def body(text):
        pdf.set_font("Helvetica", "", 9.5)
        pdf.set_text_color(*MarginPDF.TEXT)
        pdf.multi_cell(0, 5.5, clean(text))
        pdf.ln(1.5)

    def bullet(text):
        body(f"- {text}")

    def table(headers, rows, widths, aligns=None, limits=None):
        aligns = aligns or ["L"] * len(headers)
        limits = limits or [24] * len(headers)
        row_h = 7
        ensure_space(row_h * 2)
        pdf.set_font("Helvetica", "B", 7.5)
        pdf.set_fill_color(*MarginPDF.NAVY)
        pdf.set_text_color(*MarginPDF.WHITE)
        pdf.set_draw_color(*MarginPDF.NAVY)
        for header, width, align in zip(headers, widths, aligns):
            pdf.cell(width, row_h, truncate(header, 24), border=1, align=align, fill=True)
        pdf.ln(row_h)
        pdf.set_font("Helvetica", "", 7)
        pdf.set_draw_color(*MarginPDF.BORDER)
        if not rows:
            rows = [["No records flagged"] + [""] * (len(headers) - 1)]
        for idx, row in enumerate(rows):
            ensure_space(row_h)
            pdf.set_fill_color(*(MarginPDF.ZEBRA if idx % 2 else MarginPDF.WHITE))
            pdf.set_text_color(*MarginPDF.TEXT)
            for value, width, align, limit in zip(row, widths, aligns, limits):
                pdf.cell(width, row_h, truncate(value, limit), border=1, align=align, fill=True)
            pdf.ln(row_h)
        pdf.ln(3)

    h1("Internal Client Briefing Playbook")
    body(f"Client: {safe_text(client_name, 'Local Client')}. Audit period: {audit_month} {audit_year}. Courier: {courier_provider}.")
    body("Use this as internal call preparation and training material. It is not a legal claim letter and should not be forwarded as-is to the client.")

    h2("1. Executive Situation Brief")
    table(
        ["Metric", "Value", "How to explain it"],
        [
            ["Courier spend", format_currency(total_spend, currency_prefix), "Total uploaded courier cost base."],
            ["Direct recovery candidates", format_currency(direct_recovery, currency_prefix), "Evidence-backed claim candidates only."],
            ["Operational exposure", format_currency(operational_exposure, currency_prefix), "Control and prevention opportunity, not automatic refund."],
            ["Packaging leak", format_currency(packaging_leak, currency_prefix), "Warehouse-side avoidable volumetric cost signal."],
            ["Financial/routing queries", format_currency(financial_query, currency_prefix), "Account-manager explanation or credit review."],
            ["Average confidence", f"{avg_confidence:.0f}/100", "Higher after cleaner data, rate cards, and outcomes."],
        ],
        [45, 35, 100],
        ["L", "R", "L"],
        [26, 18, 58],
    )
    body(f"Interpretation: lead with {pitch_angle['Angle'].lower()}. {pitch_angle['Why']} {pitch_angle['Caution']}")

    h2("2. What We Can Safely Say")
    bullet("Direct recovery means evidence-backed claim candidates, usually accepted-vs-charged money differences or high-confidence rate-card reproduction.")
    bullet("Operational exposure is a prevention and governance signal. It should not be described as money the courier definitely owes.")
    bullet("Tier B rate-card findings require the client's signed rate card and courier confirmation before being treated as final recoveries.")
    bullet("If a row is Tier C, Tier D, or Tier E, use it to ask better questions rather than to promise a credit note.")

    h2("3. Evidence Tier Teaching Notes")
    table(
        ["Tier", "Meaning", "How to teach it"],
        [
            ["A", "Direct monetary proof", "Closest to a dispute-ready claim, still subject to courier validation."],
            ["B", "Rate-card reproducible", "Strong if the signed rate card matches; otherwise query."],
            ["C", "Statistical/same-lane query", "Useful for account-manager questions, not proof by itself."],
            ["D", "Operational estimate", "Shows preventable leakage from packaging, weight, or process."],
            ["E", "Data-quality warning", "Fix source data before relying on row-level math."],
        ],
        [18, 48, 114],
        ["L", "L", "L"],
        [10, 28, 70],
    )

    h2("4. Top Findings to Discuss")
    table(
        ["Waybill", "Order", "Tier", "Amount", "Confidence", "Talking point"],
        [[r["Waybill"], r["Order"], r["Tier"], format_currency(r["Amount"], currency_prefix), f"{r['Confidence']:.0f}", r["Talk_Point"]] for r in format_top_finding_rows(findings_df, 8)],
        [25, 25, 18, 25, 20, 67],
        ["L", "L", "L", "R", "R", "L"],
        [14, 14, 10, 14, 8, 42],
    )

    h2("5. Benchmark Interpretation")
    if benchmark_table.empty:
        body("No benchmark table was available for this upload. Treat the conversation as a baseline audit and ask for the next dataset to establish trend history.")
    else:
        table(
            ["Metric", "Value", "Status", "Commercial meaning"],
            [[row.get("Metric", ""), row.get("Display", ""), row.get("Benchmark_Status", ""), row.get("Recommendation", "")] for _, row in benchmark_table.head(8).iterrows()],
            [45, 25, 25, 85],
            ["L", "R", "L", "L"],
            [26, 12, 12, 50],
        )
        body("Benchmark statuses are internal SME heuristics. Use them to explain relative control risk, not as an industry-certified benchmark claim.")

    h2("6. Client-Specific Pitch Angle")
    body(f"Recommended angle: {pitch_angle['Angle']}.")
    body(pitch_angle["Why"])
    body(f"Caution: {pitch_angle['Caution']}")

    h2("7. Discovery Questions to Ask")
    for question in [
        "Do you have the signed courier rate card and current surcharge schedule?",
        "Are accepted rates binding or only provisional quotes?",
        "Who owns courier disputes internally and who approves credit-note follow-up?",
        "How often are weight, dimension, fuel, or zone disputes credited?",
        "Are packaging choices rule-based at the pack station or left to manual judgement?",
        "Has the courier changed fuel, remote-zone, or service-level surcharges recently?",
    ]:
        bullet(question)

    h2("8. Objections and Responses")
    for objection, response in build_objection_responses(currency_prefix):
        body(f"Objection: {objection}\nResponse: {response}")

    h2("9. Call Script")
    for line in call_script:
        bullet(line)

    h2("10. Next-Step Checklist")
    checklist = [
        "Export the evidence-scored dispute pack and send only dispute-ready claims for courier review.",
        "Request the signed rate card and surcharge rules before escalating Tier B items.",
        "Pick the highest-value packaging rule and convert it into a pack-station instruction.",
        "Use financial/routing anomalies as account-manager questions, not guaranteed credits.",
        "Schedule a monthly audit so outcomes become a proprietary benchmark trail.",
    ]
    if not priority_actions.empty:
        for _, action in priority_actions.head(5).iterrows():
            checklist.append(f"Priority action: {safe_text(action.get('Action', ''), '')} - {safe_text(action.get('Next Step', ''), '')}")
    for item in checklist:
        bullet(item)

    if not capital_trap_skus.empty:
        h2("Bonus: SKU Conversation Starters")
        table(
            ["SKU", "Class", "Orders", "Shipping cost", "Talk track"],
            [[row.get("SKU", ""), row.get("Velocity_Class", ""), f"{int(safe_float(row.get('SKU_Order_Frequency', 0))):,}", format_currency(row.get("SKU_Total_Shipping_Cost_ZAR", 0), currency_prefix), "Bundle, liquidate, reduce MOQ, or review margin."] for _, row in capital_trap_skus.head(6).iterrows()],
            [45, 18, 22, 30, 65],
            ["L", "L", "R", "R", "L"],
            [28, 8, 10, 14, 38],
        )

    return bytes(pdf.output())


def generate_client_audit_memo_pdf(
    client_name,
    courier_provider,
    analysis_data,
    findings_df,
    benchmark_table,
    priority_actions,
    currency_prefix,
    audit_month,
    audit_year,
    offer,
):
    """Generate a concise client-facing commercial audit memo."""
    analysis_data = analysis_data.copy() if analysis_data is not None else pd.DataFrame()
    findings_df = findings_df.copy() if findings_df is not None else pd.DataFrame()
    benchmark_table = benchmark_table.copy() if benchmark_table is not None else pd.DataFrame()
    priority_actions = priority_actions.copy() if priority_actions is not None else pd.DataFrame()
    offer = offer or {}

    total_spend = safe_float(analysis_data.get("Billed_Cost_ZAR", pd.Series(dtype=float)).sum()) if not analysis_data.empty else 0.0
    direct_recovery = safe_float(findings_df.get("Direct_Recovery_ZAR", pd.Series(dtype=float)).sum()) if not findings_df.empty else 0.0
    operational_exposure = safe_float(findings_df.get("Operational_Exposure_ZAR", pd.Series(dtype=float)).sum()) if not findings_df.empty else 0.0
    dispute_ready = int(findings_df.get("Dispute_Ready", pd.Series(dtype=bool)).sum()) if not findings_df.empty else 0
    avg_confidence = safe_float(findings_df.get("Confidence_Score", pd.Series(dtype=float)).mean()) if not findings_df.empty else 0.0

    pdf = MarginPDF()
    pdf.alias_nb_pages()
    pdf.set_margins(15, 24, 15)
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()

    def ensure_space(height):
        if pdf.get_y() + height > pdf.page_break_trigger:
            pdf.add_page()

    def clean(text):
        return str(text).replace("–", "-").replace("—", "-").replace("‘", "'").replace("’", "'").replace("“", '"').replace("”", '"')

    def h1(text):
        ensure_space(18)
        pdf.set_font("Helvetica", "B", 18)
        pdf.set_text_color(*MarginPDF.NAVY)
        pdf.multi_cell(0, 9, clean(text), align="C")
        pdf.ln(2)

    def h2(text):
        ensure_space(15)
        pdf.ln(3)
        pdf.set_font("Helvetica", "B", 12)
        pdf.set_text_color(*MarginPDF.NAVY)
        pdf.multi_cell(0, 7, clean(text))
        pdf.set_draw_color(*MarginPDF.BORDER)
        pdf.line(pdf.l_margin, pdf.get_y(), pdf.w - pdf.r_margin, pdf.get_y())
        pdf.ln(3)

    def body(text):
        pdf.set_font("Helvetica", "", 10)
        pdf.set_text_color(*MarginPDF.TEXT)
        pdf.multi_cell(0, 6, clean(text))
        pdf.ln(1.5)

    def bullet(text):
        body(f"- {text}")

    def metric_row(label, value, explanation):
        ensure_space(8)
        pdf.set_font("Helvetica", "B", 9)
        pdf.set_text_color(*MarginPDF.NAVY)
        pdf.cell(50, 7, clean(label), border=1)
        pdf.set_font("Helvetica", "", 9)
        pdf.set_text_color(*MarginPDF.TEXT)
        pdf.cell(35, 7, clean(value), border=1, align="R")
        pdf.cell(95, 7, clean(explanation)[:80], border=1)
        pdf.ln(7)

    h1("Courier Leak Audit Memo")
    body(f"Client: {safe_text(client_name, 'Local Client')} | Audit period: {audit_month} {audit_year} | Courier: {courier_provider}")
    body("This memo summarises the audit findings and recommended commercial next step. Amounts are review candidates and require courier/rate-card confirmation before being treated as final recoveries.")

    h2("1. What We Found")
    metric_row("Courier spend", format_currency(total_spend, currency_prefix), "Uploaded courier cost base reviewed.")
    metric_row("Recovery candidates", format_currency(direct_recovery, currency_prefix), "Evidence-backed query/claim candidates.")
    metric_row("Operational exposure", format_currency(operational_exposure, currency_prefix), "Prevention opportunity, not automatic refund.")
    metric_row("Dispute-ready rows", f"{dispute_ready:,}", "Rows closest to courier submission.")
    metric_row("Avg confidence", f"{avg_confidence:.0f}/100", "Improves with clean data, rate cards, and outcomes.")

    h2("2. What This Means")
    body("The audit separates evidence-backed recovery candidates from operational leakage. This protects the client from overclaiming while still showing where courier billing, rate-card control, data quality, or packaging rules may be costing margin.")
    body("Direct recovery candidates should be submitted or queried with supporting waybill evidence. Operational exposure should drive prevention: packaging rules, rate-card checks, cleaner exports, and monthly exception monitoring.")

    h2("3. Recommended Next Step")
    if priority_actions.empty:
        bullet("Use this upload as a baseline and request the latest courier invoice/export plus signed rate card.")
    else:
        for _, action in priority_actions.head(4).iterrows():
            bullet(f"{safe_text(action.get('Action', ''), 'Review action')}: {safe_text(action.get('Next Step', ''), 'Confirm next step with the client.')}")

    h2("4. Introduction-Stage Commercial Option")
    body(f"Recommended package: {offer.get('package_name', 'Starter Leak Snapshot')}")
    body(f"One-time audit/setup fee: {format_currency_whole(offer.get('recommended_audit_fee', 4500), currency_prefix)}. Optional monthly monitoring: {format_currency_whole(offer.get('recommended_retainer', 3500), currency_prefix)}/mo.")
    body(f"Market range guide: audit {offer.get('audit_range', 'R2,500-R9,500')}; monitoring {offer.get('retainer_range', 'R1,500-R7,500/mo')}. {offer.get('success_fee', 'Optional success fee can be agreed on credited recoveries.')}")
    body(safe_text(offer.get("positioning", "Low-risk courier leakage audit with conservative evidence and clear next actions."), "Low-risk courier leakage audit with conservative evidence and clear next actions."))

    h2("5. What We Need From You")
    for item in [
        "Signed courier rate card and current surcharge schedule.",
        "Courier account manager contact details.",
        "Latest invoice/export covering the period to review.",
        "Approval to submit the dispute pack or query list for courier review.",
        "Permission to track credited/rejected outcomes so future audits become more accurate.",
    ]:
        bullet(item)

    h2("6. CTA")
    body(offer.get("cta", "Book a Courier Leak Audit. Send us your courier export and we'll separate recoverable claim candidates from operational leakage within 48 hours."))

    if not benchmark_table.empty:
        h2("7. Benchmark Note")
        top = benchmark_table.head(3)
        for _, row in top.iterrows():
            bullet(f"{row.get('Metric', '')}: {row.get('Display', '')} ({row.get('Benchmark_Status', '')}) - {row.get('Recommendation', '')}")
        body("Benchmark bands are internal heuristics until enough audited outcomes are collected.")

    return bytes(pdf.output())


def generate_margin_pdf(data, anomaly_rows, packaging_rows, capital_trap_skus, courier_name, divisor, pillar_a_loss, pillar_b_loss, currency_prefix):
    pdf = MarginPDF()
    pdf.alias_nb_pages()
    pdf.set_margins(15, 24, 15)
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()

    def ensure_space(height):
        if pdf.get_y() + height > pdf.page_break_trigger:
            pdf.add_page()

    def money(value):
        return format_currency(value, currency_prefix)

    def truncate(value, limit):
        value = "" if pd.isna(value) else str(value)
        return value if len(value) <= limit else value[: limit - 3] + "..."

    def h1(text):
        ensure_space(18)
        pdf.set_font("Helvetica", "B", 18)
        pdf.set_text_color(*MarginPDF.NAVY)
        pdf.multi_cell(0, 9, text, align="C")
        pdf.ln(2)

    def h2(text):
        ensure_space(15)
        pdf.ln(3)
        pdf.set_font("Helvetica", "B", 12)
        pdf.set_text_color(*MarginPDF.NAVY)
        pdf.multi_cell(0, 7, text)
        pdf.set_draw_color(*MarginPDF.BORDER)
        pdf.line(pdf.l_margin, pdf.get_y(), pdf.w - pdf.r_margin, pdf.get_y())
        pdf.ln(3)

    def body(text):
        pdf.set_font("Helvetica", "", 10)
        pdf.set_text_color(*MarginPDF.TEXT)
        pdf.multi_cell(0, 6, str(text))
        pdf.ln(1.5)

    def table(headers, rows, widths, aligns=None, limits=None):
        aligns = aligns or ["L"] * len(headers)
        limits = limits or [24] * len(headers)
        row_h = 7
        ensure_space(row_h * 2)
        pdf.set_font("Helvetica", "B", 8)
        pdf.set_fill_color(*MarginPDF.NAVY)
        pdf.set_text_color(*MarginPDF.WHITE)
        pdf.set_draw_color(*MarginPDF.NAVY)
        for header, width, align in zip(headers, widths, aligns):
            pdf.cell(width, row_h, truncate(header, 25), border=1, align=align, fill=True)
        pdf.ln(row_h)
        pdf.set_font("Helvetica", "", 7.5)
        pdf.set_draw_color(*MarginPDF.BORDER)
        if not rows:
            rows = [["No records flagged"] + [""] * (len(headers) - 1)]
        for idx, row in enumerate(rows):
            ensure_space(row_h)
            pdf.set_fill_color(*(MarginPDF.ZEBRA if idx % 2 else MarginPDF.WHITE))
            pdf.set_text_color(*MarginPDF.TEXT)
            for value, width, align, limit in zip(row, widths, aligns, limits):
                pdf.cell(width, row_h, truncate(value, limit), border=1, align=align, fill=True)
            pdf.ln(row_h)
        pdf.ln(4)

    h1("360° Margin Diagnostic & Recovery Blueprint")
    body(f"Courier configuration: {courier_name}. Volumetric divisor used: {divisor}.")
    body(f"Total direct exposure identified: {money(pillar_a_loss + pillar_b_loss)}. Direct courier recovery candidates: {money(pillar_a_loss)}. Packaging workflow exposure: {money(pillar_b_loss)}.")

    h2("Chapter 1: Courier Billing Anomalies")
    body("These are the exact orders to query with the courier where the export shows a direct accepted-vs-charged rate delta. Weight-only and dimension-only findings are kept as operational exposure, not automatic credits.")
    table(
        ["Order ID", "SKU", "Actual kg", "Billed Vol kg", "Reason", "Recoverable"],
        [[r["Order_ID"], r["SKU"], f"{safe_float(r['Actual_Weight_KG']):.2f}", f"{safe_float(r['Billed_Vol_KG']):.2f}", r["Anomaly_Reason"], money(r["Recoverable_Overcharge_ZAR"])] for _, r in anomaly_rows.head(30).iterrows()],
        [25, 38, 20, 23, 52, 25],
        ["L", "L", "R", "R", "L", "R"],
        [18, 26, 10, 12, 34, 14],
    )

    h2("Chapter 2: Warehouse Packaging Inefficiencies")
    body("Single-item leaks show flyer-fit opportunities. Mixed-basket leaks show orders where the billed volumetric weight exceeds combined product volume plus a 20% void-fill allowance.")
    table(
        ["Order", "SKU", "Basket", "Reason", "Recommended", "Avoidable"],
        [[r["Order_ID"], r["SKU"], "Multi" if r["Is_Multi_Item"] else "Single", r["Packaging_Reason"], r["Recommended_Package"], money(r["Avoidable_Volumetric_Leak_ZAR"])] for _, r in packaging_rows.head(30).iterrows()],
        [23, 34, 16, 52, 38, 24],
        ["L", "L", "L", "L", "L", "R"],
        [16, 24, 8, 34, 26, 14],
    )

    h2("Chapter 3: ABC Inventory Intelligence")
    body("These C-Class SKUs have low order velocity and high shipping-cost exposure to distant zones. Treat them as capital traps until margin and reorder policy are reviewed.")
    table(
        ["SKU", "Class", "Orders", "Shipping Cost", "Distant Lines", "Action"],
        [[r["SKU"], r["Velocity_Class"], f"{int(safe_float(r['SKU_Order_Frequency'])):,}", money(r["SKU_Total_Shipping_Cost_ZAR"]), f"{int(safe_float(r['SKU_Distant_Zone_Lines'])):,}", "Bundle / liquidate / reduce MOQ"] for _, r in capital_trap_skus.head(30).iterrows()],
        [40, 18, 20, 28, 25, 48],
        ["L", "L", "R", "R", "R", "L"],
        [28, 8, 10, 14, 12, 30],
    )

    h2("Evidence & Confidence Methodology")
    body("All calculations are deterministic. Dirty numeric fields are coerced safely; rows with missing critical data are skipped from pillar math and shown in the data-quality table. Direct courier recovery requires a money-field delta such as accepted rate versus charged/confirmed rate. Weight-only differences are retained as operational exposure so the report does not overclaim courier credits. Multi-item logic uses combined physical item volume plus a 20% allowance for void fill/bubble wrap before flagging packaging bloat.")
    body("Evidence tiers: A = direct monetary proof; B = rate-card reproducible once a client-specific rate card is loaded; C = statistical or same-lane anomaly for account-manager query; D = operational estimate; E = data-quality warning. Dispute packs include only dispute-ready Tier A claims by default.")
    return bytes(pdf.output())


def row_identity_series(df):
    if df is None or df.empty:
        return pd.Series(dtype=str)
    for column in ["Waybill_ID", "Tracking_Number", "Order_ID"]:
        if column in df.columns:
            return df[column].fillna("").astype(str)
    return pd.Series(df.index.astype(str), index=df.index)


def calculate_leakage_totals(analysis_data, findings_df, financial_anomaly_rows=None, packaging_rows=None):
    """Return non-overlapping leakage buckets used by headline metrics."""
    analysis = analysis_data.copy() if analysis_data is not None else pd.DataFrame()
    findings = findings_df.copy() if findings_df is not None else pd.DataFrame()
    financial = financial_anomaly_rows.copy() if financial_anomaly_rows is not None else pd.DataFrame()

    spend_base = safe_float(analysis.get("Billed_Cost_ZAR", pd.Series(dtype=float)).sum()) if not analysis.empty else 0.0
    financial_spend = safe_float(financial.get("cost", pd.Series(dtype=float)).sum()) if not financial.empty else 0.0
    spend_base = max(spend_base, financial_spend)

    direct_recovery = 0.0
    account_query = 0.0
    financial_query = safe_float(financial.get("Financial_Excess_ZAR", pd.Series(dtype=float)).sum()) if not financial.empty else 0.0
    claimed_ids = set()

    if not findings.empty and "Recovery_Type" in findings.columns:
        direct_mask = findings["Recovery_Type"].eq(RECOVERY_DIRECT)
        query_mask = findings["Recovery_Type"].eq(RECOVERY_QUERY)
        financial_finding_mask = findings.get("Finding_Category", pd.Series("", index=findings.index)).eq("SKYNET_SAME_LANE_SPIKE")

        direct_recovery = safe_float(pd.to_numeric(findings.loc[direct_mask, "Direct_Recovery_ZAR"], errors="coerce").fillna(0).sum()) if "Direct_Recovery_ZAR" in findings.columns else 0.0
        account_query = safe_float(pd.to_numeric(findings.loc[query_mask & ~financial_finding_mask, "Direct_Recovery_ZAR"], errors="coerce").fillna(0).sum()) if "Direct_Recovery_ZAR" in findings.columns else 0.0
        claimed_ids = set(row_identity_series(findings.loc[direct_mask | query_mask]).replace("", pd.NA).dropna().tolist())

    operational_exposure = 0.0
    if not analysis.empty:
        operational_candidates = []
        for column in ["Operational_Exposure_ZAR", "Avoidable_Volumetric_Leak_ZAR"]:
            if column in analysis.columns:
                operational_candidates.append(pd.to_numeric(analysis[column], errors="coerce").fillna(0))
        if operational_candidates:
            operational_frame = pd.concat(operational_candidates, axis=1)
            operational_per_row = operational_frame.max(axis=1)
            analysis_ids = row_identity_series(analysis)
            direct_or_query_rows = analysis_ids.isin(claimed_ids) if claimed_ids else pd.Series(False, index=analysis.index)
            operational_exposure = safe_float(operational_per_row.loc[~direct_or_query_rows].sum())

    headline_leakage = direct_recovery + account_query + financial_query + operational_exposure
    leakage_rate = headline_leakage / spend_base if spend_base else 0.0
    return {
        "direct_recovery_zar": round(direct_recovery, 2),
        "account_manager_query_zar": round(account_query, 2),
        "financial_query_zar": round(financial_query, 2),
        "operational_exposure_zar": round(operational_exposure, 2),
        "headline_leakage_zar": round(headline_leakage, 2),
        "spend_base_zar": round(spend_base, 2),
        "leakage_rate": leakage_rate,
    }


def detect_repeated_issue_patterns(findings_df, packaging_rows, financial_anomaly_rows):
    rows = []
    findings = findings_df.copy() if findings_df is not None else pd.DataFrame()
    if not findings.empty:
        group_cols = [col for col in ["Finding_Category", "Courier", "Service_Level", "Destination"] if col in findings.columns]
        if group_cols:
            frame = findings.copy()
            frame["Issue_Amount"] = pd.to_numeric(frame.get("Direct_Recovery_ZAR", 0), errors="coerce").fillna(0) + pd.to_numeric(frame.get("Operational_Exposure_ZAR", 0), errors="coerce").fillna(0)
            grouped = frame.groupby(group_cols, dropna=False).agg(Count=("Issue_Amount", "size"), Amount=("Issue_Amount", "sum")).reset_index()
            for _, row in grouped[grouped["Count"] >= 2].sort_values(["Amount", "Count"], ascending=False).head(8).iterrows():
                issue_type = humanize_label(row.get("Finding_Category", "Issue"))
                context = " / ".join(safe_text(row.get(col, ""), "") for col in group_cols if col != "Finding_Category" and safe_text(row.get(col, ""), ""))
                rows.append({
                    "Issue Type": issue_type,
                    "Context": context or "Multiple rows",
                    "Count": int(row["Count"]),
                    "Amount_ZAR": round(safe_float(row["Amount"]), 2),
                    "Why It Matters": "Repeated current-month signal; check whether this becomes a monthly control rule.",
                    "Next Control Action": "Track this issue next month and compare against courier/rate-card changes.",
                })
    packaging = packaging_rows.copy() if packaging_rows is not None else pd.DataFrame()
    if not packaging.empty and "Packaging_Reason" in packaging.columns:
        grouped = packaging.groupby("Packaging_Reason", dropna=False).agg(Count=("Packaging_Reason", "size"), Amount_ZAR=("Avoidable_Volumetric_Leak_ZAR", "sum")).reset_index()
        for _, row in grouped[grouped["Count"] >= 2].sort_values(["Amount_ZAR", "Count"], ascending=False).head(5).iterrows():
            rows.append({
                "Issue Type": "Packaging / Volumetric Exposure",
                "Context": safe_text(row.get("Packaging_Reason", "Packaging issue"), "Packaging issue"),
                "Count": int(row["Count"]),
                "Amount_ZAR": round(safe_float(row["Amount_ZAR"]), 2),
                "Why It Matters": "Repeated packaging signal can become a pack-station rule.",
                "Next Control Action": "Add to next month's packaging checklist and confirm whether the rule reduced billed weight.",
            })
    financial = financial_anomaly_rows.copy() if financial_anomaly_rows is not None else pd.DataFrame()
    if not financial.empty and "Service_Level" in financial.columns:
        grouped = financial.groupby("Service_Level", dropna=False).agg(Count=("Service_Level", "size"), Amount_ZAR=("Financial_Excess_ZAR", "sum")).reset_index()
        for _, row in grouped[grouped["Count"] >= 2].sort_values(["Amount_ZAR", "Count"], ascending=False).head(5).iterrows():
            rows.append({
                "Issue Type": "Financial / Routing Query",
                "Context": f"Service {safe_text(row.get('Service_Level', 'Unknown'), 'Unknown')}",
                "Count": int(row["Count"]),
                "Amount_ZAR": round(safe_float(row["Amount_ZAR"]), 2),
                "Why It Matters": "Repeated service-level cost spikes should be queried with the courier account manager.",
                "Next Control Action": "Ask for surcharge/routing explanation and compare again next month.",
            })
    return pd.DataFrame(rows, columns=["Issue Type", "Context", "Count", "Amount_ZAR", "Why It Matters", "Next Control Action"])


def build_monthly_monitoring_summary(analysis_data, findings_df, leakage_totals, outcome_summary, evidence_readiness, rate_card_matched_count):
    analysis = analysis_data if analysis_data is not None else pd.DataFrame()
    findings = findings_df if findings_df is not None else pd.DataFrame()
    leakage_totals = leakage_totals or {}
    outcome_summary = outcome_summary or {}
    evidence_readiness = evidence_readiness or {}
    query_mask = findings["Recovery_Type"].eq(RECOVERY_QUERY) if not findings.empty and "Recovery_Type" in findings.columns else pd.Series(False, index=findings.index)
    direct_mask = findings["Recovery_Type"].eq(RECOVERY_DIRECT) if not findings.empty and "Recovery_Type" in findings.columns else pd.Series(False, index=findings.index)
    return {
        "courier_spend": safe_float(leakage_totals.get("spend_base_zar", analysis.get("Billed_Cost_ZAR", pd.Series(dtype=float)).sum() if not analysis.empty else 0)),
        "direct_recovery": safe_float(leakage_totals.get("direct_recovery_zar", 0)),
        "account_manager_query": safe_float(leakage_totals.get("account_manager_query_zar", 0)),
        "operational_exposure": safe_float(leakage_totals.get("operational_exposure_zar", 0)),
        "credited_outcomes": safe_float(outcome_summary.get("credited_amount", 0)),
        "rejected_rows": int(safe_float(outcome_summary.get("rejection_count", 0))),
        "rate_card_matched_rows": int(safe_float(rate_card_matched_count)),
        "evidence_score": int(safe_float(evidence_readiness.get("score", 0))),
        "readiness_band": evidence_readiness.get("band", "Not scored"),
        "issue_count": len(findings),
        "direct_rows": int(direct_mask.sum()),
        "query_rows": int(query_mask.sum()),
    }


def build_monitoring_retainer_pitch(summary, repeated_patterns, currency_prefix):
    repeated_count = len(repeated_patterns) if repeated_patterns is not None else 0
    return (
        f"This month shows {format_currency(summary.get('direct_recovery', 0) + summary.get('account_manager_query', 0), currency_prefix)} in recovery/query candidates and "
        f"{format_currency(summary.get('operational_exposure', 0), currency_prefix)} in operational exposure. "
        f"There are {repeated_count} current-month repeat signals to watch next month. A once-off audit catches today's leakage; monthly monitoring checks whether the same issues repeat, whether courier responses become credits or rejections, and whether operational fixes reduce future spend. "
        "The economic reason for continuous monitoring is simple: courier rate cards, fuel surcharges, service levels, remote/outlying rules, VAT treatment, and weight rounding rules can change as fuel prices, inflation, exchange rates, labour costs, and courier network costs shift. "
        "The tool keeps the client aligned to the latest rate-card context by rechecking new exports and newly supplied rate cards each month, so the business does not keep paying old assumptions against moving courier rules."
    )


def build_next_month_checklist(cleaned_rate_card):
    checklist = [
        "Latest courier export for the new month",
        "Latest courier invoice/statement for the same period",
        "Updated signed rate card or confirmation that rates did not change",
        "Fuel surcharge, VAT, remote/outlying and service surcharge updates",
        "Volumetric divisor and weight-rounding rule confirmation",
        "Courier responses: credited, rejected with reason, or needs more evidence",
        "Review repeated issue signals from this month",
        "Check whether packaging rules changed billed/volumetric weight",
        "Refresh client memo, dispute pack, and monitoring score",
    ]
    if cleaned_rate_card is None or cleaned_rate_card.empty:
        checklist.insert(2, "Request the missing client rate card before making strong rate-card claims")
    return checklist


def build_monthly_monitoring_pack_text(summary, repeated_patterns, retainer_pitch, checklist, client_message, currency_prefix):
    lines = [
        "Monthly Monitoring Pack",
        "=======================",
        "",
        "SUMMARY",
        f"Courier spend: {format_currency(summary.get('courier_spend', 0), currency_prefix)}",
        f"Recovery/query candidates: {format_currency(summary.get('direct_recovery', 0) + summary.get('account_manager_query', 0), currency_prefix)}",
        f"Operational exposure: {format_currency(summary.get('operational_exposure', 0), currency_prefix)}",
        f"Credited outcomes: {format_currency(summary.get('credited_outcomes', 0), currency_prefix)}",
        f"Rate-card matched rows: {summary.get('rate_card_matched_rows', 0):,}",
        f"Evidence score: {summary.get('evidence_score', 0)}/100",
        "",
        "WHY MONTHLY MONITORING",
        retainer_pitch,
        "",
        "CURRENT-MONTH REPEAT SIGNALS",
    ]
    if repeated_patterns is not None and not repeated_patterns.empty:
        for _, row in repeated_patterns.iterrows():
            lines.append(f"- {row.get('Issue Type')}: {row.get('Count')} rows, {format_currency(row.get('Amount_ZAR', 0), currency_prefix)} — {row.get('Next Control Action')}")
    else:
        lines.append("- No repeated issue signals detected yet. Use this month as the baseline.")
    lines.extend(["", "NEXT MONTH CHECKLIST"])
    lines.extend(f"[ ] {item}" for item in checklist)
    lines.extend(["", "CLIENT MESSAGE", client_message])
    return "\n".join(lines).strip() + "\n"


MONTHLY_SNAPSHOT_COLUMNS = [
    "Client", "Audit_Month", "Audit_Year", "Courier", "Courier_Spend_ZAR", "Direct_Recovery_ZAR",
    "Account_Manager_Query_ZAR", "Operational_Exposure_ZAR", "Credited_Outcomes_ZAR", "Rejected_Rows",
    "Evidence_Score", "Readiness_Band", "Rate_Card_Matched_Rows", "Issue_Count", "Top_Issue_Type",
    "Top_Issue_Context", "Leakage_Rate", "Updated_At",
]
MONTH_NAME_ORDER = {
    "january": 1,
    "jan": 1,
    "february": 2,
    "feb": 2,
    "march": 3,
    "mar": 3,
    "april": 4,
    "apr": 4,
    "may": 5,
    "june": 6,
    "jun": 6,
    "july": 7,
    "jul": 7,
    "august": 8,
    "aug": 8,
    "september": 9,
    "sep": 9,
    "october": 10,
    "oct": 10,
    "november": 11,
    "nov": 11,
    "december": 12,
    "dec": 12,
}


def month_sort_key(month_label):
    text = safe_text(month_label, "").strip().lower()
    if text in MONTH_NAME_ORDER:
        return MONTH_NAME_ORDER[text]
    numeric_month = safe_float(text, 0)
    return int(numeric_month) if 1 <= numeric_month <= 12 else 0


def build_monthly_snapshot(client_name, audit_month, audit_year, courier_provider, monitoring_summary, leakage_totals, repeated_patterns):
    monitoring_summary = monitoring_summary or {}
    leakage_totals = leakage_totals or {}
    top_issue = repeated_patterns.iloc[0] if repeated_patterns is not None and not repeated_patterns.empty else {}
    return {
        "Client": safe_text(client_name, "Local Client"),
        "Audit_Month": safe_text(audit_month, "Unknown"),
        "Audit_Year": int(safe_float(audit_year, 0)),
        "Courier": safe_text(courier_provider, "Unknown"),
        "Courier_Spend_ZAR": safe_float(monitoring_summary.get("courier_spend", 0)),
        "Direct_Recovery_ZAR": safe_float(monitoring_summary.get("direct_recovery", 0)),
        "Account_Manager_Query_ZAR": safe_float(monitoring_summary.get("account_manager_query", 0)),
        "Operational_Exposure_ZAR": safe_float(monitoring_summary.get("operational_exposure", 0)),
        "Credited_Outcomes_ZAR": safe_float(monitoring_summary.get("credited_outcomes", 0)),
        "Rejected_Rows": int(safe_float(monitoring_summary.get("rejected_rows", 0))),
        "Evidence_Score": int(safe_float(monitoring_summary.get("evidence_score", 0))),
        "Readiness_Band": safe_text(monitoring_summary.get("readiness_band", "Not scored"), "Not scored"),
        "Rate_Card_Matched_Rows": int(safe_float(monitoring_summary.get("rate_card_matched_rows", 0))),
        "Issue_Count": int(safe_float(monitoring_summary.get("issue_count", 0))),
        "Top_Issue_Type": safe_text(top_issue.get("Issue Type", "No repeated signal") if isinstance(top_issue, pd.Series) else "No repeated signal", "No repeated signal"),
        "Top_Issue_Context": safe_text(top_issue.get("Context", "") if isinstance(top_issue, pd.Series) else "", ""),
        "Leakage_Rate": safe_float(leakage_totals.get("leakage_rate", 0)),
        "Updated_At": f"{safe_text(audit_month, 'Unknown')} {int(safe_float(audit_year, 0))}",
    }


def empty_monthly_snapshots():
    return pd.DataFrame(columns=MONTHLY_SNAPSHOT_COLUMNS)


def load_monthly_snapshots():
    if not MONTHLY_SNAPSHOT_PATH.exists():
        return empty_monthly_snapshots()
    try:
        history = pd.read_csv(MONTHLY_SNAPSHOT_PATH)
    except Exception:
        return empty_monthly_snapshots()
    for column in MONTHLY_SNAPSHOT_COLUMNS:
        if column not in history.columns:
            history[column] = pd.NA
    return history[MONTHLY_SNAPSHOT_COLUMNS]


def save_monthly_snapshot(snapshot):
    history = load_monthly_snapshots()
    snapshot_frame = pd.DataFrame([snapshot], columns=MONTHLY_SNAPSHOT_COLUMNS)
    if not history.empty:
        duplicate_mask = (
            history["Client"].astype(str).eq(str(snapshot["Client"]))
            & history["Audit_Month"].astype(str).eq(str(snapshot["Audit_Month"]))
            & pd.to_numeric(history["Audit_Year"], errors="coerce").fillna(0).astype(int).eq(int(snapshot["Audit_Year"]))
            & history["Courier"].astype(str).eq(str(snapshot["Courier"]))
        )
        history = history.loc[~duplicate_mask].copy()
    updated = pd.concat([history, snapshot_frame], ignore_index=True)
    updated.to_csv(MONTHLY_SNAPSHOT_PATH, index=False)
    return updated


def compare_monthly_snapshots(current_snapshot, history):
    if history is None or history.empty:
        return {"has_previous": False, "message": "No prior monthly snapshot saved for comparison yet."}
    frame = history.copy()
    same_client = frame["Client"].astype(str).eq(str(current_snapshot["Client"])) & frame["Courier"].astype(str).eq(str(current_snapshot["Courier"]))
    frame = frame[same_client].copy()
    if frame.empty:
        return {"has_previous": False, "message": "No prior snapshot for this client/courier yet."}
    frame["Audit_Year"] = pd.to_numeric(frame["Audit_Year"], errors="coerce").fillna(0).astype(int)
    current_year = int(current_snapshot["Audit_Year"])
    current_month = str(current_snapshot["Audit_Month"])
    prior = frame[~(frame["Audit_Year"].eq(current_year) & frame["Audit_Month"].astype(str).eq(current_month))].copy()
    if prior.empty:
        return {"has_previous": False, "message": "Only the current month is saved so far."}
    prior["Audit_Month_Order"] = prior["Audit_Month"].apply(month_sort_key)
    previous = prior.sort_values(["Audit_Year", "Audit_Month_Order", "Audit_Month"]).iloc[-1].to_dict()
    def delta(column):
        return safe_float(current_snapshot.get(column, 0)) - safe_float(previous.get(column, 0))
    return {
        "has_previous": True,
        "previous": previous,
        "spend_delta": delta("Courier_Spend_ZAR"),
        "recovery_query_delta": delta("Direct_Recovery_ZAR") + delta("Account_Manager_Query_ZAR"),
        "operational_delta": delta("Operational_Exposure_ZAR"),
        "credited_delta": delta("Credited_Outcomes_ZAR"),
        "evidence_score_delta": delta("Evidence_Score"),
        "leakage_rate_delta": delta("Leakage_Rate"),
        "top_issue_repeated": safe_text(current_snapshot.get("Top_Issue_Type", ""), "") == safe_text(previous.get("Top_Issue_Type", ""), ""),
    }


def build_monthly_trend_narrative(comparison, currency_prefix):
    if not comparison.get("has_previous"):
        return comparison.get("message", "No previous snapshot available yet. Save this month as the baseline.")
    previous = comparison.get("previous", {})
    direction = lambda value: "increased" if value > 0 else "decreased" if value < 0 else "stayed flat"
    return (
        f"Compared with {previous.get('Audit_Month', 'the previous month')} {previous.get('Audit_Year', '')}, courier spend {direction(comparison['spend_delta'])} by {format_currency(abs(comparison['spend_delta']), currency_prefix)}. "
        f"Recovery/query candidates {direction(comparison['recovery_query_delta'])} by {format_currency(abs(comparison['recovery_query_delta']), currency_prefix)}, while operational exposure {direction(comparison['operational_delta'])} by {format_currency(abs(comparison['operational_delta']), currency_prefix)}. "
        f"Credited outcomes {direction(comparison['credited_delta'])} by {format_currency(abs(comparison['credited_delta']), currency_prefix)} and the evidence score changed by {comparison['evidence_score_delta']:.0f} points. "
        f"Top issue repeated: {'yes' if comparison.get('top_issue_repeated') else 'no'}."
    )


def calculate_market_gap_score(data, anomaly_rows, packaging_rows, financial_anomaly_rows, sku_summary, leakage_totals=None):
    """Score the market gap: outsourced shipping-margin control for reachable SMEs."""
    row_count = len(data) if data is not None else 0
    if leakage_totals is None:
        leakage_totals = calculate_leakage_totals(data, pd.DataFrame(), financial_anomaly_rows, packaging_rows)
    leakage_rate = safe_float(leakage_totals.get("leakage_rate", 0))
    packaging_rate = len(packaging_rows) / row_count if row_count else 0
    anomaly_rate = (len(anomaly_rows) + len(financial_anomaly_rows)) / row_count if row_count else 0
    sku_count = sku_summary["SKU"].nunique() if sku_summary is not None and not sku_summary.empty and "SKU" in sku_summary.columns else 0
    multi_item_rate = data["Is_Multi_Item"].mean() if row_count and "Is_Multi_Item" in data.columns else 0

    score = 0
    score += min(25, int(leakage_rate * 250))
    score += min(20, int(packaging_rate * 100))
    score += min(20, int(anomaly_rate * 160))
    score += min(15, int(sku_count / 4))
    score += min(10, int(multi_item_rate * 40))
    score += 10 if row_count >= 100 else max(0, int(row_count / 10))
    return min(100, score), leakage_rate, packaging_rate, anomaly_rate, sku_count, multi_item_rate


def build_priority_actions(data, anomaly_rows, packaging_rows, financial_anomaly_rows, sku_summary, currency_prefix):
    actions = []
    if not anomaly_rows.empty:
        amount = safe_float(anomaly_rows["Recoverable_Overcharge_ZAR"].sum())
        actions.append({
            "Priority": 1,
            "Action": "Send courier dispute pack",
            "Why": f"Potential recoverable overcharges of {format_currency(amount, currency_prefix)} were isolated at waybill level.",
            "Owner": "Finance / courier account manager",
            "Next Step": "Export the dispute pack and ask the courier to review the listed waybills.",
        })
    if not financial_anomaly_rows.empty:
        amount = safe_float(financial_anomaly_rows["Financial_Excess_ZAR"].sum())
        actions.append({
            "Priority": 2,
            "Action": "Query unusual surcharge and routing charges",
            "Why": f"Financial/routing anomalies total {format_currency(amount, currency_prefix)}.",
            "Owner": "Finance / operations",
            "Next Step": "Send the financial anomalies workbook to the account manager for explanation or credit review.",
        })
    if not packaging_rows.empty:
        amount = safe_float(packaging_rows["Avoidable_Volumetric_Leak_ZAR"].sum())
        top_package = safe_text(packaging_rows.get("Recommended_Package", pd.Series(["packaging rule review"])).mode().iloc[0] if "Recommended_Package" in packaging_rows.columns and not packaging_rows["Recommended_Package"].mode().empty else "packaging rule review")
        actions.append({
            "Priority": 3,
            "Action": "Fix packaging rules at the pack station",
            "Why": f"Avoidable packaging leakage totals {format_currency(amount, currency_prefix)}. Most common recommendation: {top_package}.",
            "Owner": "Warehouse / fulfilment",
            "Next Step": "Create a one-page packaging matrix for the top leaking SKUs and remove manual box choice where possible.",
        })
    if sku_summary is not None and not sku_summary.empty:
        c_skus = sku_summary[sku_summary["Velocity_Class"].eq("C")] if "Velocity_Class" in sku_summary.columns else pd.DataFrame()
        if not c_skus.empty:
            actions.append({
                "Priority": 4,
                "Action": "Review slow SKUs with high delivery cost",
                "Why": f"{len(c_skus):,} C-class SKUs may be tying up cash or creating weak delivery economics.",
                "Owner": "Founder / merchandising",
                "Next Step": "Bundle, reprice, reduce reorder quantities, or restrict free delivery for the worst C-class SKUs.",
            })
    if not actions:
        actions.append({
            "Priority": 1,
            "Action": "Keep monitoring monthly",
            "Why": "No major leakage was detected in this upload, which is a good baseline.",
            "Owner": "Founder / operations",
            "Next Step": "Run the same audit next month to catch changes before they become expensive habits.",
        })
    return pd.DataFrame(actions).sort_values("Priority")


def calculate_benchmark_status(metric_name, value):
    bands = {
        "Direct recovery rate": [(0.01, "controlled"), (0.03, "monitor"), (0.06, "material"), (float("inf"), "urgent")],
        "Operational exposure rate": [(0.03, "controlled"), (0.08, "monitor"), (0.15, "material"), (float("inf"), "urgent")],
        "Data quality skip rate": [(0.02, "controlled"), (0.08, "monitor"), (0.15, "material"), (float("inf"), "urgent")],
        "Dispute-ready rate": [(0.01, "controlled"), (0.04, "monitor"), (0.08, "material"), (float("inf"), "urgent")],
        "Average confidence score": [(50, "urgent"), (70, "material"), (85, "monitor"), (float("inf"), "controlled")],
    }
    for threshold, status in bands.get(metric_name, [(float("inf"), "monitor")]):
        if value <= threshold:
            return status
    return "monitor"


def build_benchmark_table(analysis_data, findings_df, total_spend, skipped_rows):
    row_count = len(analysis_data) if analysis_data is not None else 0
    skipped_count = len(skipped_rows) if skipped_rows is not None else 0
    if findings_df is not None and not findings_df.empty:
        direct_mask = findings_df["Recovery_Type"].eq(RECOVERY_DIRECT) if "Recovery_Type" in findings_df.columns else pd.Series(True, index=findings_df.index)
        direct_recovery = safe_float(findings_df.loc[direct_mask, "Direct_Recovery_ZAR"].sum()) if "Direct_Recovery_ZAR" in findings_df.columns else 0
        operational_exposure = safe_float(findings_df.get("Operational_Exposure_ZAR", pd.Series(dtype=float)).sum())
        dispute_ready_count = int(findings_df.get("Dispute_Ready", pd.Series(dtype=bool)).sum())
        avg_confidence = safe_float(findings_df.get("Confidence_Score", pd.Series(dtype=float)).mean())
    else:
        direct_recovery = 0
        operational_exposure = 0
        dispute_ready_count = 0
        avg_confidence = 0
    metrics = [
        {"Metric": "Direct recovery rate", "Value": direct_recovery / total_spend if total_spend else 0, "Display": f"{(direct_recovery / total_spend if total_spend else 0):.1%}", "Recommendation": "Submit high-confidence claim pack and track courier credits."},
        {"Metric": "Operational exposure rate", "Value": operational_exposure / total_spend if total_spend else 0, "Display": f"{(operational_exposure / total_spend if total_spend else 0):.1%}", "Recommendation": "Use packaging/rate-card review to convert exposure into preventable savings."},
        {"Metric": "Data quality skip rate", "Value": skipped_count / max(row_count + skipped_count, 1), "Display": f"{(skipped_count / max(row_count + skipped_count, 1)):.1%}", "Recommendation": "Request cleaner exports or fix source columns before client-facing claims."},
        {"Metric": "Dispute-ready rate", "Value": dispute_ready_count / max(row_count, 1), "Display": f"{(dispute_ready_count / max(row_count, 1)):.1%}", "Recommendation": "High rates indicate strong account-manager recovery workflow potential."},
        {"Metric": "Average confidence score", "Value": avg_confidence, "Display": f"{avg_confidence:.0f}/100", "Recommendation": "Improve confidence with rate cards, cleaner dimensions, and dispute outcomes."},
    ]
    table = pd.DataFrame(metrics)
    table["Benchmark_Status"] = table.apply(lambda row: calculate_benchmark_status(row["Metric"], row["Value"]), axis=1)
    table["Benchmark_Context"] = "Internal SME heuristic band until enough anonymized client history is available."
    return table


OUTCOME_STATUSES = ["not_submitted", "sent_to_courier", "accepted", "credited", "rejected", "needs_more_evidence"]


def build_outcome_tracker_template(findings_df, currency_prefix):
    """Build an editable outcome tracker from actionable findings."""
    columns = [
        "Finding_Category", "Evidence_Tier", "Recovery_Type", "Waybill_ID", "Order_ID",
        "Direct_Recovery_ZAR", "Confidence_Score", "Submission_Status", "Courier_Response",
        "Credited_Amount_ZAR", "Rejected_Reason", "Lesson_Learned", "Rule_To_Add_Next_Audit",
        "Case_Study_Permission", "Testimonial_Permission", "Follow_Up_Date", "Notes",
    ]
    if findings_df is None or findings_df.empty:
        return pd.DataFrame(columns=columns)

    frame = findings_df.copy()
    actionable = pd.Series(True, index=frame.index)
    if "Recovery_Type" in frame.columns:
        actionable = frame["Recovery_Type"].isin([RECOVERY_DIRECT, RECOVERY_QUERY])
    if "Dispute_Ready" in frame.columns:
        actionable = actionable | frame["Dispute_Ready"].eq(True)
    frame = frame[actionable].copy()
    if frame.empty:
        return pd.DataFrame(columns=columns)

    for column in columns:
        if column not in frame.columns:
            frame[column] = "" if column not in ["Direct_Recovery_ZAR", "Confidence_Score", "Credited_Amount_ZAR"] else 0.0
    frame["Submission_Status"] = frame["Submission_Status"].replace("", "not_submitted")
    frame["Case_Study_Permission"] = frame["Case_Study_Permission"].replace("", "No")
    frame["Testimonial_Permission"] = frame["Testimonial_Permission"].replace("", "No")
    frame["Credited_Amount_ZAR"] = pd.to_numeric(frame["Credited_Amount_ZAR"], errors="coerce").fillna(0.0)
    return frame[columns].sort_values("Direct_Recovery_ZAR", ascending=False).reset_index(drop=True)


def summarize_outcomes(outcome_df, currency_prefix):
    if isinstance(outcome_df, dict):
        outcome_df = outcome_df.get("data", pd.DataFrame())
    if outcome_df is None or not hasattr(outcome_df, "empty") or outcome_df.empty:
        return {
            "submitted_amount": 0.0,
            "credited_amount": 0.0,
            "conversion_rate": 0.0,
            "rejection_count": 0,
            "needs_evidence_count": 0,
            "credited_count": 0,
            "best_tier": "n/a",
        }
    frame = outcome_df.copy()
    frame["Direct_Recovery_ZAR"] = pd.to_numeric(frame.get("Direct_Recovery_ZAR", 0), errors="coerce").fillna(0.0)
    frame["Credited_Amount_ZAR"] = pd.to_numeric(frame.get("Credited_Amount_ZAR", 0), errors="coerce").fillna(0.0)
    statuses = frame.get("Submission_Status", pd.Series("not_submitted", index=frame.index)).astype(str)
    submitted_mask = statuses.ne("not_submitted")
    submitted_amount = safe_float(frame.loc[submitted_mask, "Direct_Recovery_ZAR"].sum())
    credited_amount = safe_float(frame["Credited_Amount_ZAR"].sum())
    credited_rows = frame[frame["Credited_Amount_ZAR"].gt(0)]
    best_tier = "n/a"
    if not credited_rows.empty and "Evidence_Tier" in credited_rows.columns:
        best_tier = safe_text(credited_rows.groupby("Evidence_Tier")["Credited_Amount_ZAR"].sum().sort_values(ascending=False).index[0], "n/a")
    return {
        "submitted_amount": submitted_amount,
        "credited_amount": credited_amount,
        "conversion_rate": credited_amount / submitted_amount if submitted_amount else 0.0,
        "rejection_count": int(statuses.eq("rejected").sum()),
        "needs_evidence_count": int(statuses.eq("needs_more_evidence").sum()),
        "credited_count": int((statuses.eq("credited") | frame["Credited_Amount_ZAR"].gt(0)).sum()),
        "best_tier": best_tier,
    }


def permission_granted(value):
    return str(value or "").strip().lower() in {"yes", "y", "true", "approved", "granted", "permission granted"}


def outcome_permission_summary(outcome_df):
    if outcome_df is None or not hasattr(outcome_df, "empty") or outcome_df.empty:
        return {"case_study": False, "testimonial": False}
    frame = outcome_df.copy()
    return {
        "case_study": frame.get("Case_Study_Permission", pd.Series("No", index=frame.index)).apply(permission_granted).any(),
        "testimonial": frame.get("Testimonial_Permission", pd.Series("No", index=frame.index)).apply(permission_granted).any(),
    }


def build_case_study_snapshot(client_name, outcome_summary, currency_prefix, outcome_df=None, named=False):
    client_name = safe_text(client_name, "this client") or "this client"
    credited = safe_float(outcome_summary.get("credited_amount", 0))
    submitted = safe_float(outcome_summary.get("submitted_amount", 0))
    conversion = safe_float(outcome_summary.get("conversion_rate", 0))
    permission = outcome_permission_summary(outcome_df)
    can_name = named and (permission["case_study"] or permission["testimonial"])
    display_name = client_name if can_name else "an e-commerce client"
    if credited <= 0:
        return (
            f"Internal case study draft for {client_name}: outcome tracking has started, but no credited recovery has been recorded yet. "
            "Use this as a learning record until courier responses and credit notes are confirmed."
        )
    permission_note = "Client permission is recorded for named use." if can_name else "Use anonymously unless the client approves named publication."
    return (
        f"Case study draft: {display_name} recorded {format_currency(credited, currency_prefix)} in credited courier outcomes against "
        f"{format_currency(submitted, currency_prefix)} submitted/query candidates, a {conversion:.1%} recorded conversion rate. "
        "The audit separated recovery candidates from operational exposure, then used outcome tracking to turn courier responses into a reusable control loop. "
        f"{permission_note}"
    )


def build_proof_asset_pack(client_name, courier_provider, findings_df, outcome_df, outcome_summary, currency_prefix, offer):
    client_name = safe_text(client_name, "the client") or "the client"
    offer = offer or {}
    outcome_summary = outcome_summary or summarize_outcomes(outcome_df, currency_prefix)
    credited = safe_float(outcome_summary.get("credited_amount", 0))
    submitted = safe_float(outcome_summary.get("submitted_amount", 0))
    conversion = safe_float(outcome_summary.get("conversion_rate", 0))
    credited_count = int(outcome_summary.get("credited_count", 0))
    rejection_count = int(outcome_summary.get("rejection_count", 0))
    direct_candidates = safe_float(findings_df.get("Direct_Recovery_ZAR", pd.Series(dtype=float)).sum()) if findings_df is not None and not findings_df.empty else 0.0
    permission = outcome_permission_summary(outcome_df)
    anonymized_case = build_case_study_snapshot(client_name, outcome_summary, currency_prefix, outcome_df, named=False)
    named_case = build_case_study_snapshot(client_name, outcome_summary, currency_prefix, outcome_df, named=True)
    proof_line = (
        f"Recorded outcome: {format_currency(credited, currency_prefix)} credited from {format_currency(submitted, currency_prefix)} submitted/query candidates."
        if credited > 0 else
        f"Audit baseline: {format_currency(direct_candidates, currency_prefix)} in recovery/query candidates identified; courier outcomes still being tracked."
    )
    monitoring_fee = format_currency_whole(offer.get("recommended_retainer", 3500), currency_prefix)
    return {
        "anonymized_case_study": anonymized_case,
        "named_case_study": named_case if permission["case_study"] or permission["testimonial"] else "Named publication is not approved yet. Use the anonymized case study until written permission is recorded.",
        "linkedin_proof_post": f"Courier billing control is not just about finding a once-off refund. It is about creating an evidence loop: audit the export, separate claim candidates from operational exposure, send clean courier queries, and track credited/rejected outcomes. {proof_line} This is why e-commerce brands need courier margin monitoring, not just another dashboard.",
        "whatsapp_proof_snippet": f"Quick outcome update: {proof_line} Next step is to keep the monthly audit loop running so repeat issues are caught earlier.",
        "website_proof_block": f"Courier Leak Audit outcome: {proof_line} The process produced a client-readable dispute pack, courier follow-up workflow, and monthly monitoring baseline.",
        "referral_partner_proof": f"A recent audit created a clear courier evidence trail: {proof_line} If your e-commerce clients have rising delivery costs, we can run the same low-risk baseline and share the action pack.",
        "client_follow_up_email": f"Hi {{Name}},\n\nI have attached the dispute/query pack and outcome tracker. The key point is that we are separating courier recovery candidates from operational exposure, so we do not overclaim.\n\nPlease send the dispute-ready/query rows to {courier_provider} and record each response as credited, rejected, or needing more evidence. That outcome history is what makes next month's audit sharper.\n\nCurrent tracked outcome: {proof_line}\n\nKyle",
        "courier_follow_up_email": f"Hi {{Courier contact}},\n\nPlease review the attached waybill-level query pack for {client_name}. The rows are evidence-scored and include the relevant accepted/charged/expected rate context where available.\n\nCould you please confirm for each row whether it will be credited, rejected with reason, or whether you need further evidence?\n\nRegards,\nKyle",
        "follow_up_3_day": "Hi {Name}, checking whether the courier has acknowledged the query pack. If not, I would send a short reminder asking for row-level accepted/rejected/needs-evidence responses.",
        "follow_up_7_day": "Hi {Name}, it has been a week since the dispute/query pack was sent. The important next step is getting a row-level response so we can update credited, rejected, and needs-evidence outcomes.",
        "credit_note_confirmation_request": "Hi {Name}, once the courier issues a credit note, please send the credit note or statement line so I can mark the row as credited and improve the benchmark history.",
        "rejection_learning_request": f"Hi {{Name}}, if {courier_provider} rejects any rows, please ask for the exact reason: surcharge, rate-card rule, service mismatch, fuel/VAT, remote zone, or data issue. Rejections are useful because they teach the next audit what not to overclaim.",
        "monthly_monitoring_conversion": f"The first audit created the evidence baseline. The moat is the monthly loop: audit new exports, query exceptions, record credited/rejected outcomes, and prevent repeat leakage. Recommended next step for {client_name}: monthly monitoring at {monitoring_fee}/mo so courier billing control becomes routine.",
        "proof_metrics": {
            "submitted": format_currency(submitted, currency_prefix),
            "credited": format_currency(credited, currency_prefix),
            "conversion": f"{conversion:.1%}",
            "credited_rows": credited_count,
            "rejected_rows": rejection_count,
        },
    }


def build_proof_pack_text(proof_pack):
    proof_pack = proof_pack or {}
    lines = ["Case Study & Follow-Up Pack", "===========================", ""]
    for key, value in proof_pack.items():
        if key == "proof_metrics":
            lines.extend(["PROOF METRICS"])
            for metric, metric_value in value.items():
                lines.append(f"- {humanize_label(metric)}: {metric_value}")
            lines.append("")
            continue
        lines.extend([humanize_label(key).upper(), str(value), ""])
    return "\n".join(lines).strip() + "\n"


def build_sales_readiness_copy(market_gap_score, leakage_rate, total_leakage, currency_prefix):
    if market_gap_score >= 70:
        return (
            "Strong sales story",
            f"This dataset shows a visible shipping-margin control gap. Lead with the non-overlapping leakage/exposure amount: {format_currency(total_leakage, currency_prefix)} identified and a leakage/exposure rate of {leakage_rate:.1%}.",
        )
    if market_gap_score >= 40:
        return (
            "Good operational improvement story",
            f"This is a practical SME control case. Lead with fewer mistakes, better packaging rules, and monthly monitoring. Non-overlapping leakage/exposure rate: {leakage_rate:.1%}.",
        )
    if safe_float(total_leakage) > 0:
        return (
            "Operational-control story",
            f"This upload has {format_currency(total_leakage, currency_prefix)} in non-overlapping leakage/exposure, but the evidence may be operational rather than immediately recoverable. Lead with control, prevention, and evidence gathering.",
        )
    return (
        "Baseline / trust-building story",
        "This upload is useful as a low-risk baseline. Sell the tool as ongoing insurance rather than a dramatic one-off recovery claim.",
    )


def limited_preview(df, limit=100):
    if df is None or df.empty:
        return pd.DataFrame()
    return df.head(limit).copy()


def render_prepare_download(label, generator, filename, mime_type, key, help_text=None):
    prepare_label = f"Prepare {label}"
    if st.button(prepare_label, key=f"prepare_{key}", help=help_text, use_container_width=True):
        try:
            with st.spinner(f"Preparing {label}..."):
                st.session_state[key] = generator()
            st.success(f"{label} is ready to download.")
        except Exception as exc:
            st.session_state.pop(key, None)
            st.error(f"Could not prepare {label}: {exc}")
    if key in st.session_state:
        st.download_button(
            f"Download {label}",
            st.session_state[key],
            filename,
            mime_type,
            key=f"download_{key}",
            use_container_width=True,
            help=help_text,
        )


def render_step_header(step, title, guidance):
    st.markdown(
        f"""
        <div class="step-card">
        <div class="step-kicker">Step {step}</div>
        <h3>{title}</h3>
        <p>{guidance}</p>
        </div>
        """,
        unsafe_allow_html=True,
    )


def build_client_conversation_guide(client_name, courier_provider, pitch_angle, direct_recovery, operational_exposure, packaging_leak, financial_query, currency_prefix):
    script = build_call_script(client_name, courier_provider, pitch_angle, direct_recovery, operational_exposure, packaging_leak, financial_query, currency_prefix)
    return {
        "say_first": script[0],
        "explain_numbers": script[2],
        "do_not_overclaim": "Do not say all leakage is refundable. Direct recovery candidates require courier confirmation; operational exposure is for prevention and negotiation.",
        "close": script[-1],
        "questions": [
            "Can you send the signed courier rate card and surcharge schedule?",
            "Who currently follows up on courier credits and rejected disputes?",
            "Are packaging choices rule-based or decided manually at dispatch?",
            "Can we track credited and rejected outcomes after submission?",
        ],
    }


st.markdown(
    """
    <style>
    .stApp {background: #ffffff; color: #1f2937;}
    [data-testid="stSidebar"] {background: #f8fafc; border-right: 1px solid #e5e7eb;}
    h1, h2, h3, p, label, span {color: #1f2937;}
    div[data-testid="stMetric"] {
        background: #ffffff;
        border: 1px solid #d1d5db;
        border-radius: 14px;
        padding: 18px;
        box-shadow: 0 8px 24px rgba(15, 23, 42, 0.08);
    }
    div[data-testid="stMetricLabel"] p {color: #374151; font-weight: 700;}
    div[data-testid="stMetricValue"] {color: #111827;}
    .info-card {background: #f9fafb; border: 1px solid #e5e7eb; border-radius: 14px; padding: 18px; color: #1f2937;}
    .gap-card {background: #eef6ff; border: 1px solid #bfdbfe; border-left: 6px solid #2563eb; border-radius: 18px; padding: 22px; color: #0f172a; box-shadow: 0 10px 28px rgba(15, 23, 42, 0.08);}
    .gap-card h3, .gap-card p, .gap-card li, .gap-card b {color: #0f172a;}
    .step-card {background: #ffffff; border: 1px solid #cbd5e1; border-left: 7px solid #2563eb; border-radius: 18px; padding: 18px 20px; margin: 8px 0 18px; box-shadow: 0 8px 22px rgba(15, 23, 42, 0.06);}
    .step-card h3 {margin: 0 0 6px 0; color: #0f172a;}
    .step-card p {margin: 0; color: #334155;}
    .step-kicker {font-size: 0.82rem; font-weight: 800; color: #2563eb; letter-spacing: .06em; text-transform: uppercase; margin-bottom: 4px;}
    .action-card {background: #ffffff; border: 1px solid #bbf7d0; border-left: 6px solid #16a34a; border-radius: 14px; padding: 16px; margin-bottom: 12px; color: #0f172a;}
    .warning-card {background: #fffbeb; border: 1px solid #fde68a; border-left: 6px solid #f59e0b; border-radius: 14px; padding: 16px; margin-bottom: 12px; color: #0f172a;}
    .talk-track-card {background: #f8fafc; border: 1px solid #bfdbfe; border-left: 6px solid #0ea5e9; border-radius: 14px; padding: 16px; margin-bottom: 12px; color: #0f172a;}
    .export-card {background: #f9fafb; border: 1px solid #d1d5db; border-radius: 14px; padding: 14px; margin-bottom: 12px; color: #0f172a;}
    .muted-card {background: #f8fafc; border: 1px solid #e2e8f0; border-radius: 14px; padding: 14px; color: #0f172a;}
    </style>
    """,
    unsafe_allow_html=True,
)

st.title("360° Margin Diagnostic Engine")
st.caption("A light-mode diagnostic dashboard for South African and Botswana e-commerce brands using Bob Go, The Courier Guy, and messy courier exports.")

selected_region = st.sidebar.selectbox("Select Region", ["South Africa", "Botswana"])
REGION_CONFIG = {
    "South Africa": {"currency_symbol": "ZAR", "currency_prefix": "R", "avg_shipping_cost": 120},
    "Botswana": {"currency_symbol": "BWP", "currency_prefix": "P", "avg_shipping_cost": 80},
}
region_config = REGION_CONFIG[selected_region]
currency_symbol = region_config["currency_symbol"]
currency_prefix = region_config["currency_prefix"]
avg_shipping_cost = region_config["avg_shipping_cost"]
st.sidebar.caption(f"Currency: {currency_symbol} | Average shipping cost: {format_currency(avg_shipping_cost, currency_prefix)}")
st.sidebar.divider()

supabase = get_supabase_client()
if supabase is None:
    st.sidebar.info("Supabase is not configured. CRM save/history is disabled; local analysis still works.")

st.sidebar.header("Courier Rate Card Configuration")
courier_provider = st.sidebar.selectbox(
    "Courier Provider",
    ["The Courier Guy (Divisor: 5000)", "Bob Go Aggregated (Divisor: 4000)", "Aramex (Divisor: 5000)", "Skynet", "Custom"],
    help="Select the courier rate-card family. This sets the default volumetric divisor but you can still override it below.",
)
provider_divisors = {
    "The Courier Guy (Divisor: 5000)": 5000,
    "Bob Go Aggregated (Divisor: 4000)": 4000,
    "Aramex (Divisor: 5000)": 5000,
    "Skynet": 5000,
    "Custom": 5000,
}
volumetric_divisor = st.sidebar.number_input(
    "Volumetric Divisor",
    min_value=1000,
    max_value=10000,
    value=provider_divisors[courier_provider],
    step=100,
    help="Courier formula divisor: Length × Width × Height ÷ Divisor = volumetric kilograms. Keep this aligned with the client's actual rate card.",
)
excess_penalty_per_kg = st.sidebar.number_input(
    f"Average Excess Penalty per Kg ({currency_symbol})",
    min_value=0.0,
    value=15.0,
    step=0.5,
    format="%.2f",
    help=f"The estimated {currency_symbol} cost per excess billed kilogram. Used to convert volumetric leakage into {currency_symbol} exposure.",
)
negotiated_divisor = st.sidebar.slider(
    "Packaging Simulation Divisor",
    3000,
    7000,
    int(volumetric_divisor),
    step=100,
    help="Used in the packaging simulation to test negotiated courier terms without removing your custom divisor logic.",
)
st.sidebar.divider()
uploaded_file = st.sidebar.file_uploader(
    "Upload courier shipment CSV, Excel, PDF, or HTML",
    type=["csv", "xlsx", "pdf", "html", "htm"],
    help="Upload messy Bob Go, The Courier Guy, or Aramex HTML exports. The auto-mapper will normalize chaotic column names automatically.",
)
fast_view = st.sidebar.toggle(
    "Fast founder view",
    value=True,
    help="Loads the app faster by showing executive actions first and hiding heavy full-data tables unless you open them.",
)

selected_client_id = None
selected_client_name = "Local Client"
audit_month = "June"
audit_year = 2026
if supabase is not None:
    st.sidebar.divider()
    st.sidebar.header("Cloud CRM")
    try:
        clients = supabase.table("clients").select("id, client_name").order("client_name").execute().data or []
    except Exception as exc:
        clients = []
        st.sidebar.warning(f"Cloud CRM is unavailable; continuing in local mode. {exc}")
    client_options = {client["client_name"]: client["id"] for client in clients}
    selected_client_name = st.sidebar.selectbox("Select Client", list(client_options.keys()) if client_options else ["Local Client"], help="Select the client account for saving this audit snapshot.")
    selected_client_id = client_options.get(selected_client_name)
    audit_month = st.sidebar.text_input("Audit Month", value="June", help="Month label used when saving this audit to the CRM.")
    audit_year = st.sidebar.number_input("Audit Year", min_value=2020, max_value=2100, value=2026, step=1, help="Year label used when saving this audit to the CRM.")

selected_provider = courier_provider.lower()
is_skynet_pipeline = "skynet" in selected_provider
is_standard_pipeline = "courier guy" in selected_provider or "aramex" in selected_provider or "bob go" in selected_provider or "custom" in selected_provider

try:
    file_name = uploaded_file.name if uploaded_file else None
    if is_skynet_pipeline:
        st.info("Routing through dedicated Skynet processing engine...")
        if uploaded_file is None:
            st.error("Upload a Skynet Excel file before running the Skynet financial-only pipeline.")
            st.stop()
        raw_data = process_skynet_file(uploaded_file)
    elif is_standard_pipeline:
        st.info("Routing through standard volumetric processing engine...")
        file_bytes = uploaded_file.getvalue() if uploaded_file else None
        raw_data = load_data(file_name, file_bytes)
        raw_data.attrs["pipeline"] = "generic"
    else:
        st.warning("Please select a supported courier provider from the sidebar.")
        st.stop()
    st.toast(f"Upload/load checkpoint complete: {len(raw_data):,} rows loaded.")
    st.write(f"✅ Upload/load checkpoint complete: {len(raw_data):,} rows loaded from {file_name or 'mock data'}.")
except Exception as exc:
    pipeline_name = "Skynet" if is_skynet_pipeline else "Standard"
    st.error(f"{pipeline_name} Pipeline Error: {exc}")
    st.stop()

with st.expander("Packaging Matrix Configuration", expanded=False):
    packaging_matrix = st.data_editor(
        DEFAULT_PACKAGING_MATRIX,
        num_rows="dynamic",
        use_container_width=True,
        column_config={
            "Package": st.column_config.TextColumn("Package", help="Name of the packaging option used in the repack simulation."),
            "L": st.column_config.NumberColumn("Length (cm)", min_value=0.1, help="Outer package length in centimeters."),
            "W": st.column_config.NumberColumn("Width (cm)", min_value=0.1, help="Outer package width in centimeters."),
            "H": st.column_config.NumberColumn("Height (cm)", min_value=0.1, help="Outer package height in centimeters."),
            "cost": st.column_config.NumberColumn(f"Cost ({currency_symbol})", min_value=0.0, help="Packaging material cost subtracted from savings."),
            "Fragile/Void Fill Required": st.column_config.CheckboxColumn("Fragile/Void Fill Required", help="If checked, the model reserves 15% internal space for protection."),
        },
    )

rate_card_uploads = st.sidebar.file_uploader(
    "Upload client rate card(s) (Excel, CSV, PDF, HTML)",
    type=["csv", "xlsx", "xls", "pdf", "html", "htm"],
    accept_multiple_files=True,
    help="Upload one or more client/courier rate cards. HEL039-style main, local, and regional CSV matrices can be uploaded together.",
)
st.sidebar.caption("Rate cards are optional. Without one, the audit still runs but avoids rate-card-based recovery claims.")
rate_card_source = DEFAULT_RATE_CARD.copy()
rate_card_candidates = []
rate_card_upload_message = "No rate card uploaded."
rate_card_mapped_columns = {}
rate_card_source_summaries = []
normalized_rate_frames = []
seen_rate_card_names = set()
for rate_card_upload in rate_card_uploads or []:
    if rate_card_upload.name in seen_rate_card_names:
        continue
    seen_rate_card_names.add(rate_card_upload.name)
    candidates, upload_message = read_rate_card_upload(rate_card_upload)
    rate_card_candidates.extend({**candidate, "source_name": rate_card_upload.name} for candidate in candidates)
    if candidates:
        best_rate_card = candidates[0]
        normalized_frame, mapped_columns = normalize_rate_card_candidate(best_rate_card, courier_provider, volumetric_divisor, rate_card_upload.name)
        if not normalized_frame.empty:
            normalized_rate_frames.append(normalized_frame)
        rate_card_source_summaries.append({
            "File": rate_card_upload.name,
            "Selected source": best_rate_card["label"],
            "Mode": best_rate_card.get("kind", "flat"),
            "Score": best_rate_card["score"],
            "Signals": ", ".join(best_rate_card["matched"]),
            "Rows": len(normalized_frame),
            "Mapped columns": ", ".join(f"{k} ← {v}" for k, v in mapped_columns.items()) if mapped_columns else "No confident column mapping yet.",
        })
        rate_card_mapped_columns.update(mapped_columns)
    else:
        rate_card_source_summaries.append({"File": rate_card_upload.name, "Selected source": upload_message, "Mode": "unread", "Score": 0, "Signals": "", "Rows": 0, "Mapped columns": ""})
if normalized_rate_frames:
    rate_card_source = pd.concat(normalized_rate_frames, ignore_index=True)
    rate_card_upload_message = f"Loaded {len(normalized_rate_frames):,} rate-card source file(s) with {len(rate_card_source):,} normalized row(s)."
    st.sidebar.success(rate_card_upload_message)
elif rate_card_uploads:
    rate_card_upload_message = "Rate-card uploads were detected, but no usable rows could be normalized."
    st.sidebar.warning(rate_card_upload_message)

with st.expander("Rate Card Lab: upload wizard + expected-rate reconstruction", expanded=False):
    st.warning("Use the client's actual courier agreement/rate card. The wizard helps normalize messy files, but weak/blank rows do not create Tier B claims.")
    st.caption(rate_card_upload_message)
    if rate_card_source_summaries:
        st.dataframe(pd.DataFrame(rate_card_source_summaries), use_container_width=True, hide_index=True)
        hel_files = [summary["File"].lower() for summary in rate_card_source_summaries if "hel039" in summary["File"].lower()]
        hel_buckets = {bucket for bucket in ["main", "local", "regional"] for file_name in hel_files if bucket in file_name}
        if hel_files and hel_buckets != {"main", "local", "regional"}:
            st.warning("HEL039 looks partially loaded. Upload the main, local, and regional files together for complete expected-rate reconstruction.")
    if len(rate_card_candidates) > len(rate_card_source_summaries):
        with st.expander("Other detected tables/sheets", expanded=False):
            st.dataframe(pd.DataFrame([{ "File": c.get("source_name", ""), "Candidate": c["label"], "Mode": c.get("kind", "flat"), "Score": c["score"], "Signals": ", ".join(c["matched"]) } for c in rate_card_candidates]), use_container_width=True, hide_index=True)
    col_template, col_request = st.columns(2)
    with col_template:
        st.download_button(
            "Download blank rate-card template",
            data=generate_rate_card_template_xlsx(courier_provider, volumetric_divisor),
            file_name="client_rate_card_template.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )
    with col_request:
        request_preview = build_rate_card_request_pack({"missing_evidence": ["Signed courier rate card", "Surcharge schedule", "Volumetric divisor and VAT rules"]})
        st.download_button(
            "Download rate-card request email",
            data=request_preview["email"],
            file_name="rate_card_request_email.txt",
            mime="text/plain",
            use_container_width=True,
        )
    rate_card_editor = st.data_editor(
        rate_card_source,
        num_rows="dynamic",
        use_container_width=True,
        column_config={
            "Courier": st.column_config.TextColumn("Courier"),
            "Service_Level": st.column_config.TextColumn("Service Level / All"),
            "Origin_Zone": st.column_config.TextColumn("Origin Zone / All"),
            "Destination_Zone": st.column_config.TextColumn("Destination Zone / All"),
            "Min_Weight_KG": st.column_config.NumberColumn("Min kg", min_value=0.0),
            "Max_Weight_KG": st.column_config.NumberColumn("Max kg", min_value=0.0),
            "Base_Rate_ZAR": st.column_config.NumberColumn(f"Base ({currency_symbol})", min_value=0.0),
            "Per_KG_Rate_ZAR": st.column_config.NumberColumn(f"Per kg ({currency_symbol})", min_value=0.0),
            "Minimum_Charge_ZAR": st.column_config.NumberColumn(f"Minimum ({currency_symbol})", min_value=0.0),
            "Fuel_Surcharge_Pct": st.column_config.NumberColumn("Fuel %", min_value=0.0),
            "VAT_Included": st.column_config.CheckboxColumn("VAT included"),
            "Rounding_Increment_KG": st.column_config.NumberColumn("Round kg", min_value=0.01),
            "Volumetric_Divisor": st.column_config.NumberColumn("Divisor", min_value=1000),
        },
    )
    cleaned_rate_card_preview = clean_rate_card(rate_card_editor, volumetric_divisor)
    rate_card_warnings = validate_rate_card_readiness(cleaned_rate_card_preview)
    rc1, rc2, rc3, rc4, rc5 = st.columns(5)
    rc1.metric("Source files", f"{len(seen_rate_card_names):,}")
    rc2.metric("Usable rate rows", f"{len(cleaned_rate_card_preview):,}")
    rc3.metric("Service levels", f"{cleaned_rate_card_preview['Service_Level'].nunique() if not cleaned_rate_card_preview.empty else 0:,}")
    rc4.metric("Weight bands", f"{len(cleaned_rate_card_preview[['Min_Weight_KG', 'Max_Weight_KG']].drop_duplicates()) if not cleaned_rate_card_preview.empty else 0:,}")
    rc5.metric("Route buckets", f"{cleaned_rate_card_preview['Destination_Zone'].nunique() if not cleaned_rate_card_preview.empty else 0:,}")
    for warning in rate_card_warnings:
        if warning.startswith("Rate card looks usable"):
            st.success(warning)
        else:
            st.info(warning)
    if rate_card_uploads and cleaned_rate_card_preview.empty:
        st.warning("These rate cards were uploaded, but I cannot safely infer weight bands and rates yet. Paste the relevant table below or use the template download.")

if is_skynet_pipeline:
    st.toast("Skynet financial-only pipeline: bypassing legacy cleanup and packaging calculations.")
    st.write("✅ Skynet financial-only pipeline: bypassing legacy dirty-data cleanup, volumetric checks, and packaging calculations.")

    analysis_data = raw_data.copy()
    skipped_rows = pd.DataFrame()
    sku_summary = pd.DataFrame(
        columns=[
            "SKU",
            "Velocity_Class",
            "SKU_Order_Frequency",
            "SKU_Total_Shipping_Cost_ZAR",
            "SKU_Distant_Zone_Lines",
        ]
    )
    for column, default in {
        "Order_ID": analysis_data.get("Tracking_Number", pd.Series([f"SKYNET-{i + 1}" for i in range(len(analysis_data))], index=analysis_data.index)),
        "SKU": "Skynet Shipment",
        "Province": analysis_data.get("Destination", "Unknown"),
        "Billed_Cost_ZAR": analysis_data.get("cost", 0.0),
        "Anomaly_Flag": False,
        "Recoverable_Overcharge_ZAR": 0.0,
        "Packaging_Flag": False,
        "Avoidable_Volumetric_Leak_ZAR": 0.0,
        "Capital_Trap_Flag": False,
        "Is_Multi_Item": False,
    }.items():
        if column not in analysis_data.columns:
            analysis_data[column] = default

    financial_anomaly_rows = calculate_financial_anomalies(analysis_data)
else:
    st.toast("Packaging checkpoint: starting packaging and diagnostic calculations.")
    st.write("⏳ Packaging checkpoint: starting packaging and diagnostic calculations...")
    try:
        analysis_data, skipped_rows, sku_summary = run_triple_pillar_engine(
            raw_data,
            packaging_matrix,
            excess_penalty_per_kg,
            volumetric_divisor,
            negotiated_divisor,
        )
        st.toast(f"Packaging checkpoint complete: {len(analysis_data):,} valid rows analysed.")
        st.write(f"✅ Packaging checkpoint complete: {len(analysis_data):,} valid rows analysed.")
    except Exception as exc:
        st.error(f"Diagnostic engine failed: {exc}")
        st.exception(exc)
        raise

    if analysis_data.empty:
        st.error("No valid rows remained after dirty-data cleanup. Check that the file contains weight, billed volumetric weight, cost, and dimensions.")
        if not skipped_rows.empty:
            st.dataframe(limited_preview(skipped_rows), use_container_width=True, hide_index=True)
        st.stop()

    financial_anomaly_rows = calculate_financial_anomalies(raw_data)

analysis_data, cleaned_rate_card = apply_rate_card_analysis(analysis_data, rate_card_editor, courier_provider, volumetric_divisor)
rate_card_matched_count = int(analysis_data["Rate_Card_Matched"].sum()) if "Rate_Card_Matched" in analysis_data.columns else 0
rate_card_delta_total = safe_float(analysis_data.get("Rate_Card_Delta_ZAR", pd.Series(dtype=float)).sum()) if "Rate_Card_Delta_ZAR" in analysis_data.columns else 0.0

anomaly_rows = analysis_data[analysis_data["Anomaly_Flag"]].sort_values("Recoverable_Overcharge_ZAR", ascending=False)
packaging_rows = analysis_data[analysis_data["Packaging_Flag"]].sort_values("Avoidable_Volumetric_Leak_ZAR", ascending=False)
capital_trap_skus = sku_summary[
    (sku_summary["Velocity_Class"].eq("C"))
    & (sku_summary["SKU_Total_Shipping_Cost_ZAR"] >= (sku_summary["SKU_Total_Shipping_Cost_ZAR"].median() if not sku_summary.empty else 0))
    & (sku_summary["SKU_Distant_Zone_Lines"] > 0)
].sort_values("SKU_Total_Shipping_Cost_ZAR", ascending=False)

findings_df = build_dispute_findings(analysis_data, financial_anomaly_rows, packaging_rows, skipped_rows, courier_provider, currency_prefix)
dispute_ready_findings = findings_df[findings_df["Dispute_Ready"]] if not findings_df.empty else pd.DataFrame()
manual_review_findings = findings_df[findings_df["Manual_Review_Required"]] if not findings_df.empty else pd.DataFrame()
query_findings = findings_df[findings_df["Recovery_Type"].eq(RECOVERY_QUERY)] if not findings_df.empty else pd.DataFrame()

direct_recovery_mask = findings_df["Recovery_Type"].eq(RECOVERY_DIRECT) if not findings_df.empty and "Recovery_Type" in findings_df.columns else pd.Series(False, index=findings_df.index)
query_recovery_mask = findings_df["Recovery_Type"].eq(RECOVERY_QUERY) if not findings_df.empty and "Recovery_Type" in findings_df.columns else pd.Series(False, index=findings_df.index)
leakage_totals = calculate_leakage_totals(analysis_data, findings_df, financial_anomaly_rows, packaging_rows)
pillar_a_loss = leakage_totals["direct_recovery_zar"]
pillar_b_loss = safe_float(packaging_rows["Avoidable_Volumetric_Leak_ZAR"].sum()) if not packaging_rows.empty and "Avoidable_Volumetric_Leak_ZAR" in packaging_rows.columns else 0.0
financial_anomaly_loss = leakage_totals["financial_query_zar"] + leakage_totals["account_manager_query_zar"]
operational_exposure_loss = leakage_totals["operational_exposure_zar"]
avg_confidence_score = safe_float(findings_df["Confidence_Score"].mean()) if not findings_df.empty else 0.0
client_summary_anomalies = pd.concat([financial_anomaly_rows, anomaly_rows], ignore_index=True, sort=False)
capital_trap_count = analysis_data["Capital_Trap_Flag"].sum()
multi_item_orders = analysis_data[analysis_data["Is_Multi_Item"]]["Order_ID"].nunique()
total_direct_leakage = leakage_totals["headline_leakage_zar"]
leakage_spend_base = leakage_totals["spend_base_zar"]
market_gap_score, leakage_rate, packaging_rate, anomaly_rate, sku_count, multi_item_rate = calculate_market_gap_score(
    analysis_data,
    anomaly_rows,
    packaging_rows,
    financial_anomaly_rows,
    sku_summary,
    leakage_totals,
)
priority_actions = build_priority_actions(
    analysis_data,
    anomaly_rows,
    packaging_rows,
    financial_anomaly_rows,
    sku_summary,
    currency_prefix,
)
sales_readiness_label, sales_readiness_copy = build_sales_readiness_copy(
    market_gap_score,
    leakage_rate,
    total_direct_leakage,
    currency_prefix,
)
benchmark_table = build_benchmark_table(analysis_data, findings_df, safe_float(analysis_data.get("Billed_Cost_ZAR", pd.Series(dtype=float)).sum()), skipped_rows)
base_outcome_summary = {"submitted_amount": 0.0, "credited_amount": 0.0, "conversion_rate": 0.0, "rejection_count": 0, "needs_evidence_count": 0, "credited_count": 0, "best_tier": "n/a"}
evidence_readiness = calculate_evidence_readiness(analysis_data, findings_df, skipped_rows, cleaned_rate_card, base_outcome_summary, rate_card_matched_count)

st.sidebar.divider()
consultant_password = st.sidebar.text_input(
    " ",
    type="password",
    placeholder="",
    label_visibility="collapsed",
    key="consultant_mode_password",
)

if consultant_password == "pixie":
    st.sidebar.divider()
    st.sidebar.header("Consultant Module")

    estimated_monthly_spend = (
        pd.to_numeric(analysis_data["Billed_Cost_ZAR"], errors="coerce")
        .fillna(0)
        .sum()
    )
    direct_monthly_leakage = float(pillar_a_loss + leakage_totals["account_manager_query_zar"])
    modelled_monthly_risk = estimated_monthly_spend * 0.05
    annual_margin_risk = modelled_monthly_risk * 12

    intro_offer = build_intro_stage_offer(
        estimated_monthly_spend,
        direct_monthly_leakage,
        operational_exposure_loss,
        currency_prefix,
    )
    recommended_setup = intro_offer["recommended_audit_fee"]
    recommended_retainer = intro_offer["recommended_retainer"]

    st.sidebar.caption(f"Detected offer: {intro_offer['package_name']}")
    st.sidebar.caption(intro_offer["condition"])
    st.sidebar.caption(intro_offer["intro_note"])

    setup_fee = st.sidebar.number_input(
        "Audit / Setup Fee",
        min_value=0,
        value=recommended_setup,
        step=500,
        format="%d",
        key="consultant_setup_fee",
        help="Introduction-stage fee. Keep friction low while collecting outcomes and testimonials.",
    )

    retainer_fee = st.sidebar.number_input(
        "Optional Monthly Monitoring",
        min_value=0,
        value=recommended_retainer,
        step=500,
        format="%d",
        key="consultant_retainer_fee",
        help="Editable monthly monitoring fee for follow-up audits, dispute tracking, and rate-card control.",
    )

    annual_retainer = retainer_fee * 12
    intro_offer["recommended_audit_fee"] = setup_fee
    intro_offer["recommended_retainer"] = retainer_fee

    st.sidebar.subheader("Introduction-Stage Pricing")
    col1, col2 = st.sidebar.columns(2)
    col1.metric("Audit", format_currency_whole(setup_fee, currency_prefix))
    col2.metric("Monitor", f"{format_currency_whole(retainer_fee, currency_prefix)}/mo")

    st.sidebar.metric("Estimated Monthly Courier Spend", format_currency_whole(estimated_monthly_spend, currency_prefix))
    st.sidebar.metric("Observed Direct / Query Leakage", format_currency_whole(direct_monthly_leakage, currency_prefix))
    st.sidebar.metric(
        "Operational Exposure",
        format_currency_whole(operational_exposure_loss, currency_prefix),
        help="Control/prevention opportunity, not automatic refund value.",
    )
    st.sidebar.info(
        f"Market guide: audit {intro_offer['audit_range']}; monitoring {intro_offer['retainer_range']}. "
        f"Selected annual monitoring: {format_currency_whole(annual_retainer, currency_prefix)}/year. "
        f"{intro_offer['success_fee']}"
    )

    if retainer_fee > direct_monthly_leakage and direct_monthly_leakage > 0:
        st.sidebar.warning(
            "Monitoring fee is above observed direct leakage. Pitch this as governance, prevention, and beta monitoring — not immediate monthly savings."
        )


k1, k2, k3, k4 = st.columns(4)
k1.metric("Pillar A: Direct Courier Recovery", format_currency(pillar_a_loss, currency_prefix), f"{len(anomaly_rows):,} rows", help="Evidence-backed money deltas, such as charged/courier-confirmed rate exceeding the accepted rate. Weight-only estimates are separated from this number.")
k2.metric("Pillar B: Packaging Leak", format_currency(pillar_b_loss, currency_prefix), f"{len(packaging_rows):,} rows", help="Avoidable warehouse-side leakage from single-item flyer opportunities and multi-item packaging bloat.")
k3.metric("Financial & Routing", format_currency(financial_anomaly_loss, currency_prefix), f"{len(financial_anomaly_rows):,} rows", help="Surcharge spikes and same-weight shipments with unusually high base charges.")
k4.metric("Pillar C: Capital Traps", f"{int(capital_trap_count):,}", "flagged lines", help="C-Class low-velocity SKUs that also incur high shipping costs to distant or remote zones.")

courier_explanation = get_courier_explanation(courier_provider)
with st.expander(f"Courier cost logic: {courier_explanation['title']}", expanded=False):
    for note in courier_explanation["notes"]:
        st.write(f"• {note}")

command_center_tab, setup_tab, summary_tab, command_tab, operations_tab, conversation_tab, exports_tab, outcome_tab, sales_tab, pipeline_tab, readiness_tab, monitoring_tab = st.tabs([
    "0 Command Center",
    "1 Setup & Load",
    "2 Executive Diagnosis",
    "3 Evidence & Disputes",
    "4 Fix the Operation",
    "5 Client Conversation",
    "6 Exports & Advanced",
    "7 Outcome Tracker",
    "8 Sales Kit",
    "9 Client Pipeline",
    "10 Evidence & SWOT",
    "11 Monthly Monitoring",
])

with command_center_tab:
    render_step_header("0", "Client Command Center", "One operational view: current stage, blockers, next action, assets to send, and the exact message to use.")
    command_outcome_summary = summarize_outcomes(st.session_state.get("outcome_tracker_editor", pd.DataFrame()), currency_prefix)
    command_state = determine_client_stage(
        file_name,
        cleaned_rate_card,
        findings_df,
        dispute_ready_findings,
        query_findings,
        skipped_rows,
        evidence_readiness,
        command_outcome_summary,
        rate_card_matched_count,
    )
    command_message_pack = build_next_action_message(selected_client_name, courier_provider, command_state, currency_prefix)
    cc1, cc2, cc3, cc4 = st.columns(4)
    cc1.metric("Current stage", command_state["stage"])
    cc2.metric("Evidence score", f"{command_state['evidence_score']}/100")
    cc3.metric("Readiness", command_state["readiness_band"])
    cc4.metric("Follow-up", command_state["follow_up_timing"])
    st.markdown(f"""
    <div class="action-card">
    <b>Next action</b><br>{command_state['next_action']}<br><br>
    <b>Why</b><br>{command_state['stage_reason']}<br><br>
    <b>Primary angle</b><br>{command_state['primary_angle']}
    </div>
    """, unsafe_allow_html=True)
    bcol, acol = st.columns(2)
    with bcol:
        st.subheader("Blockers / evidence gaps")
        for blocker in command_state["blockers"]:
            st.write(f"• {blocker}")
    with acol:
        st.subheader("Assets to send now")
        for asset in command_state["assets_to_send"]:
            st.write(f"• {asset}")
    st.subheader(command_message_pack["message_title"])
    msg1, msg2 = st.columns(2)
    with msg1:
        st.markdown("**Email**")
        st.code(command_message_pack["email"], language="text")
    with msg2:
        st.markdown("**WhatsApp**")
        st.code(command_message_pack["whatsapp"], language="text")
    st.markdown("**Internal note**")
    st.code(command_message_pack["internal_note"], language="text")
    st.download_button(
        "Download command center next-action pack",
        data=build_command_center_pack_text(command_state, command_message_pack),
        file_name="client_command_center_next_action.txt",
        mime="text/plain",
        use_container_width=True,
    )

    delivery_pack = build_delivery_pack(command_state, evidence_readiness)
    st.subheader("Client Delivery Pack")
    st.markdown(f"""
    <div class="export-card">
    <b>{delivery_pack['pack_name']}</b><br>
    {delivery_pack['objective']}<br>
    <small><b>Follow-up:</b> {delivery_pack['follow_up_timing']} | <b>Next stage:</b> {delivery_pack['next_stage']}</small>
    </div>
    """, unsafe_allow_html=True)
    dcol1, dcol2, dcol3 = st.columns(3)
    with dcol1:
        st.markdown("**Attach / generate**")
        for item in delivery_pack["attach_or_generate"]:
            st.write(f"[ ] {item}")
    with dcol2:
        st.markdown("**Ask for**")
        for item in delivery_pack["ask_for"]:
            st.write(f"[ ] {item}")
    with dcol3:
        st.markdown("**Warnings**")
        for item in delivery_pack["warnings"]:
            st.write(f"• {item}")
    with st.expander("Delivery checklist", expanded=False):
        for item in delivery_pack["delivery_checklist"]:
            st.write(f"[ ] {item}")
    st.download_button(
        "Download client delivery pack instructions",
        data=build_delivery_pack_text(delivery_pack, command_message_pack),
        file_name="client_delivery_pack_instructions.txt",
        mime="text/plain",
        use_container_width=True,
    )

with setup_tab:
    render_step_header("1", "Setup & Load", "Confirm the courier, upload status, data quality, and advanced rate/packaging inputs before you explain anything to a client.")
    s1, s2, s3, s4 = st.columns(4)
    s1.metric("Raw rows", f"{len(raw_data):,}")
    s2.metric("Valid rows", f"{len(analysis_data):,}")
    s3.metric("Skipped rows", f"{len(skipped_rows):,}")
    s4.metric("Divisor", f"{volumetric_divisor:,}")
    st.info(f"Loaded {file_name or 'mock data'} through the {'Skynet financial-only' if is_skynet_pipeline else 'standard volumetric'} pipeline. Start here when checking whether the audit data is credible enough for a client conversation.")
    key_fields = ["Waybill_ID", "Order_ID", "Actual_Weight_KG", "Billed_Vol_KG", "Billed_Cost_ZAR", "Accepted_Rate_ZAR", "Charged_Rate_ZAR"]
    present_fields = [field for field in key_fields if field in analysis_data.columns]
    missing_fields = [field for field in key_fields if field not in analysis_data.columns]
    st.write(f"**Present fields:** {', '.join(present_fields) if present_fields else 'None'}")
    st.write(f"**Missing fields:** {', '.join(missing_fields) if missing_fields else 'None'}")
    with st.expander("Skipped / dirty rows preview", expanded=False):
        st.dataframe(limited_preview(skipped_rows), use_container_width=True, hide_index=True)


rate_tab = benchmark_tab = gap_tab = anomaly_tab = financial_tab = packaging_tab = velocity_tab = exports_tab

with summary_tab:
    render_step_header("2", "Executive Diagnosis", "Use this screen to understand the client story in plain English before opening detailed tables.")
    st.markdown(
        f"""
        <div class="gap-card">
        <h3>Market gap detected: outsourced shipping-margin control for SMEs</h3>
        <p>Most small e-commerce brands can see their total courier bill, but they cannot see which waybills, package choices, routes or SKUs created the leak. This dashboard turns that invisible gap into a ranked action list.</p>
        <p><b>Sales readiness:</b> {sales_readiness_label}. {sales_readiness_copy}</p>
        </div>
        """,
        unsafe_allow_html=True,
    )

    f1, f2, f3, f4 = st.columns(4)
    f1.metric("Market Gap Score", f"{market_gap_score}/100", help="Scores how strong the client story is based on leakage, packaging issues, anomalies, SKU complexity and data volume.")
    f2.metric("Leakage / Exposure Rate", f"{leakage_rate:.1%}", format_currency(total_direct_leakage, currency_prefix), help=f"Non-overlapping leakage divided by courier spend base of {format_currency(leakage_spend_base, currency_prefix)}. Direct recovery and query amounts are counted once; operational exposure is included only for rows not already counted as direct/query recovery.")
    f3.metric("SKU Complexity", f"{sku_count:,} SKUs", f"{multi_item_rate:.0%} multi-item", help="More SKU and basket complexity means a stronger need for outsourced logistics control.")
    f4.metric("Action Count", f"{len(priority_actions):,}", "ranked fixes", help="The number of practical next steps generated from this upload.")

    st.subheader("Do these first")
    for _, action in priority_actions.head(4).iterrows():
        st.markdown(
            f"""
            <div class="action-card">
            <b>{int(action['Priority'])}. {action['Action']}</b><br>
            <span>{action['Why']}</span><br>
            <small><b>Owner:</b> {action['Owner']} | <b>Next:</b> {action['Next Step']}</small>
            </div>
            """,
            unsafe_allow_html=True,
        )

    leakage_column = f"Leakage_{currency_symbol}"
    chart_data = pd.DataFrame([
        {"Pillar": "Courier Billing", leakage_column: pillar_a_loss},
        {"Pillar": "Financial/Routing", leakage_column: financial_anomaly_loss},
        {"Pillar": "Packaging", leakage_column: pillar_b_loss},
    ])
    st.plotly_chart(px.bar(chart_data, x="Pillar", y=leakage_column, text_auto=".2s", color="Pillar", color_discrete_sequence=["#2563eb", "#16a34a", "#f97316"], template="plotly_white"), use_container_width=True)

    st.subheader("Client Email Summary")
    client_summary_text = generate_client_summary(raw_data, client_summary_anomalies)
    st.info(client_summary_text)
    st.code(client_summary_text, language="text")

    c1, c2, c3 = st.columns(3)
    with c1:
        render_prepare_download(
            "PDF Blueprint",
            lambda: generate_margin_pdf(analysis_data, anomaly_rows, packaging_rows, capital_trap_skus, courier_provider, volumetric_divisor, pillar_a_loss, pillar_b_loss, currency_prefix),
            "Margin_Diagnostic_Recovery_Blueprint.pdf",
            "application/pdf",
            "export_pdf_blueprint",
            "Generates the white-background PDF report only when needed.",
        )
    with c2:
        render_prepare_download(
            "Action Plan Excel",
            lambda: convert_df_to_styled_excel(priority_actions),
            "founder_action_plan.xlsx",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "export_action_plan",
            "Exports ranked next actions only when needed.",
        )
    if supabase is not None:
        if c3.button("Save Audit to CRM", disabled=selected_client_id is None, use_container_width=True, help="Saves the evidence-scored audit values to Supabase for client history tracking."):
            audit_payload = {
                "client_id": selected_client_id,
                "audit_month": audit_month,
                "audit_year": int(audit_year),
                "total_loss_zar": float(pillar_a_loss + pillar_b_loss),
                "dead_stock_value_zar": float(capital_trap_skus["SKU_Total_Shipping_Cost_ZAR"].sum() if not capital_trap_skus.empty else 0),
                "direct_recovery_zar": float(pillar_a_loss),
                "operational_exposure_zar": float(operational_exposure_loss),
                "packaging_leak_zar": float(pillar_b_loss),
                "financial_anomaly_zar": float(financial_anomaly_loss),
                "avg_confidence_score": float(avg_confidence_score),
                "dispute_ready_count": int(len(dispute_ready_findings)),
                "manual_review_count": int(len(manual_review_findings)),
                "leakage_rate": float(leakage_rate),
                "courier_provider": courier_provider,
                "rows_raw": int(len(raw_data)),
                "rows_valid": int(len(analysis_data)),
                "rows_skipped": int(len(skipped_rows)),
            }
            audit_response = supabase.table("audits").insert(audit_payload).execute()
            audit_id = (audit_response.data or [{}])[0].get("id")
            if audit_id and not findings_df.empty:
                finding_payload = []
                for _, finding in findings_df.head(100).iterrows():
                    finding_payload.append({
                        "audit_id": audit_id,
                        "finding_category": safe_text(finding.get("Finding_Category", ""), ""),
                        "evidence_tier": safe_text(finding.get("Evidence_Tier", ""), ""),
                        "confidence_score": float(safe_float(finding.get("Confidence_Score", 0))),
                        "recovery_type": safe_text(finding.get("Recovery_Type", ""), ""),
                        "waybill_id": safe_text(finding.get("Waybill_ID", ""), ""),
                        "order_id": safe_text(finding.get("Order_ID", ""), ""),
                        "direct_recovery_zar": float(safe_float(finding.get("Direct_Recovery_ZAR", 0))),
                        "operational_exposure_zar": float(safe_float(finding.get("Operational_Exposure_ZAR", 0))),
                        "explanation": safe_text(finding.get("Explainability_Text", ""), ""),
                        "recommended_action": safe_text(finding.get("Recommended_Action", ""), ""),
                    })
                if finding_payload:
                    findings_response = supabase.table("audit_findings").insert(finding_payload).execute()
                    st.session_state["last_saved_findings"] = findings_response.data or []
            st.session_state["last_saved_audit_id"] = audit_id
            st.success(f"Saved evidence-scored {audit_month} {audit_year} audit for {selected_client_name}.")

    with st.expander("Advanced exports and dirty-data detail", expanded=not fast_view):
        render_prepare_download(
            "Full Professional Excel",
            lambda: convert_df_to_styled_excel(analysis_data),
            "full_margin_diagnostic.xlsx",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "export_full_professional_excel",
            "Exports every cleaned and calculated row only when needed.",
        )
        st.caption("Rows shown here were not used in pillar calculations because critical numeric data was missing or invalid.")
        st.dataframe(display_currency_columns(limited_preview(skipped_rows), currency_symbol), use_container_width=True, hide_index=True)
        if not fast_view:
            st.subheader("Full Cleaned Diagnostic Data")
            st.dataframe(display_currency_columns(limited_preview(analysis_data, 250), currency_symbol), use_container_width=True, hide_index=True)

with command_tab:
    render_step_header("3", "Evidence & Dispute Pack", "Separate what can be queried now from what must stay as operational exposure or data-quality cleanup.")
    st.caption("Evidence-scored findings split direct recovery from account-manager queries, operational exposure, and data-quality warnings.")

    d1, d2, d3, d4, d5, d6 = st.columns(6)
    d1.metric("Dispute-ready", format_currency(safe_float(dispute_ready_findings.get("Direct_Recovery_ZAR", pd.Series(dtype=float)).sum()), currency_prefix), f"{len(dispute_ready_findings):,} rows")
    d2.metric("Avg confidence", f"{avg_confidence_score:.0f}/100", help="Confidence uses evidence tier, waybill presence, known courier context, and source quality.")
    d3.metric("Tier A claims", f"{int(findings_df['Evidence_Tier'].eq(EVIDENCE_TIERS['A']).sum()) if not findings_df.empty else 0:,}")
    d4.metric("Tier B claims", f"{int(findings_df['Evidence_Tier'].eq(EVIDENCE_TIERS['B']).sum()) if not findings_df.empty else 0:,}")
    d5.metric("Manual review", f"{len(manual_review_findings):,}")
    d6.metric("Rate-card matched", f"{rate_card_matched_count:,}", format_currency(rate_card_delta_total, currency_prefix))

    provenance_summary = build_recovery_provenance_summary(analysis_data, findings_df, currency_prefix)
    sanity_checks = build_direct_recovery_sanity_checks(analysis_data, findings_df)
    st.markdown(f"""
    <div class="talk-track-card">
    <b>Direct Recovery Sanity Check</b><br>
    {provenance_summary['verdict']}<br>
    <small>{provenance_summary['formula']} Amounts remain candidates until courier confirmation.</small>
    </div>
    """, unsafe_allow_html=True)
    sc1, sc2, sc3, sc4 = st.columns(4)
    sc1.metric("Direct candidates", format_currency(provenance_summary["direct_total"], currency_prefix), f"{provenance_summary['direct_rows']:,} rows")
    sc2.metric("Operational separated", format_currency(provenance_summary["operational_total"], currency_prefix), f"{provenance_summary['operational_rows']:,} rows")
    sc3.metric("Tier B rate-card", format_currency(provenance_summary["tier_b_total"], currency_prefix), f"{provenance_summary['tier_b_rows']:,} rows")
    sc4.metric("Money-field rows", f"{provenance_summary['direct_with_money_fields']:,}", "accepted + charged")
    st.dataframe(sanity_checks, use_container_width=True, hide_index=True)

    with st.expander("Ingestion quality and confidence methodology", expanded=True):
        raw_count = len(raw_data) if raw_data is not None else 0
        valid_count = len(analysis_data) if analysis_data is not None else 0
        skipped_count = len(skipped_rows) if skipped_rows is not None else 0
        skipped_rate = skipped_count / raw_count if raw_count else 0
        key_fields = ["Waybill_ID", "Order_ID", "Actual_Weight_KG", "Billed_Vol_KG", "Billed_Cost_ZAR", "Accepted_Rate_ZAR", "Charged_Rate_ZAR"]
        present_fields = [field for field in key_fields if field in analysis_data.columns]
        missing_fields = [field for field in key_fields if field not in analysis_data.columns]
        iq1, iq2, iq3, iq4 = st.columns(4)
        iq1.metric("Raw rows", f"{raw_count:,}")
        iq2.metric("Valid rows", f"{valid_count:,}")
        iq3.metric("Skipped rows", f"{skipped_count:,}", f"{skipped_rate:.1%}")
        iq4.metric("Mapped key fields", f"{len(present_fields)}/{len(key_fields)}")
        st.write(f"**Present fields:** {', '.join(present_fields) if present_fields else 'None'}")
        st.write(f"**Missing fields:** {', '.join(missing_fields) if missing_fields else 'None'}")
        st.write("**Evidence tiers:** A = direct money proof; B = rate-card reproducible; C = statistical/same-lane query; D = operational estimate; E = data-quality warning.")
        if findings_df.empty or not findings_df["Dispute_Ready"].any():
            st.warning("No dispute-ready Tier A claims were found. Treat any weight/dimension values as operational improvement evidence until a rate card or courier explanation supports recovery.")

    if findings_df.empty:
        st.info("No findings were generated for this upload.")
    else:
        filter_cols = st.columns(4)
        tier_values = sorted(findings_df["Evidence_Tier"].dropna().unique().tolist())
        recovery_values = sorted(findings_df["Recovery_Type"].dropna().unique().tolist())
        tier_options = {"All": "All"} | {humanize_label(value): value for value in tier_values}
        recovery_options = {"All": "All"} | {humanize_label(value): value for value in recovery_values}
        selected_tier_label = filter_cols[0].selectbox("Evidence tier", list(tier_options.keys()))
        selected_recovery_label = filter_cols[1].selectbox("Recovery type", list(recovery_options.keys()))
        selected_tier = tier_options[selected_tier_label]
        selected_recovery = recovery_options[selected_recovery_label]
        min_confidence = filter_cols[2].slider("Minimum confidence", 0, 100, 0, 5)
        dispute_only = filter_cols[3].checkbox("Dispute-ready only", value=False)

        filtered_findings = findings_df.copy()
        if selected_tier != "All":
            filtered_findings = filtered_findings[filtered_findings["Evidence_Tier"].eq(selected_tier)]
        if selected_recovery != "All":
            filtered_findings = filtered_findings[filtered_findings["Recovery_Type"].eq(selected_recovery)]
        filtered_findings = filtered_findings[filtered_findings["Confidence_Score"] >= min_confidence]
        if dispute_only:
            filtered_findings = filtered_findings[filtered_findings["Dispute_Ready"]]

        display_cols = [
            "Finding_Category", "Evidence_Tier", "Confidence_Score", "Recovery_Type", "Dispute_Ready",
            "Direct_Recovery_ZAR", "Operational_Exposure_ZAR", "Expected_Rate_ZAR", "Rate_Card_Delta_ZAR",
            "Rate_Card_Confidence", "Waybill_ID", "Order_ID", "Service_Level",
            "Reason", "Recommended_Action", "Explainability_Text",
        ]
        display_findings = filtered_findings[[col for col in display_cols if col in filtered_findings.columns]]
        st.dataframe(limited_preview(prepare_client_facing_findings(display_findings, currency_symbol), 250), use_container_width=True, hide_index=True)
        render_prepare_download(
            "Evidence-Scored Dispute Pack",
            lambda: generate_dispute_pack_xlsx(findings_df, currency_symbol, currency_prefix),
            "evidence_scored_dispute_pack.xlsx",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "export_evidence_dispute_pack",
            "Prepares a client-readable workbook that separates dispute-ready claims, account-manager queries, operational exposure, data-quality exceptions, and methodology as separate sheets.",
        )

with rate_tab:
    render_step_header("6", "Exports & Advanced Data", "Use this final step for detailed analyst views, rate-card diagnostics, benchmarks, and export preparation. Heavy files are generated only when requested.")
    st.subheader("Rate Card Lab")
    st.caption("Load or edit a client-specific rate card to reconstruct expected charges and create Tier B rate-card-reproducible findings.")
    if cleaned_rate_card.empty:
        st.info("No usable client rate card rows are active. The app will not create Tier B claims from blank/default templates.")
    else:
        r1, r2, r3, r4 = st.columns(4)
        r1.metric("Usable rate rows", f"{len(cleaned_rate_card):,}")
        r2.metric("Matched shipments", f"{rate_card_matched_count:,}")
        r3.metric("Positive deltas", f"{int((analysis_data['Rate_Card_Delta_ZAR'] > 0).sum()) if 'Rate_Card_Delta_ZAR' in analysis_data.columns else 0:,}")
        r4.metric("Rate-card delta", format_currency(rate_card_delta_total, currency_prefix))
        st.subheader("Active Rate Card")
        st.dataframe(cleaned_rate_card, use_container_width=True, hide_index=True)
        st.subheader("Matched Shipment Sample")
        rate_cols = ["Order_ID", "Waybill_ID", "Service_Level", "Billed_Vol_KG", "Billed_Cost_ZAR", "Expected_Rate_ZAR", "Rate_Card_Delta_ZAR", "Rate_Card_Confidence", "Rate_Card_Reason"]
        matched_sample = analysis_data[analysis_data["Rate_Card_Matched"]].head(100) if "Rate_Card_Matched" in analysis_data.columns else pd.DataFrame()
        st.dataframe(display_currency_columns(matched_sample[[col for col in rate_cols if col in matched_sample.columns]], currency_symbol), use_container_width=True, hide_index=True)
    st.warning("Tier B findings are only as strong as the uploaded client rate card. Verify effective dates, VAT mode, fuel surcharge, and account-specific exceptions before submitting to a courier.")

with benchmark_tab:
    st.subheader("Benchmark & Outcome Intelligence")
    st.caption("Internal SME heuristic benchmarks until enough anonymized client history exists. This turns one upload into a repeatable market-positioning signal.")
    if benchmark_table.empty:
        st.info("No benchmark metrics available for this upload.")
    else:
        status_counts = benchmark_table["Benchmark_Status"].value_counts().to_dict()
        b1, b2, b3, b4 = st.columns(4)
        b1.metric("Urgent", f"{status_counts.get('urgent', 0):,}")
        b2.metric("Material", f"{status_counts.get('material', 0):,}")
        b3.metric("Monitor", f"{status_counts.get('monitor', 0):,}")
        b4.metric("Controlled", f"{status_counts.get('controlled', 0):,}")
        st.dataframe(benchmark_table[["Metric", "Display", "Benchmark_Status", "Recommendation", "Benchmark_Context"]], use_container_width=True, hide_index=True)
        st.info("Next moat step: save dispute outcomes and credited amounts so these internal heuristic bands become proprietary recovery benchmarks by courier, industry, and order-volume segment.")

with gap_tab:
    st.subheader("Market Gap Analysis")
    st.markdown(
        """
        <div class="info-card">
        <b>Marketing logic used:</b> The best gap is not another generic courier dashboard. It is a concentrated, founder-friendly control system for SMEs that do not have a logistics analyst. The notes support this through segmentation, targeting, positioning, relationship marketing, measurable segments, accessibility, and B2B segmentation by firm size, location and buying situation.
        </div>
        """,
        unsafe_allow_html=True,
    )
    g1, g2, g3 = st.columns(3)
    g1.metric("Segment", "SME e-commerce", "accessible + measurable")
    g2.metric("Positioning", "Shipping-margin control", "not courier software")
    g3.metric("Differentiator", "Action plan first", "not raw analytics")
    st.dataframe(priority_actions, use_container_width=True, hide_index=True)
    st.markdown(
        f"""
        <div class="muted-card">
        <b>How to sell this result:</b> Tell the prospect that the audit found a {leakage_rate:.1%} combined direct-and-operational leakage rate and a {market_gap_score}/100 market-gap score. Separate direct courier recovery from packaging/weight exposure so the client sees what can be disputed now versus what needs operational follow-up.
        </div>
        """,
        unsafe_allow_html=True,
    )

with operations_tab:
    render_step_header("4", "Fix the Operation", "Turn the audit into practical actions: courier queries, packaging fixes, financial/routing questions, and SKU decisions.")
    st.subheader("Priority Actions")
    st.dataframe(priority_actions, use_container_width=True, hide_index=True)

with anomaly_tab:
    st.subheader("Pillar A: Direct Invoice & Billing Anomalies")
    with st.expander("ELI5: What does this mean and why should I care?", expanded=True):
        st.write("This tab is the dispute list only where the file gives direct money evidence, such as accepted rate versus charged or courier-confirmed rate. Weight-only differences are still tracked, but they belong in operational exposure until the courier rate card confirms a refundable amount.")
    cols = ["Order_ID", "SKU", "Province", "Actual_Weight_KG", "Billed_Vol_KG", "Accepted_Rate_ZAR", "Charged_Rate_ZAR", "Rate_Delta_ZAR", "Anomaly_Reason", "Courier_Recoverable_ZAR", "Billed_Cost_ZAR"]
    st.dataframe(display_currency_columns(limited_preview(anomaly_rows[[col for col in cols if col in anomaly_rows.columns]], 250), currency_symbol), use_container_width=True, hide_index=True)
    render_prepare_download(
        "Courier Dispute Pack",
        lambda: generate_dispute_pack_xlsx(findings_df, currency_symbol, currency_prefix),
        "courier_dispute_pack.xlsx",
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "export_courier_dispute_pack",
        "Prepares a professional Excel dispute pack for direct courier recovery rows.",
    )

with financial_tab:
    st.subheader("Financial & Routing Anomalies")
    with st.expander("ELI5: What does this mean and why should I care?", expanded=True):
        st.write("This tab catches invoice glitches that are not weight mistakes: unusual surcharge spikes and shipments with the same weight but very different base charges. These are good candidates for account-manager queries even when the physics audit shows R0.00.")
    financial_cols = [
        "Tracking_Number",
        "Destination",
        "Shipper_Ref",
        "Order_ID",
        "Waybill_ID",
        "Financial_Weight_KG",
        "Financial_Base_Charge_ZAR",
        "Financial_Other_Charges_ZAR",
        "Financial_Excess_ZAR",
        "Financial_Anomaly_Reason",
    ]
    st.dataframe(display_currency_columns(limited_preview(financial_anomaly_rows[[col for col in financial_cols if col in financial_anomaly_rows.columns]], 250), currency_symbol), use_container_width=True, hide_index=True)
    render_prepare_download(
        "Financial Anomalies Excel",
        lambda: convert_df_to_styled_excel(financial_anomaly_rows),
        "financial_routing_anomalies.xlsx",
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "export_financial_anomalies",
        "Prepares surcharge spikes and routing/base-charge anomalies as a styled Excel workbook.",
    )

with packaging_tab:
    st.subheader("Pillar B: Workflow & Packaging Inefficiency")
    with st.expander("ELI5: What does this mean and why should I care?", expanded=True):
        st.write("This tab separates single-item mistakes from mixed-basket mistakes. A single t-shirt in a big box should become a flyer rule. For multi-item orders, we do not guess a box; we check whether the courier's billed size is bigger than all products combined plus a fair 20% packing buffer.")
    cols = ["Order_ID", "SKU", "Is_Multi_Item", "Packaging_Reason", "Recommended_Package", "Billed_Vol_KG", "Optimized_Vol_KG", "Avoidable_Volumetric_Leak_ZAR"]
    st.dataframe(display_currency_columns(limited_preview(packaging_rows[[col for col in cols if col in packaging_rows.columns]], 250), currency_symbol), use_container_width=True, hide_index=True)
    if not packaging_rows.empty:
        avoidable_column = f"Avoidable_Volumetric_Leak_{currency_symbol}"
        by_reason = display_currency_columns(packaging_rows, currency_symbol).groupby("Packaging_Reason", as_index=False)[avoidable_column].sum()
        st.plotly_chart(px.pie(by_reason, names="Packaging_Reason", values=avoidable_column, color_discrete_sequence=px.colors.qualitative.Safe, template="plotly_white"), use_container_width=True)
    render_prepare_download(
        "Packaging Leak Excel",
        lambda: convert_df_to_styled_excel(packaging_rows),
        "packaging_leak_rows.xlsx",
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "export_packaging_leak",
        "Prepares single-item flyer opportunities and multi-item packaging bloat rows as a styled Excel workbook.",
    )

with conversation_tab:
    render_step_header("5", "Client Conversation", "This is the guided talk track: what to say first, what not to overclaim, what to ask, and what to send after the call.")
    conversation_angle = choose_pitch_angle(findings_df, operational_exposure_loss, pillar_a_loss, skipped_rows, benchmark_table)
    conversation_guide = build_client_conversation_guide(selected_client_name, courier_provider, conversation_angle, pillar_a_loss, operational_exposure_loss, pillar_b_loss, financial_anomaly_loss, currency_prefix)
    st.markdown(f"""
    <div class="talk-track-card"><b>Evidence posture for this call</b><br>{evidence_readiness['band']} — {evidence_readiness['recommended_next_action']}</div>
    <div class="talk-track-card"><b>What to say first</b><br>{conversation_guide['say_first']}</div>
    <div class="warning-card"><b>What not to overclaim</b><br>{conversation_guide['do_not_overclaim']}</div>
    <div class="talk-track-card"><b>Explain the numbers</b><br>{conversation_guide['explain_numbers']}</div>
    <div class="action-card"><b>Close the call</b><br>{conversation_guide['close']}</div>
    """, unsafe_allow_html=True)
    st.subheader("Discovery questions")
    for question in conversation_guide["questions"]:
        st.write(f"• {question}")
    st.subheader("Objection responses")
    for objection, response in build_objection_responses(currency_prefix):
        st.write(f"**{objection}** {response}")
    if consultant_password != "pixie":
        st.info("Enter the consultant password in the sidebar to unlock pricing, offer builder, pitch text, and client memo exports.")

if consultant_password == "pixie":
    with conversation_tab:
        st.subheader("Automated Pitch Engine")
        pitch_company_name = st.text_input(
            "Company Name",
            value=selected_client_name if selected_client_name != "Local Client" else "",
            key="pitch_company_name",
        )
        pitch_founder_name = st.text_input(
            "Founder Name",
            key="pitch_founder_name",
        )
        pitch_brand_context = st.text_area(
            "Brand Context",
            placeholder="e.g., Premium athleisure brand dominating the Cape Town market with sustainable materials...",
            key="pitch_brand_context",
        )

        worst_sku = "your highest-risk SKU"
        if not capital_trap_skus.empty and "SKU" in capital_trap_skus.columns:
            worst_sku = safe_text(capital_trap_skus.iloc[0]["SKU"], worst_sku)
        elif not packaging_rows.empty and "SKU" in packaging_rows.columns:
            worst_sku = safe_text(packaging_rows.iloc[0]["SKU"], worst_sku)

        total_recoverable = safe_float(pillar_a_loss)
        total_avoidable = safe_float(pillar_b_loss)
        total_operational = safe_float(analysis_data.get("Operational_Exposure_ZAR", pd.Series(dtype=float)).sum()) if "Operational_Exposure_ZAR" in analysis_data.columns else 0.0
        annualized_loss = (total_recoverable + total_avoidable) * 12
        company_for_pitch = pitch_company_name.strip() or "your company"
        founder_for_pitch = pitch_founder_name.strip() or "there"
        context_for_pitch = pitch_brand_context.strip() or "the brand positioning and operational momentum you are building"

        pitch_text = f"""Hi {founder_for_pitch},

I just finished running the historical logistics audit on the dataset you sent over. First off, I love what {company_for_pitch} is doing—{context_for_pitch} is a massive differentiator right now.

I wanted to get straight to the numbers. The engine flagged two major structural leaks in your current shipping setup:

1. Direct Courier Recovery: I isolated {format_currency(total_recoverable, currency_prefix)} in evidence-backed accepted-vs-charged rate differences. I've attached the Dispute Pack for the account manager to review.
2. Packaging Bloat: We found that oversized packaging, specifically on orders containing the {worst_sku}, is triggering an additional {format_currency(total_avoidable, currency_prefix)} in unnecessary volumetric penalties.

The model also shows {format_currency(total_operational, currency_prefix)} in weight/dimension operational exposure. I would use that for rate-card and packaging conversations, not as an automatic refund claim. If we annualize the direct recovery plus packaging figures, {company_for_pitch} is quietly losing over {format_currency(annualized_loss, currency_prefix)} this year just to un-optimized dispatch logistics.

I’d love to jump on a quick 15-minute Google Meet this week to walk you through the visual dashboard and show you exactly how to plug these holes.

Let me know what day works best for you."""

        st.code(pitch_text, language="text")

        st.subheader("Courier Leak Audit Offer Builder")
        st.markdown(
            f"""
            <div class="action-card">
            <b>{intro_offer['package_name']}</b><br>
            <span>{intro_offer['positioning']}</span><br>
            <small><b>Audit:</b> {format_currency_whole(setup_fee, currency_prefix)} | <b>Monitoring:</b> {format_currency_whole(retainer_fee, currency_prefix)}/mo | <b>Success fee:</b> {intro_offer['success_fee']}</small><br>
            <small><b>CTA:</b> {intro_offer['cta']}</small>
            </div>
            """,
            unsafe_allow_html=True,
        )
        st.code(
            f"Offer: {intro_offer['package_name']}\n"
            f"Audit fee: {format_currency_whole(setup_fee, currency_prefix)}\n"
            f"Monthly monitoring: {format_currency_whole(retainer_fee, currency_prefix)}/mo\n"
            f"Success fee: {intro_offer['success_fee']}\n"
            f"CTA: {intro_offer['cta']}",
            language="text",
        )

        render_prepare_download(
            "Client Audit Memo",
            lambda: generate_client_audit_memo_pdf(company_for_pitch, courier_provider, analysis_data, findings_df, benchmark_table, priority_actions, currency_prefix, audit_month, audit_year, intro_offer),
            "client_audit_memo.pdf",
            "application/pdf",
            "export_client_audit_memo",
            "Prepares the client-facing commercial memo with cautious next steps and introduction-stage pricing.",
        )
        render_prepare_download(
            "Client Briefing Playbook",
            lambda: generate_client_training_brief_pdf(company_for_pitch, courier_provider, analysis_data, findings_df, benchmark_table, priority_actions, anomaly_rows, packaging_rows, financial_anomaly_rows, capital_trap_skus, currency_prefix, audit_month, audit_year),
            "client_briefing_playbook.pdf",
            "application/pdf",
            "export_client_briefing_playbook",
            "Prepares the internal consultant playbook only when needed.",
        )
with velocity_tab:
    st.subheader("Pillar C: SKU Velocity & Trapped Capital")
    with st.expander("ELI5: What does this mean and why should I care?", expanded=True):
        st.write("This tab tells you which products are slow movers. If a slow product also costs a lot to ship to far-away zones, it traps cash in inventory and courier fees. These SKUs should be bundled, liquidated, or reordered less often.")
    v1, v2, v3 = st.columns(3)
    v1.metric("A-Class SKUs", f"{sku_summary[sku_summary['Velocity_Class'].eq('A')]['SKU'].nunique():,}", help="Highest-frequency SKUs. These usually deserve stock protection and fast fulfillment.")
    v2.metric("B-Class SKUs", f"{sku_summary[sku_summary['Velocity_Class'].eq('B')]['SKU'].nunique():,}", help="Middle-frequency SKUs. Monitor these for movement up or down.")
    v3.metric("C-Class SKUs", f"{sku_summary[sku_summary['Velocity_Class'].eq('C')]['SKU'].nunique():,}", help="Bottom 30% of SKUs by order frequency. These may be dead-stock or slow-moving inventory.")
    st.dataframe(display_currency_columns(limited_preview(sku_summary, 250), currency_symbol), use_container_width=True, hide_index=True)
    if not sku_summary.empty:
        shipping_cost_column = f"SKU_Total_Shipping_Cost_{currency_symbol}"
        sku_chart_data = display_currency_columns(sku_summary, currency_symbol)
        st.plotly_chart(px.treemap(sku_chart_data, path=["Velocity_Class", "SKU"], values=shipping_cost_column, color="SKU_Order_Frequency", color_continuous_scale="Blues", template="plotly_white"), use_container_width=True)
    render_prepare_download(
        "SKU Velocity Excel",
        lambda: convert_df_to_styled_excel(sku_summary),
        "sku_velocity_matrix.xlsx",
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "export_sku_velocity",
        "Prepares ABC velocity classes and shipping-cost exposure by SKU as a styled Excel workbook.",
    )

with outcome_tab:
    render_step_header("7", "Outcome Tracker", "This is how the tool becomes a moat: record what the courier accepts, rejects, credits, or asks for after you send the dispute pack.")
    outcome_template = build_outcome_tracker_template(findings_df, currency_prefix)
    if outcome_template.empty:
        st.info("No direct recovery or account-manager query findings are available for outcome tracking yet.")
    else:
        st.markdown(
            """
            <div class="talk-track-card">
            <b>Learning loop</b><br>
            Tier A credited outcomes become case-study evidence. Tier B outcomes improve rate-card confidence. Tier C rejections teach what account managers will not accept. Tier D remains prevention, not refund tracking.
            </div>
            """,
            unsafe_allow_html=True,
        )
        edited_outcomes = st.data_editor(
            limited_preview(outcome_template, 250),
            use_container_width=True,
            hide_index=True,
            num_rows="dynamic",
            column_config={
                "Submission_Status": st.column_config.SelectboxColumn("Submission Status", options=OUTCOME_STATUSES),
                "Credited_Amount_ZAR": st.column_config.NumberColumn(f"Credited Amount ({currency_symbol})", min_value=0.0),
                "Case_Study_Permission": st.column_config.SelectboxColumn("Case Study Permission", options=["No", "Yes"]),
                "Testimonial_Permission": st.column_config.SelectboxColumn("Testimonial Permission", options=["No", "Yes"]),
                "Follow_Up_Date": st.column_config.TextColumn("Follow-up Date"),
            },
            key="outcome_tracker_editor",
        )
        outcome_summary = summarize_outcomes(edited_outcomes, currency_prefix)
        o1, o2, o3, o4, o5 = st.columns(5)
        o1.metric("Submitted", format_currency(outcome_summary["submitted_amount"], currency_prefix))
        o2.metric("Credited", format_currency(outcome_summary["credited_amount"], currency_prefix))
        o3.metric("Conversion", f"{outcome_summary['conversion_rate']:.1%}")
        o4.metric("Rejected", f"{outcome_summary['rejection_count']:,}")
        o5.metric("Needs evidence", f"{outcome_summary['needs_evidence_count']:,}")
        render_prepare_download(
            "Outcome Tracker Excel",
            lambda: convert_df_to_styled_excel(edited_outcomes),
            "courier_outcome_tracker.xlsx",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "export_outcome_tracker",
            "Exports the editable outcome tracker so courier responses and credits can be tracked after submission.",
        )
        proof_offer = intro_offer if consultant_password == "pixie" else build_intro_stage_offer(
            safe_float(analysis_data.get("Billed_Cost_ZAR", pd.Series(dtype=float)).sum()),
            pillar_a_loss,
            operational_exposure_loss,
            currency_prefix,
        )
        proof_pack = build_proof_asset_pack(selected_client_name, courier_provider, findings_df, edited_outcomes, outcome_summary, currency_prefix, proof_offer)
        st.subheader("Case study draft")
        case_study = build_case_study_snapshot(selected_client_name, outcome_summary, currency_prefix, edited_outcomes)
        st.code(case_study, language="text")
        with st.expander("Proof Builder & Follow-Up Engine", expanded=True):
            st.markdown("**Anonymized case study**")
            st.code(proof_pack["anonymized_case_study"], language="text")
            st.markdown("**LinkedIn proof post**")
            st.code(proof_pack["linkedin_proof_post"], language="text")
            st.markdown("**Client follow-up email after dispute pack**")
            st.code(proof_pack["client_follow_up_email"], language="text")
            st.markdown("**Courier follow-up email**")
            st.code(proof_pack["courier_follow_up_email"], language="text")
            st.markdown("**Monthly monitoring conversion message**")
            st.code(proof_pack["monthly_monitoring_conversion"], language="text")
            st.download_button(
                "Download case study and follow-up pack",
                data=build_proof_pack_text(proof_pack),
                file_name="case_study_and_followup_pack.txt",
                mime="text/plain",
                use_container_width=True,
            )
        if st.session_state.get("last_saved_audit_id"):
            st.success(f"Last saved audit available for cloud outcome linking: {st.session_state['last_saved_audit_id']}")
        else:
            st.caption("Cloud persistence is optional. This tracker works locally even when Supabase is not configured or the audit has not been saved.")

with sales_tab:
    render_step_header("8", "Sales Kit", "Use this to turn the audit into marketable assets: landing copy, outreach messages, referral partner pitch, discovery script, and an offer sheet.")
    sales_outcome_summary = summarize_outcomes(st.session_state.get("outcome_tracker_editor", outcome_template if 'outcome_template' in locals() else pd.DataFrame()), currency_prefix)
    sales_kit = build_sales_kit(
        selected_client_name,
        courier_provider,
        pillar_a_loss,
        operational_exposure_loss,
        pillar_b_loss,
        financial_anomaly_loss,
        currency_prefix,
        intro_offer if consultant_password == "pixie" else build_intro_stage_offer(safe_float(analysis_data.get("Billed_Cost_ZAR", pd.Series(dtype=float)).sum()), pillar_a_loss, operational_exposure_loss, currency_prefix),
        sales_outcome_summary,
    )
    st.markdown(f"""
    <div class="gap-card">
    <h3>{sales_kit['headline']}</h3>
    <p>{sales_kit['subheadline']}</p>
    <p><b>Audit promise:</b> Send us your courier export and rate card. We separate direct recovery candidates from operational leakage within 48 hours.</p>
    <p><b>Proof angle:</b> {sales_kit['proof_line']}</p>
    <p><b>Readiness badge:</b> {evidence_readiness['band']}</p>
    </div>
    """, unsafe_allow_html=True)

    with st.expander("Prospect Qualifier & First Pitch", expanded=True):
        q1, q2 = st.columns(2)
        prospect_company = q1.text_input("Prospect company", value=safe_text(selected_client_name, "Prospect Brand"), key="qualifier_company")
        prospect_name = q2.text_input("Contact name", value="there", key="qualifier_contact")
        q3, q4, q5 = st.columns(3)
        prospect_spend = q3.number_input(f"Monthly courier spend ({currency_symbol})", min_value=0.0, value=float(safe_float(analysis_data.get("Billed_Cost_ZAR", pd.Series(dtype=float)).sum())), step=1000.0, key="qualifier_spend")
        prospect_orders = q4.number_input("Monthly order volume", min_value=0, value=int(max(len(analysis_data), 100)), step=50, key="qualifier_orders")
        prospect_couriers = q5.number_input("Number of couriers", min_value=1, max_value=10, value=1, step=1, key="qualifier_couriers")
        q6, q7, q8 = st.columns(3)
        has_export_access = q6.checkbox("Has export/invoice access", value=True, key="qualifier_export_access")
        has_rate_card = q6.checkbox("Has signed rate card", value=not cleaned_rate_card.empty, key="qualifier_rate_card")
        uses_aggregator = q7.checkbox("Uses aggregator/Bob Go", value="bob go" in courier_provider.lower(), key="qualifier_aggregator")
        has_dispute_process = q7.checkbox("Has dispute process", value=False, key="qualifier_dispute_process")
        has_packaging_rules = q8.checkbox("Has packaging rules", value=False, key="qualifier_packaging_rules")
        pain_level = q8.slider("Pain level", 0, 5, 3, key="qualifier_pain")
        urgency = st.slider("Urgency", 0, 5, 3, key="qualifier_urgency")
        qualification = score_prospect_fit(prospect_spend, prospect_orders, prospect_couriers, has_rate_card, has_export_access, uses_aggregator, has_dispute_process, has_packaging_rules, pain_level, urgency)
        prospect_outreach = build_prospect_outreach_from_score(prospect_name, prospect_company, qualification, sales_kit, currency_prefix)
        qs1, qs2, qs3 = st.columns(3)
        qs1.metric("Prospect score", f"{qualification['score']}/100")
        qs2.metric("Band", qualification["band"])
        qs3.metric("Pitch angle", qualification["pitch_angle"])
        st.write(qualification["why"])
        st.markdown("**Risks / watch-outs**")
        for risk in qualification["risks"]:
            st.write(f"• {risk}")
        st.markdown(f"**Recommended first ask:** {qualification['first_data_ask']}")
        st.markdown(f"**Offer hint:** {qualification['best_offer_hint']}")
        st.info(qualification["moat_note"])
        st.markdown("**First message**")
        st.code(prospect_outreach["first_message"], language="text")
        st.markdown("**Email**")
        st.code(f"Subject: {prospect_outreach['email_subject']}\n\n{prospect_outreach['email_body']}", language="text")
        st.markdown("**Follow-up**")
        st.code(prospect_outreach["follow_up"], language="text")
        st.download_button(
            "Download prospect qualification pitch pack",
            data=build_prospect_qualification_pack_text(qualification, prospect_outreach),
            file_name="prospect_qualification_pitch_pack.txt",
            mime="text/plain",
            use_container_width=True,
        )

    c1, c2 = st.columns(2)
    with c1:
        st.subheader("Landing page copy")
        st.code(
            f"Headline: {sales_kit['headline']}\n\n"
            f"Subheadline: {sales_kit['subheadline']}\n\n"
            f"Pain bullets:\n- " + "\n- ".join(sales_kit["pain_bullets"]) + "\n\n"
            f"What they get:\n- " + "\n- ".join(sales_kit["deliverables"]) + "\n\n"
            f"CTA: {sales_kit['offer_sheet']['cta']}",
            language="text",
        )
        st.subheader("Referral partner pitch")
        st.code(sales_kit["referral_pitch"], language="text")
    with c2:
        st.subheader("LinkedIn DM")
        st.code(sales_kit["linkedin_dm"], language="text")
        st.subheader("WhatsApp")
        st.code(sales_kit["whatsapp"], language="text")
        st.subheader("Email")
        st.write("**Subject options:**")
        for subject in sales_kit["email_subjects"]:
            st.write(f"• {subject}")
        st.code(sales_kit["email_body"], language="text")

    with st.expander("What I do in plain English", expanded=True):
        for label, text in sales_kit["category_explainer"].items():
            st.markdown(f"**{humanize_label(label)}**")
            st.write(text)

    with st.expander("Cold outreach sequence", expanded=True):
        for label, text in sales_kit["cold_outreach_sequence"].items():
            st.markdown(f"**{humanize_label(label)}**")
            st.code(text, language="text")

    with st.expander("Persona-specific scripts", expanded=False):
        for label, text in sales_kit["persona_scripts"].items():
            st.markdown(f"**{humanize_label(label)}**")
            st.code(text, language="text")

    with st.expander("Objection handling", expanded=False):
        for question, answer in sales_kit["objection_responses"].items():
            st.markdown(f"**{question}**")
            st.write(answer)

    with st.expander("Data request sequence", expanded=False):
        for label, text in sales_kit["data_request_sequence"].items():
            st.markdown(f"**{humanize_label(label)}**")
            st.code(text, language="text")

    with st.expander("Moat positioning", expanded=False):
        for item in sales_kit["moat_positioning"]:
            st.write(f"• {item}")
        st.markdown("**Post-audit conversion**")
        st.code(sales_kit["post_audit_conversion"], language="text")

    st.download_button(
        "Download outreach pack",
        data=build_outreach_pack_text(sales_kit),
        file_name="courier_leak_audit_outreach_pack.txt",
        mime="text/plain",
        use_container_width=True,
    )

    st.subheader("Discovery call script")
    for question in sales_kit["discovery_script"]:
        st.write(f"• {question}")

    st.subheader("Client intake checklist")
    intake_items = [
        "Courier export for the audit period",
        "Courier invoice or statement for the same period",
        "Signed/current courier rate card",
        "Fuel, VAT, remote/outlying and surcharge rules",
        "Volumetric divisor and weight rounding rules",
        "Packaging/dimensions file if available",
        "Any courier credits or disputes already submitted",
    ]
    st.code("\n".join(f"[ ] {item}" for item in intake_items), language="text")
    st.download_button(
        "Download client intake checklist",
        data="Courier Leak Audit client intake checklist\n\n" + "\n".join(f"[ ] {item}" for item in intake_items),
        file_name="courier_leak_audit_client_intake_checklist.txt",
        mime="text/plain",
        use_container_width=True,
    )

    st.subheader("One-page offer sheet")
    sheet = sales_kit["offer_sheet"]
    st.markdown(f"""
    <div class="export-card">
    <b>{sheet['package']}</b><br>
    Audit fee: {sheet['audit_fee']}<br>
    Monitoring: {sheet['monitoring']}<br>
    Timeline: {sheet['timeline']}<br>
    {sheet['outcome_line']}
    </div>
    """, unsafe_allow_html=True)
    render_prepare_download(
        "Courier Leak Audit Offer Sheet",
        lambda: generate_offer_sheet_pdf(selected_client_name, sales_kit, currency_prefix),
        "courier_leak_audit_offer_sheet.pdf",
        "application/pdf",
        "export_offer_sheet",
        "Prepares a one-page marketable offer sheet for prospects or referral partners.",
    )

with pipeline_tab:
    render_step_header("9", "Client Pipeline", "Manage the full loop: prospect, request data, audit, send memo, submit dispute pack, track outcomes, and convert to monthly monitoring.")
    pipeline_offer = intro_offer if consultant_password == "pixie" else build_intro_stage_offer(
        safe_float(analysis_data.get("Billed_Cost_ZAR", pd.Series(dtype=float)).sum()),
        pillar_a_loss,
        operational_exposure_loss,
        currency_prefix,
    )
    pipeline_outcome_summary = summarize_outcomes(st.session_state.get("outcome_tracker_editor", pd.DataFrame()), currency_prefix)
    pipeline_seed = build_client_pipeline_template(
        selected_client_name,
        courier_provider,
        safe_float(analysis_data.get("Billed_Cost_ZAR", pd.Series(dtype=float)).sum()),
        pillar_a_loss,
        pipeline_outcome_summary.get("credited_amount", 0),
        pipeline_offer,
        currency_prefix,
    )
    if "client_pipeline_df" not in st.session_state:
        st.session_state["client_pipeline_df"] = pipeline_seed
    if st.button("Add current audit to pipeline", use_container_width=True):
        st.session_state["client_pipeline_df"] = pd.concat([st.session_state["client_pipeline_df"], pipeline_seed], ignore_index=True)
        st.success("Current audit added to the pipeline.")

    pipeline_df = st.data_editor(
        st.session_state["client_pipeline_df"],
        use_container_width=True,
        hide_index=True,
        num_rows="dynamic",
        column_config={
            "Stage": st.column_config.SelectboxColumn("Stage", options=PIPELINE_STAGES),
            "Monthly_Spend_ZAR": st.column_config.NumberColumn(f"Monthly Spend ({currency_symbol})", min_value=0.0),
            "Direct_Candidates_ZAR": st.column_config.NumberColumn(f"Direct Candidates ({currency_symbol})", min_value=0.0),
            "Credited_ZAR": st.column_config.NumberColumn(f"Credited ({currency_symbol})", min_value=0.0),
            "Audit_Fee_ZAR": st.column_config.NumberColumn(f"Audit Fee ({currency_symbol})", min_value=0.0),
            "Monthly_Retainer_ZAR": st.column_config.NumberColumn(f"Monthly Retainer ({currency_symbol})", min_value=0.0),
        },
        key="client_pipeline_editor",
    )
    st.session_state["client_pipeline_df"] = pipeline_df
    pipeline_summary = summarize_pipeline(pipeline_df)
    p1, p2, p3, p4, p5 = st.columns(5)
    p1.metric("Clients", f"{pipeline_summary['client_count']:,}")
    p2.metric("Active", f"{pipeline_summary['active_count']:,}")
    p3.metric("Converted", f"{pipeline_summary['converted_count']:,}")
    p4.metric("Projected setup", format_currency(pipeline_summary["projected_setup"], currency_prefix))
    p5.metric("Projected MRR", format_currency(pipeline_summary["projected_mrr"], currency_prefix))
    st.metric("Credited outcomes in pipeline", format_currency(pipeline_summary["credited_total"], currency_prefix))

    st.subheader("Next action guidance")
    for _, row in limited_preview(pipeline_df, 10).iterrows():
        st.markdown(
            f"""
            <div class="talk-track-card">
            <b>{safe_text(row.get('Client', 'Client'), 'Client')} — {safe_text(row.get('Stage', ''), '')}</b><br>
            {recommend_pipeline_action(row.get('Stage', ''))}
            </div>
            """,
            unsafe_allow_html=True,
        )
    render_prepare_download(
        "Client Pipeline Excel",
        lambda: convert_df_to_styled_excel(pipeline_df),
        "client_pipeline_tracker.xlsx",
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "export_client_pipeline",
        "Exports the editable client pipeline for sales, delivery, follow-up, and retainer tracking.",
    )

with readiness_tab:
    render_step_header("10", "Evidence & Business SWOT", "Score how safe this audit is to present, then address the business strengths, weaknesses, opportunities, and threats.")
    current_outcome_summary = summarize_outcomes(st.session_state.get("outcome_tracker_editor", pd.DataFrame()), currency_prefix)
    current_readiness = calculate_evidence_readiness(analysis_data, findings_df, skipped_rows, cleaned_rate_card, current_outcome_summary, rate_card_matched_count)
    r1, r2, r3 = st.columns(3)
    r1.metric("Evidence score", f"{current_readiness['score']}/100")
    r2.metric("Readiness band", current_readiness["band"])
    r3.metric("Case-study ready", "Yes" if current_readiness["case_study_ready"] else "No")
    st.markdown(f"""
    <div class="talk-track-card"><b>Posture</b><br>{current_readiness['posture']}</div>
    <div class="action-card"><b>Recommended next action</b><br>{current_readiness['recommended_next_action']}</div>
    """, unsafe_allow_html=True)

    st.subheader("Missing evidence checklist")
    if current_readiness["missing_evidence"]:
        for item in current_readiness["missing_evidence"]:
            st.write(f"• {item}")
    else:
        st.success("No major evidence gaps detected. Keep tracking outcomes and credit notes.")

    request_pack = build_rate_card_request_pack(current_readiness)
    st.subheader("Rate card / evidence request copy")
    st.code(request_pack["email"], language="text")
    st.code(request_pack["whatsapp"], language="text")

    swot_offer = intro_offer if consultant_password == "pixie" else build_intro_stage_offer(
        safe_float(analysis_data.get("Billed_Cost_ZAR", pd.Series(dtype=float)).sum()),
        pillar_a_loss,
        operational_exposure_loss,
        currency_prefix,
    )
    swot_sales_kit = build_sales_kit(selected_client_name, courier_provider, pillar_a_loss, operational_exposure_loss, pillar_b_loss, financial_anomaly_loss, currency_prefix, swot_offer, current_outcome_summary)
    swot_pipeline_summary = summarize_pipeline(st.session_state.get("client_pipeline_df", pd.DataFrame()))
    business_swot = build_business_swot(current_readiness, current_outcome_summary, swot_pipeline_summary, swot_sales_kit, swot_offer)
    st.subheader("Business SWOT")
    swot_cols = st.columns(4)
    for col, section in zip(swot_cols, ["Strengths", "Weaknesses", "Opportunities", "Threats"]):
        with col:
            st.markdown(f"**{section}**")
            for item in business_swot[section]:
                st.write(f"• {item}")
    st.subheader("Address this next")
    for item in business_swot["Address Next"]:
        st.markdown(f"<div class='action-card'>{item}</div>", unsafe_allow_html=True)

with monitoring_tab:
    render_step_header("11", "Monthly Monitoring", "Turn a once-off audit into a recurring courier margin control rhythm: monitor rate-card changes, repeated issues, outcomes, and next-month actions.")
    monitoring_outcome_summary = summarize_outcomes(st.session_state.get("outcome_tracker_editor", pd.DataFrame()), currency_prefix)
    monitoring_summary = build_monthly_monitoring_summary(
        analysis_data,
        findings_df,
        leakage_totals,
        monitoring_outcome_summary,
        evidence_readiness,
        rate_card_matched_count,
    )
    repeated_patterns = detect_repeated_issue_patterns(findings_df, packaging_rows, financial_anomaly_rows)
    current_snapshot = build_monthly_snapshot(
        selected_client_name,
        audit_month,
        audit_year,
        courier_provider,
        monitoring_summary,
        leakage_totals,
        repeated_patterns,
    )
    snapshot_history = load_monthly_snapshots()
    snapshot_comparison = compare_monthly_snapshots(current_snapshot, snapshot_history)
    trend_narrative = build_monthly_trend_narrative(snapshot_comparison, currency_prefix)
    retainer_pitch = build_monitoring_retainer_pitch(monitoring_summary, repeated_patterns, currency_prefix)
    next_month_checklist = build_next_month_checklist(cleaned_rate_card)
    monitoring_client_message = (
        f"Hi {{Name}},\n\nThis month's courier audit is now the baseline for monthly monitoring. "
        f"We found {format_currency(monitoring_summary['direct_recovery'] + monitoring_summary['account_manager_query'], currency_prefix)} in recovery/query candidates and "
        f"{format_currency(monitoring_summary['operational_exposure'], currency_prefix)} in operational exposure.\n\n"
        "The reason I recommend monitoring this monthly is that courier rate cards and surcharge rules do not stand still. Fuel prices, inflation, exchange rates, remote-zone rules, service costs and courier network changes can all shift the rate-card context. "
        "By checking the latest export, invoice and rate card each month, we keep the business aligned to current courier rules instead of paying against outdated assumptions.\n\n"
        "Next step: send the latest monthly export, invoice/statement, and any updated rate-card or surcharge schedule so we can compare what changed and track repeat issues.\n\nKyle"
    )

    m1, m2, m3, m4, m5, m6 = st.columns(6)
    m1.metric("Courier spend", format_currency(monitoring_summary["courier_spend"], currency_prefix))
    m2.metric("Recovery/query", format_currency(monitoring_summary["direct_recovery"] + monitoring_summary["account_manager_query"], currency_prefix))
    m3.metric("Operational exposure", format_currency(monitoring_summary["operational_exposure"], currency_prefix))
    m4.metric("Credited", format_currency(monitoring_summary["credited_outcomes"], currency_prefix))
    m5.metric("Rate-card matched", f"{monitoring_summary['rate_card_matched_rows']:,}")
    m6.metric("Evidence score", f"{monitoring_summary['evidence_score']}/100")

    st.subheader("Snapshot History & Trends")
    if st.button("Save current month snapshot", use_container_width=True):
        snapshot_history = save_monthly_snapshot(current_snapshot)
        snapshot_comparison = compare_monthly_snapshots(current_snapshot, snapshot_history)
        trend_narrative = build_monthly_trend_narrative(snapshot_comparison, currency_prefix)
        st.success(f"Saved monthly snapshot to {MONTHLY_SNAPSHOT_PATH.name}.")
    t1, t2, t3, t4 = st.columns(4)
    t1.metric("Spend change", format_currency(snapshot_comparison.get("spend_delta", 0), currency_prefix) if snapshot_comparison.get("has_previous") else "Baseline")
    t2.metric("Recovery/query change", format_currency(snapshot_comparison.get("recovery_query_delta", 0), currency_prefix) if snapshot_comparison.get("has_previous") else "Baseline")
    t3.metric("Operational change", format_currency(snapshot_comparison.get("operational_delta", 0), currency_prefix) if snapshot_comparison.get("has_previous") else "Baseline")
    t4.metric("Evidence score change", f"{snapshot_comparison.get('evidence_score_delta', 0):.0f}" if snapshot_comparison.get("has_previous") else "Baseline")
    st.info(trend_narrative)
    client_history = snapshot_history[
        snapshot_history["Client"].astype(str).eq(str(current_snapshot["Client"]))
        & snapshot_history["Courier"].astype(str).eq(str(current_snapshot["Courier"]))
    ] if not snapshot_history.empty else pd.DataFrame()
    if client_history.empty:
        st.caption("No saved history for this client/courier yet. Save this month to create the baseline.")
    else:
        st.dataframe(display_currency_columns(client_history.sort_values(["Audit_Year", "Audit_Month"], ascending=False), currency_symbol), use_container_width=True, hide_index=True)
        st.download_button(
            "Download monthly snapshot history CSV",
            data=snapshot_history.to_csv(index=False),
            file_name="monthly_audit_snapshots.csv",
            mime="text/csv",
            use_container_width=True,
        )

    st.subheader("Current-month repeat signals")
    if repeated_patterns.empty:
        st.info("No repeated issue signals detected yet. Treat this month as the monitoring baseline.")
    else:
        st.dataframe(display_currency_columns(repeated_patterns, currency_symbol), use_container_width=True, hide_index=True)

    st.subheader("Why monthly monitoring is justified")
    st.markdown(f"<div class='talk-track-card'>{retainer_pitch}</div>", unsafe_allow_html=True)

    st.subheader("Next month checklist")
    st.code("\n".join(f"[ ] {item}" for item in next_month_checklist), language="text")

    st.subheader("Client monitoring message")
    st.code(monitoring_client_message, language="text")

    st.download_button(
        "Download monthly monitoring pack",
        data=build_monthly_monitoring_pack_text(monitoring_summary, repeated_patterns, retainer_pitch, next_month_checklist, monitoring_client_message, currency_prefix),
        file_name="monthly_monitoring_pack.txt",
        mime="text/plain",
        use_container_width=True,
    )
